import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import boto3
from botocore.exceptions import ClientError
import logging
import pytz
import finnhub
import time
import threading
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from requests.exceptions import HTTPError
from dotenv import load_dotenv
from polygon import RESTClient
from polygon.rest.models import TickerNews
import requests
from bs4 import BeautifulSoup
# Removed requests-html import - using traditional requests + BeautifulSoup instead
from bs4 import XMLParsedAsHTMLWarning
import warnings
import yfinance as yf
import re
import sys
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import holidays
from typing import Any, Union, Optional, List, Dict, Tuple
import gc
import psutil
from openpyxl.styles import PatternFill
# Removed nest_asyncio import - no longer needed
from playwright.sync_api import sync_playwright
import ta  # For RSI and Bollinger Bands


# Configure logging (file only)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s:%(levelname)s:%(message)s",
    handlers=[logging.FileHandler("stock_data.log")]
)

# Suppress XMLParsedAsHTMLWarning
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

# Lightweight in-memory cache for GET requests
_original_requests_get = requests.get

@lru_cache(maxsize=2048)
def _cached_get(url, timeout=10):
    resp = _original_requests_get(url, timeout=timeout)
    resp.raise_for_status()
    return resp

# Don't override requests.get globally - it causes issues with headers
# requests.get = _cached_get

# Rate limiter for API calls
class RateLimiter:
    def __init__(self, max_calls: int, period: float):
        self.max_calls = max_calls
        self.period = period
        self.lock = threading.Lock()
        self.calls = []

    def acquire(self):
        with self.lock:
            now = time.time()
            while self.calls and self.calls[0] <= now - self.period:
                self.calls.pop(0)
            if len(self.calls) >= self.max_calls:
                time.sleep(self.period - (now - self.calls[0]))
            self.calls.append(now)
    
    def acquire_with_lock(self):
        """Thread-safe acquire with additional locking for high-concurrency scenarios"""
        # FIXED: Remove double-locking issue
        return self.acquire()

# Instantiate global rate limiters
POLYGON_PREMIUM = os.getenv("POLYGON_PREMIUM", "false").lower() == "true"
# Optimized rate limiters for parallel processing
polygon_limiter = RateLimiter(max_calls=200 if POLYGON_PREMIUM else 30, period=1.0 if POLYGON_PREMIUM else 60.0)  # Increased for parallel processing
finnhub_limiter = RateLimiter(max_calls=80, period=60.0)  # Increased for parallel processing
logging.info(f"Polygon rate limit: {200 if POLYGON_PREMIUM else 30} calls per {'second' if POLYGON_PREMIUM else 'minute'}")

# Load environment variables
load_dotenv()

# API keys
POLYGON_API_KEY = os.getenv("POLYGON_API_KEY")
FINNHUB_API_KEY = os.getenv("FINNHUB_API_KEY")
SEC_API_KEY = os.getenv("SEC_API_KEY")

# Initialize SEC API
if SEC_API_KEY:
    logging.info("SEC API key loaded successfully")
    SEC_WORKING = True
else:
    logging.warning("SEC API key not found in environment variables")
    SEC_WORKING = False

# Float Rotation Functions
def get_intraday_cumulative_volume(ticker, date, api_key):
    """
    Get intraday cumulative volume data from Polygon API for a specific date
    Returns cumulative volume at specific time points: 5AM, 6AM, 7AM, 8AM, 9AM, 9:30AM, 10AM, 11AM, 12PM, 1PM, 4PM
    Each time point represents total volume from market open up to that time
    """
    # Normalize date handling - accept both string and datetime
    if isinstance(date, str):
        date = datetime.strptime(date, '%Y-%m-%d')
    date_str = date.strftime('%Y-%m-%d')
    
    # Polygon API endpoint for intraday data
    url = f"https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/minute/{date_str}/{date_str}?adjusted=true&sort=asc&limit=50000&apiKey={api_key}"
    
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        # Check if response has expected structure
        if 'status' not in data or data['status'] != 'OK' or 'results' not in data or not data['results']:
            logging.warning(f"No intraday data found for {ticker} on {date_str} (status: {data.get('status', 'N/A')}, has_results: {'results' in data})")
            return None
        
        # Convert to DataFrame
        df = pd.DataFrame(data['results'])
        # Convert UTC timestamps to Eastern Time (automatically handles EST/EDT)
        df['timestamp'] = pd.to_datetime(df['t'], unit='ms', utc=True)
        df['timestamp'] = df['timestamp'].dt.tz_convert(EASTERN_TZ)
        df['time'] = df['timestamp'].dt.time
        
        # Define time points for volume aggregation
        time_points = {
            '5AM': '05:00:00',
            '6AM': '06:00:00', 
            '7AM': '07:00:00',
            '8AM': '08:00:00',
            '9AM': '09:00:00',
            '9:30AM': '09:30:00',
            '10AM': '10:00:00',
            '11AM': '11:00:00',
            '12PM': '12:00:00',
            '1PM': '13:00:00',
            '4PM': '16:00:00'  # Market close
        }
        
        # Calculate cumulative volume at each time point
        volume_data = {}
        
        for label, time_str in time_points.items():
            target_time = datetime.strptime(time_str, '%H:%M:%S').time()
            
            # Filter data up to this time point
            mask = df['time'] <= target_time
            cumulative_volume = df[mask]['v'].sum()
            
            volume_data[label] = cumulative_volume
        
        return volume_data
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching intraday data for {ticker}: {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error for {ticker}: {e}")
        return None

def calculate_float_rotation(cumulative_volume_data, float_shares):
    """
    Calculate float rotation for each time point
    Float Rotation = Cumulative Volume / Float
    """
    if not cumulative_volume_data or float_shares is None or float_shares <= 0:
        return None
    
    float_rotation = {}
    for time_point, cumulative_volume in cumulative_volume_data.items():
        if cumulative_volume > 0:
            float_rotation[time_point] = cumulative_volume / float_shares
        else:
            float_rotation[time_point] = 0
    
    return float_rotation

def calculate_rvol(cumulative_volume_data, ticker, date, polygon_api_key, pm_high_time=None, open_high_time=None):
    """
    Calculate Relative Volume (RVOL) for each time point with exact time calculations
    RVOL = Cumulative Volume up to exact time point / Average Cumulative Volume up to exact time point (last 10 trading days)
    """
    if not cumulative_volume_data:
        return None
    
    try:
        # Define standard time points in order for cumulative calculation
        time_points_order = ['5AM', '6AM', '7AM', '8AM', '9AM', '9:30AM', '10AM', '11AM', '12PM', '1PM', '4PM']
        
        # Add special time points if provided
        special_time_points = {}
        if pm_high_time and pm_high_time != "Not Available":
            special_time_points['PM_High'] = pm_high_time
        if open_high_time and open_high_time != "Not Available":
            special_time_points['Open_High'] = open_high_time
        
        # Combine all time points
        all_time_points = time_points_order + list(special_time_points.keys())
        
        # Get current day's exact cumulative volumes
        current_cumulative_volumes = {}
        
        # For standard time points, use existing cumulative_volume_data
        for time_point in time_points_order:
            if time_point in cumulative_volume_data:
                current_cumulative_volumes[time_point] = cumulative_volume_data[time_point]
        
        # For special time points, calculate exact cumulative volume
        if special_time_points:
            # Get raw intraday data for exact calculations
            raw_data = get_raw_intraday_data(ticker, date, polygon_api_key)
            if raw_data is not None and not raw_data.empty:
                for time_point, time_str in special_time_points.items():
                    exact_volume = calculate_exact_cumulative_volume(raw_data, time_str)
                    current_cumulative_volumes[time_point] = exact_volume
            else:
                # Fallback: Use closest available time point for special times
                logging.debug(f"No raw intraday data available for {ticker} on {date}, using fallback for special times")
                for time_point, time_str in special_time_points.items():
                    # Try to find closest time point in cumulative_volume_data
                    if time_point == 'PM_High' and '8AM' in cumulative_volume_data:
                        current_cumulative_volumes[time_point] = cumulative_volume_data['8AM']
                    elif time_point == 'Open_High' and '9:30AM' in cumulative_volume_data:
                        current_cumulative_volumes[time_point] = cumulative_volume_data['9:30AM']
                    else:
                        current_cumulative_volumes[time_point] = None
        
        # Get historical cumulative volume data for the last 10 trading days
        end_date = datetime.strptime(date, "%Y-%m-%d")
        historical_cumulative_volumes = {}
        
        # Initialize historical data structure
        for time_point in all_time_points:
            historical_cumulative_volumes[time_point] = []
        
        # Get trading days only - more efficient and accurate
        trading_days = get_trading_days(end_date, num_days=10)
        
        for check_date in trading_days:
            check_date_str = check_date.strftime("%Y-%m-%d")
            
            try:
                # Get volume data for historical date
                hist_volume_data = get_intraday_cumulative_volume(ticker, check_date, polygon_api_key)
                if hist_volume_data:
                    # Use cumulative volumes for standard time points
                    for time_point in time_points_order:
                        if time_point in hist_volume_data:
                            historical_cumulative_volumes[time_point].append(hist_volume_data[time_point])
                    
                    # Calculate exact cumulative volumes for special time points
                    if special_time_points:
                        hist_raw_data = get_raw_intraday_data(ticker, check_date_str, polygon_api_key)
                        if hist_raw_data is not None and not hist_raw_data.empty:
                            for time_point, time_str in special_time_points.items():
                                exact_volume = calculate_exact_cumulative_volume(hist_raw_data, time_str)
                                historical_cumulative_volumes[time_point].append(exact_volume)
                        else:
                            # Fallback: Use closest available time point for historical special times
                            for time_point, time_str in special_time_points.items():
                                if time_point == 'PM_High' and '8AM' in hist_volume_data:
                                    historical_cumulative_volumes[time_point].append(hist_volume_data['8AM'])
                                elif time_point == 'Open_High' and '9:30AM' in hist_volume_data:
                                    historical_cumulative_volumes[time_point].append(hist_volume_data['9:30AM'])
                    
                    logging.debug(f"Processed trading day for {ticker}: {check_date_str}")
                    
            except Exception as e:
                logging.debug(f"Error getting historical volume for {ticker} on {check_date_str}: {e}")
                continue
        
        # Calculate average cumulative volume for each time point
        avg_cumulative_volumes = {}
        for time_point in all_time_points:
            volumes = historical_cumulative_volumes[time_point]
            if volumes:
                avg_cumulative_volumes[time_point] = sum(volumes) / len(volumes)
            else:
                avg_cumulative_volumes[time_point] = None
        
        # Calculate RVOL for each time point using cumulative volumes
        rvol_data = {}
        for time_point in all_time_points:
            if time_point in current_cumulative_volumes and time_point in avg_cumulative_volumes:
                current_cumulative = current_cumulative_volumes[time_point]
                avg_cumulative = avg_cumulative_volumes[time_point]
                
                if avg_cumulative and avg_cumulative > 0:
                    rvol_data[time_point] = current_cumulative / avg_cumulative
                else:
                    rvol_data[time_point] = None
            else:
                rvol_data[time_point] = None
        
        return rvol_data, historical_cumulative_volumes
        
    except Exception as e:
        logging.error(f"Error calculating RVOL for {ticker}: {e}")
        return None, None

def get_trading_days(end_date, num_days=10):
    """
    Get the last N trading days (excluding weekends and holidays)
    """
    trading_days = []
    days_checked = 0
    max_days_to_check = num_days * 3  # Check up to 3x the number of days needed
    
    while len(trading_days) < num_days and days_checked < max_days_to_check:
        check_date = end_date - timedelta(days=days_checked + 1)
        
        # Skip weekends
        if check_date.weekday() >= 5:  # Saturday = 5, Sunday = 6
            days_checked += 1
            continue
        
        # Skip US holidays (except Veterans Day and Columbus Day - market is open)
        us_holidays = holidays.US()
        if check_date in us_holidays:
            holiday_name = us_holidays[check_date]
            # Veterans Day and Columbus Day are trading days - market is open
            if ("Veterans Day" not in holiday_name and "Veterans" not in holiday_name and 
                "Columbus Day" not in holiday_name and "Columbus" not in holiday_name):
                days_checked += 1
                continue
        
        trading_days.append(check_date)
        days_checked += 1
    
    return trading_days

def get_raw_intraday_data(ticker, date, api_key):
    """
    Get raw intraday data from Polygon API for exact time calculations
    """
    try:
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
        date_str = date.strftime('%Y-%m-%d')
        
        url = f"https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/minute/{date_str}/{date_str}?adjusted=true&sort=asc&limit=50000&apiKey={api_key}"
        
        response = requests.get(url, timeout=5)  # Reduced timeout to prevent hanging
        response.raise_for_status()
        data = response.json()
        
        # Check if we have valid data
        if 'status' not in data or data['status'] != 'OK':
            logging.debug(f"Invalid status for {ticker} on {date_str}: {data.get('status', 'UNKNOWN')}")
            return None
            
        if 'results' not in data or not data['results']:
            logging.debug(f"No results data for {ticker} on {date_str}")
            return None
        
        # Convert to DataFrame with proper timestamp handling
        df = pd.DataFrame(data['results'])
        if df.empty:
            logging.debug(f"Empty DataFrame for {ticker} on {date_str}")
            return None
            
        # Convert UTC timestamps to Eastern Time (automatically handles EST/EDT)
        df['timestamp'] = pd.to_datetime(df['t'], unit='ms', utc=True)
        df['timestamp'] = df['timestamp'].dt.tz_convert(EASTERN_TZ)
        df['time'] = df['timestamp'].dt.time
        
        return df
        
    except requests.exceptions.Timeout:
        logging.debug(f"Timeout getting raw intraday data for {ticker} on {date_str}")
        return None
    except requests.exceptions.RequestException as e:
        logging.debug(f"Request error getting raw intraday data for {ticker}: {e}")
        return None
    except Exception as e:
        logging.debug(f"Error getting raw intraday data for {ticker}: {e}")
        return None

def calculate_exact_cumulative_volume(raw_data, time_str):
    """
    Calculate exact cumulative volume up to a specific time
    """
    try:
        if raw_data is None or raw_data.empty:
            return 0
        
        # Parse the time string (e.g., "08:45:00" or "8:45:00")
        if ':' in time_str:
            # Handle time format like "08:45:00"
            target_time = datetime.strptime(time_str, '%H:%M:%S').time()
        else:
            # Handle time format like "8:45:00"
            target_time = datetime.strptime(time_str, '%H:%M:%S').time()
        
        # Filter data up to this exact time point
        mask = raw_data['time'] <= target_time
        cumulative_volume = raw_data[mask]['v'].sum()
        
        return cumulative_volume
        
    except Exception as e:
        logging.debug(f"Error calculating exact cumulative volume for {time_str}: {e}")
        return 0

def calculate_hourly_volumes(raw_data):
    """
    Calculate volume for each hour from 6AM to 12PM
    Returns a dictionary with hourly volume data
    """
    try:
        if raw_data is None or raw_data.empty:
            return {}
        
        hourly_volumes = {}
        
        # Define hourly ranges
        hourly_ranges = {
            '6AM': ('06:00:00', '06:59:59'),
            '7AM': ('07:00:00', '07:59:59'),
            '8AM': ('08:00:00', '08:59:59'),
            '9AM': ('09:00:00', '09:59:59'),
            '10AM': ('10:00:00', '10:59:59'),
            '11AM': ('11:00:00', '11:59:59'),
            '12PM': ('12:00:00', '12:59:59')
        }
        
        for hour_label, (start_time, end_time) in hourly_ranges.items():
            try:
                start_dt = datetime.strptime(start_time, '%H:%M:%S').time()
                end_dt = datetime.strptime(end_time, '%H:%M:%S').time()
                
                # Filter data for this hour
                mask = (raw_data['time'] >= start_dt) & (raw_data['time'] <= end_dt)
                hour_volume = raw_data[mask]['v'].sum()
                
                hourly_volumes[hour_label] = hour_volume if not pd.isna(hour_volume) else 0
                
            except Exception as e:
                logging.debug(f"Error calculating volume for {hour_label}: {e}")
                hourly_volumes[hour_label] = 0
        
        return hourly_volumes
        
    except Exception as e:
        logging.error(f"Error calculating hourly volumes: {e}")
        return {}

def calculate_macd(prices, fast_period=12, slow_period=26, signal_period=9):
    """
    Calculate MACD (Moving Average Convergence Divergence)
    
    Args:
        prices (list): List of price data
        fast_period (int): Fast EMA period (default 12)
        slow_period (int): Slow EMA period (default 26)
        signal_period (int): Signal line period (default 9)
    
    Returns:
        dict: MACD values with full arrays (macd_line, signal_line, histogram)
    """
    if len(prices) < slow_period:
        return None
    
    try:
        import numpy as np
        
        prices = np.array(prices)
        
        # Calculate EMAs
        ema_fast = calculate_ema(prices, fast_period)
        ema_slow = calculate_ema(prices, slow_period)
        
        # MACD line
        macd_line = ema_fast - ema_slow
        
        # Signal line (EMA of MACD line)
        signal_line = calculate_ema(macd_line, signal_period)
        
        # Histogram
        histogram = macd_line - signal_line
        
        # Return last values from arrays to match expected scalar format
        return {
            'macd_line': macd_line[-1] if len(macd_line) > 0 else None,
            'signal_line': signal_line[-1] if len(signal_line) > 0 else None,
            'histogram': histogram[-1] if len(histogram) > 0 else None
        }
        
    except Exception as e:
        logging.error(f"Error calculating MACD: {e}")
        return None

def calculate_ema(data, period):
    """
    Calculate Exponential Moving Average
    
    Args:
        data (numpy array): Price data
        period (int): EMA period
    
    Returns:
        numpy array: EMA values
    """
    import numpy as np
    
    alpha = 2 / (period + 1)
    ema = np.zeros_like(data)
    ema[0] = data[0]
    
    for i in range(1, len(data)):
        ema[i] = alpha * data[i] + (1 - alpha) * ema[i-1]
    
    return ema

def calculate_macd_at_high(data, high_time, fast_period=12, slow_period=26, signal_period=9, hist_data=None):
    """
    Calculate MACD at a specific high time using rolling period like brokers
    
    Args:
        data (DataFrame): Current day's price data
        high_time (datetime): Time of the high
        fast_period (int): Fast EMA period (default 12)
        slow_period (int): Slow EMA period (default 26)
        signal_period (int): Signal line period (default 9)
        hist_data (DataFrame): Historical data for rolling calculation
    
    Returns:
        dict: MACD values at the high time
    """
    try:
        # Use hist_data for rolling calculation if available, but ensure final value is from target date
        if hist_data is not None and not hist_data.empty:
            # Calculate MACD using historical data for rolling window (30 days)
            min_required = slow_period + signal_period
            if len(hist_data) < min_required:
                logging.debug(f"MACD calculation failed: insufficient historical data. Available={len(hist_data)}, Required={min_required}")
                return None
            
            # Calculate MACD using ta library (already imported) or manual calculation
            try:
                from ta.trend import MACD
                # Use ta library for MACD calculation with historical data
                macd_indicator = MACD(close=hist_data['close'], 
                                     window_fast=fast_period, 
                                     window_slow=slow_period, 
                                     window_sign=signal_period)
                macd_line = macd_indicator.macd().values
                signal_line = macd_indicator.macd_signal().values
                histogram = macd_indicator.macd_diff().values
            except Exception as e:
                # Fallback to manual calculation if ta library fails
                logging.debug(f"ta library MACD failed: {e}, using manual calculation")
                macd_result = calculate_macd(hist_data['close'].tolist(), fast_period, slow_period, signal_period)
                if macd_result:
                    return macd_result
                else:
                    return None
            
            # Find the MACD values at the specific high_time from target date data
            if high_time in data.index:
                # Use target date data for the actual time point
                if len(data) < min_required:
                    logging.debug(f"MACD target data insufficient: {len(data)}, Required={min_required}")
                    return None
                
                try:
                    from ta.trend import MACD
                    # Calculate MACD for target date data
                    macd_target_indicator = MACD(close=data['close'], 
                                               window_fast=fast_period, 
                                               window_slow=slow_period, 
                                               window_sign=signal_period)
                    macd_target_line = macd_target_indicator.macd().values
                    signal_target_line = macd_target_indicator.macd_signal().values
                    histogram_target = macd_target_indicator.macd_diff().values
                except Exception as e:
                    logging.debug(f"ta library MACD target failed: {e}")
                    return None
                
                if high_time in data.index:
                    idx = data.index.get_loc(high_time)
                else:
                    # Find closest time in target date data
                    time_diff = abs(data.index - high_time)
                    idx = time_diff.argmin()
                
                # Get MACD values at the index from target date data
                macd_value = macd_target_line[idx] if idx < len(macd_target_line) else None
                signal_value = signal_target_line[idx] if idx < len(signal_target_line) else None
                hist_value = histogram_target[idx] if idx < len(histogram_target) else None
            else:
                # High time not in target date data - return None
                logging.debug(f"High time {high_time} not found in target date data")
                return None
        else:
            # No historical data available - use target date data only
            use_data = data
            if use_data.empty or pd.isna(high_time):
                logging.debug(f"MACD calculation failed: use_data empty={use_data.empty}, high_time={high_time}")
                return None
            
            # Ensure we have enough data for MACD calculation (need at least slow_period + signal_period)
            min_required = slow_period + signal_period
            if len(use_data) < min_required:
                logging.debug(f"MACD calculation failed: insufficient data. Available={len(use_data)}, Required={min_required}")
                return None
            
            # Calculate MACD using ta library (already imported) or manual calculation
            try:
                from ta.trend import MACD
                # Use ta library for MACD calculation
                macd_indicator = MACD(close=use_data['close'], 
                                     window_fast=fast_period, 
                                     window_slow=slow_period, 
                                     window_sign=signal_period)
                macd_line = macd_indicator.macd().values
                signal_line = macd_indicator.macd_signal().values
                histogram = macd_indicator.macd_diff().values
            except Exception as e:
                # Fallback to manual calculation if ta library fails
                logging.debug(f"ta library MACD failed: {e}, using manual calculation")
                macd_result = calculate_macd(use_data['close'].tolist(), fast_period, slow_period, signal_period)
                if macd_result:
                    return macd_result
                else:
                    return None
            
            # Find the MACD values at or closest to the high time
            if high_time in use_data.index:
                idx = use_data.index.get_loc(high_time)
            else:
                # Find the closest time within a reasonable range
                time_diff = abs(use_data.index - high_time)
                idx = time_diff.argmin()
            
            # Get MACD values at the index
            macd_value = macd_line[idx] if idx < len(macd_line) else None
            signal_value = signal_line[idx] if idx < len(signal_line) else None
            hist_value = histogram[idx] if idx < len(histogram) else None
        
        # Check for NaN values
        if pd.isna(macd_value) or pd.isna(signal_value) or pd.isna(hist_value):
            logging.debug(f"MACD calculation failed: NaN values at index {idx}")
            return None
        
        return {
            'macd_line': round(macd_value, 4),
            'signal_line': round(signal_value, 4),
            'histogram': round(hist_value, 4)
        }
        
    except Exception as e:
        logging.warning(f"Error calculating MACD at high: {e}")
        return None

# Configure global connection pool settings for urllib3 - optimized for parallel processing
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Optimize connection pooling for parallel requests
urllib3.util.retry.Retry.DEFAULT_ALLOWED_METHODS = frozenset(['GET', 'POST', 'PUT', 'DELETE', 'HEAD', 'OPTIONS', 'TRACE'])

# Configure session pooling for better parallel performance
session_pool = {
    'max_retries': 3,
    'backoff_factor': 0.3,
    'status_forcelist': [500, 502, 503, 504],
    'pool_connections': 100,  # Increased for parallel processing
    'pool_maxsize': 100,      # Increased for parallel processing
}

# Web Scraping Fallback for Historical Float Data
class WaybackFloatScraper:
    """Web scraping fallback for historical float data from Yahoo Finance statistics tab."""
    
    def __init__(self):
        self.session = requests.Session()
        
        # Configure session with explicit retry strategy to completely disable retries
        retry_strategy = Retry(
            total=0,  # No retries at all
            connect=0,  # No connection retries
            read=0,     # No read retries
            redirect=0, # No redirect retries
            status=0,   # No status retries
            other=0     # No other retries
        )
        
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=100,  # Increased to match global session for 20 workers
            pool_maxsize=100,      # Increased to match global session for 20 workers
            max_retries=retry_strategy,  # Use explicit retry strategy
            pool_block=False
        )
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
    
    def get_float_from_wayback(self, ticker, target_date):
        """
        Try to get float data from Wayback Machine for historical data.
        
        Args:
            ticker (str): Stock ticker symbol
            target_date (str): Target date in YYYY-MM-DD format
            
        Returns:
            dict: Float data or None
        """
        # Commented out - using Excel float data only
        # try:
        #     timeout_count = 0  # Track consecutive timeouts
        #     # Skip after first timeout to fall back to existing float fallback process
        #     max_consecutive_timeouts = 1  # Skip ticker after 1 timeout
        #     # Try different URL formats and dates around the target (expanded for better coverage)
        #     base_urls = [
        #         f"https://finance.yahoo.com/quote/{ticker}/key-statistics/",
        #         f"https://au.finance.yahoo.com/quote/{ticker}/key-statistics/",
        #         f"https://finance.yahoo.com/quote/{ticker}/statistics/",
        #         f"https://au.finance.yahoo.com/quote/{ticker}/statistics/",
        #         f"https://uk.finance.yahoo.com/quote/{ticker}/key-statistics/",
        #         f"https://uk.finance.yahoo.com/quote/{ticker}/statistics/",
        #         f"https://ca.finance.yahoo.com/quote/{ticker}/key-statistics/",
        #         f"https://ca.finance.yahoo.com/quote/{ticker}/statistics/"
        #     ]
            
        #     # Try dates up to 60 days before the target date (most current data as of that date)
        #     target_dt = datetime.strptime(target_date, "%Y-%m-%d")
        #     dates_to_try = []
        #     
        #     # Start from 60 days before and work forward to the target date
        #     for i in range(60, -1, -1):  # 60 days before to 0 days (target date)
        #         test_date = target_dt - timedelta(days=i)
        #         dates_to_try.append(test_date.strftime("%Y%m%d"))
        #     
        #     for base_url in base_urls:
        #         for date_str in dates_to_try:
        #             # Construct direct Wayback Machine URL
        #             wayback_url = f"https://web.archive.org/web/{date_str}000000/{base_url}"
        #             
        #             # Single attempt - no retries for timeouts to speed up processing
        #             try:
        #                 response = self.session.get(wayback_url, timeout=20)  # 20 second timeout
        #                 
        #                 # Reset timeout counter on successful response
        #                 timeout_count = 0
        #                 
        #                 if response.status_code == 200:
        #                     soup = BeautifulSoup(response.text, 'html.parser')
        #                     
        #                     # Check if this is a valid page (not a 404 or error)
        #                     if "not found" not in soup.get_text().lower() and "error" not in soup.get_text().lower():
        #                         float_data = self._extract_float_data(soup, ticker, date_str)
        #                         if float_data:
        #                             # Convert YYYYMMDD to YYYY-MM-DD format for display
        #                             display_date = f"{date_str[0:4]}-{date_str[4:6]}-{date_str[6:8]}"
        #                             float_data['date'] = display_date
        #                             logging.info(f"[WAYBACK] Found float data for {ticker} on {display_date}")
        #                             return float_data
        #                     
        #                 except requests.exceptions.ReadTimeout as e:
        #                     # Explicitly handle read timeouts - skip immediately
        #                     timeout_count += 1
        #                     logging.info(f"[WAYBACK] Read timeout for {ticker} on {date_str} - skipping to next URL (timeout #{timeout_count})")
        #                     if timeout_count >= max_consecutive_timeouts:
        #                         logging.info(f"[WAYBACK] First timeout for {ticker} - skipping to fallback process")
        #                         return None
        #                     continue
                    # Commented out - using Excel float data only
                    # except requests.exceptions.ConnectTimeout as e:
                    #     # Explicitly handle connection timeouts - skip immediately
                    #     timeout_count += 1
                    #     logging.info(f"[WAYBACK] Connection timeout for {ticker} on {date_str} - skipping to next URL (timeout #{timeout_count})")
                    #     if timeout_count >= max_consecutive_timeouts:
                    #         logging.info(f"[WAYBACK] First timeout for {ticker} - skipping to fallback process")
                    #         return None
                    #     continue
                    # except requests.exceptions.Timeout as e:
                    #     # Handle any other timeout - skip immediately
                    #     timeout_count += 1
                    #     logging.info(f"[WAYBACK] Timeout for {ticker} on {date_str} - skipping to next URL (timeout #{timeout_count})")
                    #     if timeout_count >= max_consecutive_timeouts:
                    #         logging.info(f"[WAYBACK] First timeout for {ticker} - skipping to fallback process")
                    #         return None
                    #     continue
                    # except Exception as e:
                    #     # Log other errors but continue to next URL
                    #     logging.warning(f"[WAYBACK] Error for {ticker} on {date_str}: {e}")
                    #     timeout_count = 0  # Reset timeout count for non-timeout errors
                    #     continue
                    
                    # Commented out - using Excel float data only
                    # Rate limiting - increased to 5s to be very conservative and respectful
                    # time.sleep(5)
            
            # Commented out - using Excel float data only
            # logging.warning(f"[WAYBACK] No float data found for {ticker} around {target_date}")
        return None
            
        # Commented out - using Excel float data only
        # except Exception as e:
        #     logging.error(f"[WAYBACK] Error in Wayback access for {ticker}: {e}")
        #     return None
        return None
    

    
    def _extract_float_data(self, soup, ticker, date_str):
        """
        Extract float data from Yahoo Finance statistics page with fixed regex patterns.
        
        Args:
            soup (BeautifulSoup): Parsed HTML
            ticker (str): Stock ticker
            date_str (str): Date string
            
        Returns:
            dict: Extracted float data
        """
        try:
            float_data = {
                'ticker': ticker,
                'date': date_str,
                'shares_outstanding': None,
                'float': None,
                'float_percentage': None,
                'institutional_ownership': None,
                'insider_ownership': None
            }
            
            # Get all text content with better formatting
            page_text = soup.get_text(separator=' ', strip=True)
            
            # Fixed patterns based on the debug output
            # The format is: "Shares Outstanding 5  508k" where "5" is column number and "508k" is the value
            patterns = {
                'shares_outstanding': [
                    # Handle column number format: "Shares Outstanding 5 55.64M" (MOST SPECIFIC FIRST)
                    r'Shares Outstanding\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Outstanding Shares\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Total Shares Outstanding\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Shares Outstanding\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    r'Outstanding Shares\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    r'Total Shares Outstanding\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    # Handle direct format: "Shares Outstanding 55.64M" (LESS SPECIFIC LAST)
                    r'Shares Outstanding\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Outstanding Shares\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Total Shares Outstanding\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Shares Outstanding\s+([\d,]+\.?\d*[MBK]?)',
                    r'Outstanding Shares\s+([\d,]+\.?\d*[MBK]?)',
                    r'Total Shares Outstanding\s+([\d,]+\.?\d*[MBK]?)'
                ],
                'float': [
                    # Handle column number format: "Float 8 19.7M" (MOST SPECIFIC FIRST)
                    r'Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Shares Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Public Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    r'Shares Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    r'Public Float\s+\d+\s+([\d,]+\.?\d*[MBK]?)',
                    # Handle direct format: "Float 19.7M" (LESS SPECIFIC LAST)
                    r'Float\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Shares Float\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Public Float\s+([\d,]+\.?\d*[MBK]?)%',
                    r'Float\s+([\d,]+\.?\d*[MBK]?)',
                    r'Shares Float\s+([\d,]+\.?\d*[MBK]?)',
                    r'Public Float\s+([\d,]+\.?\d*[MBK]?)'
                ],
                'float_percentage': [
                    r'Float %\s+\d+\s+([\d,]+\.?\d*%)',
                    r'Float Percentage\s+\d+\s+([\d,]+\.?\d*%)',
                    r'% of Float\s+\d+\s+([\d,]+\.?\d*%)'
                ],
                'institutional_ownership': [
                    r'% Held by Institutions\s+\d+\s+([\d,]+\.?\d*%)',
                    r'Institutional Ownership\s+\d+\s+([\d,]+\.?\d*%)'
                ],
                'insider_ownership': [
                    r'% Held by Insiders\s+\d+\s+([\d,]+\.?\d*%)',
                    r'Insider Ownership\s+\d+\s+([\d,]+\.?\d*%)'
                ]
            }
            
            # Extract data using fixed patterns
            for field, pattern_list in patterns.items():
                for pattern in pattern_list:
                    matches = re.findall(pattern, page_text, re.IGNORECASE)
                    if matches:
                        value = matches[0].strip()
                        float_data[field] = value
                        logging.debug(f"[WAYBACK] Found {field}: {value} with pattern: {pattern}")
                        break
            
            # Also look for table data with better parsing
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 2:
                        label = cells[0].get_text().strip().lower()
                        value = cells[1].get_text().strip()
                        
                        # Look for the actual financial values, not column numbers
                        if 'outstanding' in label and not float_data['shares_outstanding']:
                            # Extract the actual value (skip column numbers)
                            value_match = re.search(r'([\d,]+\.?\d*[MBK]?)', value)
                            if value_match:
                                float_data['shares_outstanding'] = value_match.group(1)
                        elif 'float' in label and not float_data['float']:
                            value_match = re.search(r'([\d,]+\.?\d*[MBK]?)', value)
                            if value_match:
                                float_data['float'] = value_match.group(1)
                        elif 'institutional' in label and not float_data['institutional_ownership']:
                            value_match = re.search(r'([\d,]+\.?\d*%)', value)
                            if value_match:
                                float_data['institutional_ownership'] = value_match.group(1)
                        elif 'insider' in label and not float_data['insider_ownership']:
                            value_match = re.search(r'([\d,]+\.?\d*%)', value)
                            if value_match:
                                float_data['insider_ownership'] = value_match.group(1)
            
            # Check if we found any meaningful data
            has_data = any(value is not None for key, value in float_data.items() 
                          if key not in ['ticker', 'date'])
            
            if has_data:
                return float_data
            else:
                return None
                
        except Exception as e:
            logging.error(f"[WAYBACK] Error extracting float data: {e}")
            return None
    
    def convert_to_shares(self, value_str):
        """
        Convert string values like "1.5M" or "1,500,000" to actual share counts.
        
        Args:
            value_str (str): String representation of shares
            
        Returns:
            int: Number of shares
        """
        try:
            if not value_str:
                return None
            
            # Remove commas and spaces
            value_str = value_str.replace(',', '').replace(' ', '')
            
            # Handle K, M, B suffixes
            multipliers = {'K': 1000, 'M': 1000000, 'B': 1000000000}
            
            for suffix, multiplier in multipliers.items():
                if value_str.upper().endswith(suffix):
                    number = float(value_str[:-1])
                    return int(number * multiplier)
            
            # Try to parse as regular number
            return int(float(value_str))
            
        except Exception as e:
            logging.warning(f"[WAYBACK] Error converting {value_str} to shares: {e}")
            return None
    
    def get_float_analysis(self, ticker, target_date):
        """
        Get comprehensive float analysis.
        
        Args:
            ticker (str): Stock ticker symbol
            target_date (str): Target date in YYYY-MM-DD format
            
        Returns:
            dict: Complete float analysis
        """
        try:
            logging.info(f"[WAYBACK] Getting float analysis for {ticker} on {target_date}")
            
            # Try Wayback Machine for historical data only
            float_data = self.get_float_from_wayback(ticker, target_date)
            
            if not float_data:
                return {
                    'ticker': ticker,
                    'target_date': target_date,
                    'status': 'No data found',
                    'source': None,
                    'raw_data': None,
                    'converted_data': None
                }
            
            # Convert string values to numbers
            analysis = {
                'ticker': ticker,
                'target_date': target_date,
                'status': 'Data found',
                'source': float_data.get('date', 'wayback_machine'),  # Use the actual date used
                'raw_data': float_data,
                'converted_data': {}
            }
            
            # Convert shares outstanding
            if float_data['shares_outstanding']:
                shares_outstanding = self.convert_to_shares(float_data['shares_outstanding'])
                analysis['converted_data']['shares_outstanding'] = shares_outstanding
                logging.info(f"[WAYBACK] Converted shares outstanding: {shares_outstanding:,}")
            
            # Convert float
            if float_data['float']:
                float_shares = self.convert_to_shares(float_data['float'])
                analysis['converted_data']['float_shares'] = float_shares
                logging.info(f"[WAYBACK] Converted float shares: {float_shares:,}")
            
            # Calculate float percentage if we have both values
            if (analysis['converted_data'].get('float_shares') and 
                analysis['converted_data'].get('shares_outstanding')):
                float_pct = (analysis['converted_data']['float_shares'] / 
                           analysis['converted_data']['shares_outstanding']) * 100
                analysis['converted_data']['float_percentage_calculated'] = float_pct
                logging.info(f"[WAYBACK] Calculated float percentage: {float_pct:.2f}%")
            
            return analysis
            
        except Exception as e:
            logging.error(f"[WAYBACK] Error in float analysis for {ticker}: {e}")
            return {
                'ticker': ticker,
                'target_date': target_date,
                'status': f'Error: {e}',
                'source': None,
                'raw_data': None,
                'converted_data': None
            }

# Initialize global web scraper
wayback_scraper = WaybackFloatScraper()

# Polygon Stock Financials API for Historical Outstanding Shares
def get_polygon_ticker_details(ticker, date=None):
    """
    Get current outstanding shares from Polygon's Ticker Details API using RESTClient.
    
    Args:
        ticker (str): Stock ticker symbol
        date (str): Date in YYYY-MM-DD format (ignored - uses current data)
    
    Returns:
        dict: Ticker details including outstanding shares
    """
    try:
        if not POLYGON_API_KEY:
            logging.warning("[POLYGON] API key not available")
            return None
        
        polygon_limiter.acquire()
        
        # Initialize Polygon client
        client = RESTClient(POLYGON_API_KEY)
        
        logging.info(f"[POLYGON] Fetching current ticker details for {ticker}")
        
        # Get current ticker details
        details = client.get_ticker_details(ticker)
        
        if not details:
            logging.warning(f"[POLYGON] No ticker details found for {ticker}")
            return None
        
        # Extract outstanding shares from current ticker details
        outstanding_shares = None
        source_field = None
        
        # Try weighted_shares_outstanding first (most accurate for current data)
        if hasattr(details, 'weighted_shares_outstanding') and details.weighted_shares_outstanding:
            outstanding_shares = details.weighted_shares_outstanding
            source_field = "weighted_shares_outstanding"
        
        # Fallback to shares_outstanding (total of all classes)
        elif hasattr(details, 'shares_outstanding') and details.shares_outstanding:
            outstanding_shares = details.shares_outstanding
            source_field = "shares_outstanding"
        
        if outstanding_shares and outstanding_shares > 0 and outstanding_shares < 1000000000000:
            logging.info(f"[POLYGON] Found current outstanding shares ({source_field}): {outstanding_shares:,.0f} for {ticker}")
            return {
                'outstanding_shares': outstanding_shares,
                'filing_date': 'Current',
                'source_field': source_field
            }
        else:
            logging.warning(f"[POLYGON] No valid outstanding shares found for {ticker}")
            return None
            
    except Exception as e:
        logging.error(f"[POLYGON] Error in get_polygon_ticker_details for {ticker}: {e}")
        return None
    finally:
        polygon_limiter.release()
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Increase connection pool size to prevent "Connection pool is full" errors
urllib3.util.Retry.DEFAULT_ALLOWED_METHODS = frozenset(['GET', 'POST', 'PUT', 'DELETE', 'HEAD', 'OPTIONS', 'TRACE'])
urllib3.util.Retry.DEFAULT_STATUS_FORCELIST = [500, 502, 503, 504]

# Configure requests session with larger connection pool
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Create a session with larger connection pool
session = requests.Session()
retry_strategy = Retry(
    total=3,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
)
adapter = HTTPAdapter(
    pool_connections=100,  # Increased from default 10
    pool_maxsize=100,      # Increased from default 10
    max_retries=retry_strategy
)
session.mount("http://", adapter)
session.mount("https://", adapter)

# Configure Polygon client settings with optimized pool size for 30 workers
POLYGON_NUM_POOLS = 500  # Increased for 30 workers with multiple API calls each
POLYGON_RETRIES = 3

# Create a single global client instance to reuse connections
_polygon_client = None

def get_polygon_client():
    """Get or create the global Polygon client instance."""
    global _polygon_client
    if _polygon_client is None:
        _polygon_client = RESTClient(POLYGON_API_KEY, num_pools=POLYGON_NUM_POOLS, retries=POLYGON_RETRIES)
    return _polygon_client

def cleanup_polygon_client():
    """Clean up the global Polygon client to free connections."""
    global _polygon_client
    if _polygon_client is not None:
        try:
            # Close any open connections
            if hasattr(_polygon_client, '_session'):
                _polygon_client._session.close()
        except Exception as e:
            logging.warning(f"Error cleaning up Polygon client: {e}")
        _polygon_client = None

# Memory optimization settings
def optimize_memory():
    """Optimize memory usage for the script."""
    # Force garbage collection
    gc.collect()
    
    # Set pandas memory usage optimization
    pd.options.mode.chained_assignment = None  # default='warn'
    
    # Increase memory limit for pandas
    import warnings
    warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)
    
    # Set numpy to use less memory
    np.set_printoptions(precision=6, suppress=True)
    
    # Log memory usage
    process = psutil.Process(os.getpid())
    logging.info(f"Initial memory usage: {process.memory_info().rss / 1024 / 1024:.2f} MB")

def cleanup_memory():
    """Clean up memory periodically."""
    gc.collect()
    process = psutil.Process(os.getpid())
    memory_mb = process.memory_info().rss / 1024 / 1024
    logging.info(f"Memory usage: {memory_mb:.2f} MB")
    
    # Force cleanup if memory usage is high
    if memory_mb > 4000:  # Increased from 1000 to 4000 for speed optimization
        logging.warning(f"High memory usage detected: {memory_mb:.2f} MB")
        gc.collect()
        return True
    return False

def monitor_memory():
    """Monitor memory usage and log it."""
    process = psutil.Process(os.getpid())
    memory_mb = process.memory_info().rss / 1024 / 1024
    logging.info(f"Current memory usage: {memory_mb:.2f} MB")
    return memory_mb

# Removed nest_asyncio - no longer needed since we removed requests-html
# Removed nest_asyncio.apply() - no longer needed

# Call memory optimization
optimize_memory()

# Initialize clients
finnhub_client = finnhub.Client(api_key=FINNHUB_API_KEY)

# Quandl API already configured above

# AWS S3 bucket
S3_BUCKET = "thanh-stock-data-2025"

# Time zone
EASTERN_TZ = pytz.timezone("US/Eastern")

# Cache removed for dynamic date processing

# Delisted tickers
DELISTED_TICKERS = {"ACHL": datetime(2025, 3, 20).replace(tzinfo=EASTERN_TZ)}

def colorize_output(text, use_color=True):
    """Apply purple ANSI color to text if terminal supports it and use_color is True."""
    if use_color and sys.stdout.isatty():
        return f"\033[95m{text}\033[0m"
    return text



def colorize_yfinance(text, use_color=True):
    """Apply red ANSI color for YFinance data."""
    if use_color and sys.stdout.isatty():
        return f"\033[91m{text}\033[0m"
    return text

def colorize_polygon(text, use_color=True):
    """Apply purple ANSI color for Polygon data."""
    if use_color and sys.stdout.isatty():
        return f"\033[95m{text}\033[0m"
    return text

def format_number(value):
    """Format a number with K or M suffix, handling small and large values."""
    try:
        if isinstance(value, str) and value.replace(".", "").replace("-", "").isdigit():
            value = float(value)
        if not isinstance(value, (int, float)) or value in ("Not Available", "No PM Data", "No Regular Data", "No AH Data"):
            return str(value)
        value = float(value)
        if value >= 1_000_000:
            return f"{round(value / 1_000_000, 1)}M"
        elif value >= 1_000:
            return f"{round(value / 1_000, 1)}K"
        elif value >= 1:
            return str(round(value, 1))
        else:
            return str(round(value, 3))
    except Exception as e:
        logging.warning(f"Error formatting number {value}: {e}")
        return str(value)
# Dilution tracking functionality
class DilutionExtractor:
    """Extract dilution data using the same pattern as the working SEC API code."""
    
    def __init__(self):
        self.api_key = SEC_API_KEY
        if not self.api_key:
            raise ValueError("SEC_API_KEY not found in environment variables")
        
        self.base_url = "https://api.sec-api.io"
        logging.info("DilutionExtractor initialized successfully")
    
    def extract_active_dilution(self, ticker: str, target_date: str, lookback_days: int = 1095) -> dict:
        """Extract active dilution data for a ticker as of a specific date."""
        logging.info(f"[DILUTION] Extracting dilution data for {ticker} as of {target_date}")
        
        try:
            # Convert target_date to datetime
            target_dt = datetime.strptime(target_date, "%Y-%m-%d")
            lookback_start = target_dt - timedelta(days=lookback_days)
            
            logging.info(f"[DILUTION] Lookback period: {lookback_start.strftime('%Y-%m-%d')} to {target_date}")
            
            # Get CIK for ticker (using the same method as SECAdvanced)
            cik = self._get_cik(ticker)
            if not cik:
                logging.warning(f"[DILUTION] Could not find CIK for {ticker}")
                return self._get_default_result()
            
            logging.info(f"[DILUTION] Found CIK: {cik}")
            
            # Query for dilution-related filings
            filings = self._query_dilution_filings(cik, ticker, target_date, lookback_start)
            logging.info(f"[DILUTION] Found {len(filings)} dilution-related filings")
            
            if not filings:
                logging.warning(f"[DILUTION] No filings found for {ticker} in the specified period")
                return self._get_default_result()
            
            # Extract dilution data from each filing
            all_dilution_data = []
            successful_extractions = 0
            
            for i, filing in enumerate(filings):
                logging.info(f"[DILUTION] Processing filing {i+1}/{len(filings)}: {filing.get('formType', 'Unknown')} - {filing.get('filedAt', 'Unknown')}")
                
                filing_data = self._extract_dilution_from_filing(filing, target_dt)
                if filing_data:
                    all_dilution_data.append(filing_data)
                    successful_extractions += 1
                    
                    # Log what was found
                    warrants_found = len(filing_data.get('warrants', []))
                    convertibles_found = len(filing_data.get('convertibles', []))
                    logging.info(f"[DILUTION] Extracted from {filing.get('formType', 'Unknown')}: {warrants_found} warrants, {convertibles_found} convertibles")
                else:
                    logging.debug(f"[DILUTION] No dilution data found in {filing.get('formType', 'Unknown')}")
            
            logging.info(f"[DILUTION] Successfully extracted from {successful_extractions}/{len(filings)} filings")
            
            # Filter for active instruments only
            active_dilution = self._filter_active_instruments(all_dilution_data, target_dt)
            
            # Calculate final metrics
            result = self._calculate_dilution_metrics(active_dilution)
            
            # Enhanced diagnostics (Grok's suggestion)
            logging.info(f"[DILUTION] For {ticker}: CIK={cik}, Filings={len(filings)}, Active Warrants={result['active_warrant_shares']:,}, Active Convertibles={result['active_convertible_shares']:,}")
            
            return result
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting dilution data: {str(e)}")
            import traceback
            logging.error(f"[DILUTION] Traceback: {traceback.format_exc()}")
            return self._get_default_result()
    
    def _get_cik(self, ticker: str) -> Optional[str]:
        """Get CIK for ticker using the same method as SECAdvanced."""
        try:
            # Known ticker name changes - add more as discovered
            ticker_cik_map = {
                "SHOT": "1760903",  # JUPW/SHOT (without leading zeros)
                # Add more ticker changes here as discovered
            }
            
            if ticker in ticker_cik_map:
                logging.info(f"[DILUTION] Using known CIK for {ticker}: {ticker_cik_map[ticker]}")
                return ticker_cik_map[ticker]
            
            # Try to get CIK from SEC API using direct HTTP request
            query_str = f"ticker:{ticker}"
            payload = {
                "query": query_str, 
                "from": "0", 
                "size": "1", 
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            headers = {"Authorization": self.api_key, "Content-Type": "application/json"}
            
            response = requests.post(self.base_url, json=payload, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            filings = data.get("filings", [])
            if filings:
                cik = filings[0].get("cik")
                if cik:
                    logging.info(f"[DILUTION] Found CIK for {ticker}: {cik}")
                    return cik
            
            logging.warning(f"[DILUTION] Could not find CIK for {ticker}")
            return None
            
        except Exception as e:
            logging.error(f"[DILUTION] Error getting CIK for {ticker}: {e}")
            return None
    
    def _query_dilution_filings(self, cik: str, ticker: str, target_date: str, lookback_start: datetime) -> List[Dict]:
        """Query for dilution-related filings using both Query API and Full-Text Search API."""
        try:
            all_filings = []
            
            # Query 1: 8-K and D forms with description field (supported)
            query_str_8k = f"ticker:{ticker} AND formType:(8-K OR D OR 8-K/A OR D/A) AND (description:\"warrant\" OR description:\"convertible\" OR description:\"dilution\" OR description:\"offering\" OR description:\"private placement\" OR description:\"securities\")"
            
            payload_8k = {
                "query": query_str_8k,
                "from": lookback_start.strftime("%Y-%m-%d"),
                "to": target_date,
                "page": 0,
                "size": 50,
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            headers = {"Authorization": self.api_key, "Content-Type": "application/json"}
            
            response_8k = requests.post(self.base_url, json=payload_8k, headers=headers, timeout=10)
            response_8k.raise_for_status()
            data_8k = response_8k.json()
            filings_8k = data_8k.get("filings", [])
            logging.info(f"[DILUTION] 8-K/D query found {len(filings_8k)} filings")
            all_filings.extend(filings_8k)
            
            # Query 2: Registration statements (S-1, S-3, 424B4) - broader search
            query_str_reg = f"ticker:{ticker} AND formType:(S-1 OR S-3 OR 424B4 OR S-1/A OR S-3/A)"
            
            payload_reg = {
                "query": query_str_reg,
                "from": lookback_start.strftime("%Y-%m-%d"),
                "to": target_date,
                "page": 0,
                "size": 30,
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            
            response_reg = requests.post(self.base_url, json=payload_reg, headers=headers, timeout=10)
            response_reg.raise_for_status()
            data_reg = response_reg.json()
            filings_reg = data_reg.get("filings", [])
            logging.info(f"[DILUTION] Registration query found {len(filings_reg)} filings")
            all_filings.extend(filings_reg)
            
            # Query 3: Financial statements (10-Q, 10-K) - broader search
            query_str_fin = f"ticker:{ticker} AND formType:(10-Q OR 10-K OR 10-Q/A OR 10-K/A)"
            
            payload_fin = {
                "query": query_str_fin,
                "from": lookback_start.strftime("%Y-%m-%d"),
                "to": target_date,
                "page": 0,
                "size": 30,
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            
            response_fin = requests.post(self.base_url, json=payload_fin, headers=headers, timeout=10)
            response_fin.raise_for_status()
            data_fin = response_fin.json()
            filings_fin = data_fin.get("filings", [])
            logging.info(f"[DILUTION] Financial query found {len(filings_fin)} filings")
            all_filings.extend(filings_fin)
            
            # Query 4: Full-Text Search for content-based dilution detection (NEW)
            full_text_url = f"{self.base_url}/full-text-search"
            query_str_full = f"ticker:{ticker} AND formType:(S-1 OR S-3 OR 424B4 OR 10-Q OR 10-K) AND (warrant OR convertible OR dilution OR offering OR \"private placement\")"
            payload_full = {
                "query": query_str_full,
                "from": lookback_start.strftime("%Y-%m-%d"),
                "to": target_date,
                "size": 100,  # Increased for more results
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            
            try:
                response_full = requests.post(full_text_url, json=payload_full, headers=headers, timeout=10)
                if response_full.status_code == 200:
                    data_full = response_full.json()
                    filings_full = data_full.get("filings", [])
                    logging.info(f"[DILUTION] Full-text search found {len(filings_full)} filings")
                    all_filings.extend(filings_full)
                else:
                    logging.warning(f"[DILUTION] Full-text search error: {response_full.status_code} - {response_full.text[:200]}")
            except Exception as e:
                logging.warning(f"[DILUTION] Full-text search failed: {e}")
            
            # Remove duplicates based on accession number
            unique_filings = {}
            for filing in all_filings:
                accession = filing.get('accessionNo')
                if accession and accession not in unique_filings:
                    unique_filings[accession] = filing
            
            result = list(unique_filings.values())
            logging.info(f"[DILUTION] Total unique filings found: {len(result)}")
            return result
            
        except Exception as e:
            logging.error(f"[DILUTION] Error querying filings: {str(e)}")
            return []
    
    def _extract_dilution_from_filing(self, filing: Dict, target_dt: datetime) -> Optional[Dict]:
        """Extract dilution data from a specific filing."""
        try:
            filing_type = filing.get('formType', '')
            filing_url = filing.get('linkToHtml') or filing.get('linkToTxt')
            filed_date = datetime.strptime(filing.get('filedAt', ''), "%Y-%m-%d")
            
            logging.info(f"[DILUTION] Processing {filing_type} from {filed_date.strftime('%Y-%m-%d')}")
            
            if not filing_url:
                logging.warning(f"[DILUTION] No URL found for {filing_type}")
                return None
            
            # Extract content based on filing type
            if filing_type == "8-K":
                return self._extract_from_8k(filing_url, filed_date)
            elif filing_type in ["10-Q", "10-K"]:
                return self._extract_from_financial_filing(filing_url, filed_date)
            elif filing_type in ["S-1", "S-3", "424B4"]:
                return self._extract_from_registration(filing_url, filed_date)
            elif filing_type == "D":
                return self._extract_from_form_d(filing_url, filed_date)
            else:
                logging.warning(f"[DILUTION] Unsupported filing type: {filing_type}")
                return None
                
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from {filing.get('formType', '')}: {str(e)}")
            return None
    
    def _extract_from_8k(self, filing_url: str, filed_date: datetime) -> Optional[Dict]:
        """Extract from 8-K immediate events with enhanced section coverage."""
        try:
            # Try different 8-K sections
            sections = [
                "Item 1.01", "Item 3.02", "Item 3.03", "Item 4.01", "Item 4.02",
                "Description of Securities", "Securities", "Warrants", "Convertible Securities"
            ]
            
            for section in sections:
                content = self._get_section_content(filing_url, section)
                if content:
                    warrant_data = self._parse_warrant_text(content)
                    convertible_data = self._parse_convertible_text(content)
                    
                    if warrant_data or convertible_data:
                        logging.info(f"[DILUTION] Found dilution data in 8-K section: {section}")
                        return {
                            'filing_date': filed_date,
                            'filing_type': '8-K',
                            'warrants': warrant_data or [],
                            'convertibles': convertible_data or []
                        }
            
            # Fallback: try full-text search if section extraction fails
            logging.info(f"[DILUTION] No dilution found in 8-K sections, trying full-text fallback")
            return self._extract_from_full_text(filing_url, filed_date, '8-K')
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from 8-K: {str(e)}")
            return None
    
    def _extract_from_financial_filing(self, filing_url: str, filed_date: datetime) -> Optional[Dict]:
        """Extract from 10-Q/10-K financial footnotes with enhanced section coverage."""
        try:
            # Try different financial statement sections
            sections = [
                "Notes to Financial Statements", "Note 1", "Note 2", "Note 3", "Note 4", "Note 5",
                "Stockholders' Equity", "Capital Stock", "Warrants", "Convertible Securities",
                "Long-term Debt", "Notes Payable", "Derivative Instruments"
            ]
            
            for section in sections:
                content = self._get_section_content(filing_url, section)
                if content:
                    warrant_data = self._parse_warrant_text(content)
                    convertible_data = self._parse_convertible_text(content)
                    
                    if warrant_data or convertible_data:
                        logging.info(f"[DILUTION] Found dilution data in financial section: {section}")
                        return {
                            'filing_date': filed_date,
                            'filing_type': '10-Q/10-K',
                            'warrants': warrant_data or [],
                            'convertibles': convertible_data or []
                        }
            
            # Fallback: try full-text search if section extraction fails
            logging.info(f"[DILUTION] No dilution found in financial sections, trying full-text fallback")
            return self._extract_from_full_text(filing_url, filed_date, '10-Q/10-K')
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from financial filing: {str(e)}")
            return None
    
    def _extract_from_registration(self, filing_url: str, filed_date: datetime) -> Optional[Dict]:
        """Extract from S-1/S-3/424B4 registration statements with enhanced section coverage."""
        try:
            # Try different registration statement sections
            sections = [
                "Description of Securities", "Description of Capital Stock", "Description of Common Stock",
                "Warrants", "Convertible Securities", "Use of Proceeds", "Plan of Distribution",
                "Exhibit 4.1", "Exhibit 4.2", "Exhibit 4.3", "Exhibit 4.4", "Exhibit 4.5"
            ]
            
            for section in sections:
                content = self._get_section_content(filing_url, section)
                if content:
                    warrant_data = self._parse_warrant_text(content)
                    convertible_data = self._parse_convertible_text(content)
                    
                    if warrant_data or convertible_data:
                        logging.info(f"[DILUTION] Found dilution data in registration section: {section}")
                        return {
                            'filing_date': filed_date,
                            'filing_type': 'S-1/S-3/424B4',
                            'warrants': warrant_data or [],
                            'convertibles': convertible_data or []
                        }
            
            # Fallback: try full-text search if section extraction fails
            logging.info(f"[DILUTION] No dilution found in registration sections, trying full-text fallback")
            return self._extract_from_full_text(filing_url, filed_date, 'S-1/S-3/424B4')
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from registration: {str(e)}")
            return None
    
    def _extract_from_form_d(self, filing_url: str, filed_date: datetime) -> Optional[Dict]:
        """Extract from Form D private placement."""
        try:
            content = self._get_section_content(filing_url, "Item 4. Description of Securities")
            if content:
                warrant_data = self._parse_warrant_text(content)
                convertible_data = self._parse_convertible_text(content)
                
                if warrant_data or convertible_data:
                    return {
                        'filing_date': filed_date,
                        'filing_type': 'Form D',
                        'warrants': warrant_data or [],
                        'convertibles': convertible_data or []
                    }
            
            return None
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from Form D: {str(e)}")
            return None
    
    def _get_section_content(self, filing_url: str, section: str) -> Optional[str]:
        """Get section content using Extractor API."""
        try:
            extractor_url = f"{self.base_url}/extract"
            params = {
                "url": filing_url,
                "section": section
            }
            headers = {"Authorization": self.api_key}
            
            response = requests.get(extractor_url, params=params, headers=headers, timeout=10)
            if response.status_code == 200:
                return response.text
            else:
                logging.warning(f"[DILUTION] Extractor API returned {response.status_code} for section {section}")
                return None
                
        except Exception as e:
            logging.error(f"[DILUTION] Error getting section content: {str(e)}")
            return None
    
    def _extract_from_full_text(self, filing_url: str, filed_date: datetime, filing_type: str) -> Optional[Dict]:
        """Extract dilution data from full filing text as fallback."""
        try:
            # Get full filing text
            full_text = self._get_full_filing_text(filing_url)
            if not full_text:
                return None
            
            # Parse the full text for dilution data
            warrant_data = self._parse_warrant_text(full_text)
            convertible_data = self._parse_convertible_text(full_text)
            
            if warrant_data or convertible_data:
                logging.info(f"[DILUTION] Found dilution data in full text of {filing_type}")
                return {
                    'filing_date': filed_date,
                    'filing_type': filing_type,
                    'warrants': warrant_data or [],
                    'convertibles': convertible_data or []
                }
            
            return None
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting from full text: {str(e)}")
            return None
    
    def _get_full_filing_text(self, filing_url: str) -> Optional[str]:
        """Get full filing text using the filing endpoint."""
        try:
            # Use the filing endpoint to get full text
            filing_url_api = f"{self.base_url}/filing"
            params = {
                "url": filing_url
            }
            headers = {"Authorization": self.api_key}
            
            response = requests.get(filing_url_api, params=params, headers=headers, timeout=15)
            if response.status_code == 200:
                return response.text
            else:
                logging.warning(f"[DILUTION] Filing API returned {response.status_code}")
                return None
                
        except Exception as e:
            logging.error(f"[DILUTION] Error getting full filing text: {str(e)}")
            return None
    
    def _parse_warrant_text(self, text: str) -> List[Dict]:
        """Parse warrant information from text with enhanced patterns."""
        warrants = []
        
        # Enhanced patterns for warrant extraction
        patterns = [
            # Standard patterns
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|warrants?)\s*(?:of\s+common\s+stock\s+)?(?:warrants?|to\s+purchase)",
            r"warrants?\s*(?:to\s+purchase\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*shares?",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?\s*(?:exercisable|expiring)",
            
            # Enhanced patterns for various phrasings
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|warrants?)\s*(?:issuable\s+upon\s+exercise|to\s+purchase|exercisable\s+for|underlying\s+warrants?)",
            r"warrants?\s*(?:entitling\s+the\s+holder\s+to\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*shares?",
            r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?\s+of\s+common\s+stock\s+)?(?:underlying\s+warrants?|warrant\s+shares?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?\s*(?:to\s+acquire|for\s+the\s+purchase\s+of)",
            r"warrants?\s*(?:representing\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|common\s+stock)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?\s+issuable\s+upon\s+exercise\s+of\s+warrants?)",
            
            # Patterns with units (M, K, million, thousand)
            r"(\d+(?:\.\d+)?)\s*(?:M|million)\s*(?:shares?|warrants?)\s*(?:of\s+common\s+stock\s+)?(?:warrants?|to\s+purchase)",
            r"(\d+(?:\.\d+)?)\s*(?:K|thousand)\s*(?:shares?|warrants?)\s*(?:of\s+common\s+stock\s+)?(?:warrants?|to\s+purchase)",
            r"warrants?\s*(?:to\s+purchase\s+)?(\d+(?:\.\d+)?)\s*(?:M|million)\s*shares?",
            r"warrants?\s*(?:to\s+purchase\s+)?(\d+(?:\.\d+)?)\s*(?:K|thousand)\s*shares?",
            
            # NEW: Small-cap specific patterns (Grok's suggestions)
            r"issued\s+a\s+total\s+of\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:warrants?|shares?\s+upon\s+exercise)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:warrants?\s+exercisable|underlying\s+warrants?|convertible\s+notes?\s+warrants?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?\s*(?:outstanding|issued|granted)",
            r"total\s+of\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?\s*(?:were\s+issued|have\s+been\s+issued)",
            
            # Patterns for exercised warrants (to exclude)
            r"exercised\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:warrants?|shares?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*warrants?\s*exercised",
        ]
        
        for pattern in patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                # Skip if this is an exercised warrant pattern
                if "exercised" in pattern.lower():
                    continue
                    
                shares = self._parse_number(match.group(1))
                if shares > 0:
                    expiration = self._extract_expiration_date(text, match.start())
                    
                    # Check if this warrant is already expired or exercised
                    if self._is_warrant_active(text, match.start(), expiration):
                        warrants.append({
                            'shares': shares,
                            'expiration_date': expiration,
                            'type': 'warrant',
                            'context': text[max(0, match.start()-50):match.end()+50]  # Add context for debugging
                        })
        
        return warrants
    
    def _parse_convertible_text(self, text: str) -> List[Dict]:
        """Parse convertible note information from text with enhanced patterns."""
        convertibles = []
        
        # Enhanced patterns for convertible extraction
        patterns = [
            # Standard patterns
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|notes?)\s*(?:of\s+common\s+stock\s+)?(?:convertible|notes?)",
            r"convertible\s*(?:notes?|debt)\s*(?:of\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*shares?",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)",
            
            # Enhanced patterns for various phrasings
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|notes?)\s*(?:issuable\s+upon\s+conversion|convertible\s+into|underlying\s+convertible)",
            r"convertible\s*(?:notes?|debt|securities)\s*(?:entitling\s+the\s+holder\s+to\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*shares?",
            r"(\d+(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?\s+of\s+common\s+stock\s+)?(?:underlying\s+convertible|convertible\s+shares?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)\s*(?:to\s+acquire|for\s+the\s+conversion\s+of)",
            r"convertible\s*(?:notes?|securities)\s*(?:representing\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?|common\s+stock)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares?\s+issuable\s+upon\s+conversion\s+of\s+convertible)",
            
            # Patterns with units (M, K, million, thousand)
            r"(\d+(?:\.\d+)?)\s*(?:M|million)\s*(?:shares?|notes?)\s*(?:of\s+common\s+stock\s+)?(?:convertible|notes?)",
            r"(\d+(?:\.\d+)?)\s*(?:K|thousand)\s*(?:shares?|notes?)\s*(?:of\s+common\s+stock\s+)?(?:convertible|notes?)",
            r"convertible\s*(?:notes?|debt)\s*(?:of\s+)?(\d+(?:\.\d+)?)\s*(?:M|million)\s*shares?",
            r"convertible\s*(?:notes?|debt)\s*(?:of\s+)?(\d+(?:\.\d+)?)\s*(?:K|thousand)\s*shares?",
            
            # NEW: Small-cap specific patterns (Grok's suggestions)
            r"issued\s+a\s+total\s+of\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:convertible|notes?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:convertible\s+notes?\s+outstanding|underlying\s+convertible)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)\s*(?:outstanding|issued|granted)",
            r"total\s+of\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)\s*(?:were\s+issued|have\s+been\s+issued)",
            
            # Patterns for converted notes (to exclude)
            r"converted\s+(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:convertible|notes?|shares?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*convertible\s*(?:notes?|securities)\s*converted",
        ]
        
        for pattern in patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                # Skip if this is a converted note pattern
                if "converted" in pattern.lower():
                    continue
                    
                shares = self._parse_number(match.group(1))
                if shares > 0:
                    maturity = self._extract_maturity_date(text, match.start())
                    
                    # Check if this convertible is still active
                    if self._is_convertible_active(text, match.start(), maturity):
                        convertibles.append({
                            'shares': shares,
                            'maturity_date': maturity,
                            'type': 'convertible',
                            'context': text[max(0, match.start()-50):match.end()+50]  # Add context for debugging
                        })
        
        return convertibles
    
    def _parse_number(self, text: str) -> int:
        """Parse number from text with enhanced unit handling."""
        try:
            text = text.replace(',', '').strip().upper()
            
            # Handle units (M, K, million, thousand)
            if 'M' in text or 'MILLION' in text:
                # Extract the number part
                number_part = re.sub(r'[^\d.]', '', text)
                return int(float(number_part) * 1_000_000)
            elif 'K' in text or 'THOUSAND' in text:
                # Extract the number part
                number_part = re.sub(r'[^\d.]', '', text)
                return int(float(number_part) * 1_000)
            else:
                # Regular number parsing
                return int(float(text))
        except Exception as e:
            logging.debug(f"[DILUTION] Error parsing number '{text}': {e}")
            return 0
    
    def _extract_expiration_date(self, text: str, start_pos: int) -> Optional[datetime]:
        """Extract expiration date near warrant mention with enhanced format support."""
        try:
            # Use larger context window for better date detection
            context = text[max(0, start_pos-500):start_pos+500]
            
            date_patterns = [
                # Standard formats
                r"expir(?:es?|ing)\s+(?:on\s+)?(\d{1,2}/\d{1,2}/\d{4})",
                r"expir(?:es?|ing)\s+(?:on\s+)?(\d{4}-\d{2}-\d{2})",
                r"(\d{1,2}/\d{1,2}/\d{4})\s*(?:expiration|expires)",
                r"(\d{4}-\d{2}-\d{2})\s*(?:expiration|expires)",
                
                # Enhanced formats
                r"expir(?:es?|ing)\s+(?:on\s+)?(\w+\s+\d{1,2},?\s+\d{4})",  # "December 31, 2025"
                r"expir(?:es?|ing)\s+(?:on\s+)?(\d{1,2}\s+\w+\s+\d{4})",    # "31 December 2025"
                r"(\w+\s+\d{1,2},?\s+\d{4})\s*(?:expiration|expires)",
                r"(\d{1,2}\s+\w+\s+\d{4})\s*(?:expiration|expires)",
                
                # Alternative phrasings
                r"expir(?:es?|ing)\s+(?:date\s+of\s+)?(\d{1,2}/\d{1,2}/\d{4})",
                r"expir(?:es?|ing)\s+(?:date\s+of\s+)?(\d{4}-\d{2}-\d{2})",
                r"expir(?:es?|ing)\s+(?:date\s+of\s+)?(\w+\s+\d{1,2},?\s+\d{4})",
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, context, re.IGNORECASE)
                if match:
                    date_str = match.group(1)
                    try:
                        # Try different date formats
                        if '/' in date_str:
                            return datetime.strptime(date_str, "%m/%d/%Y")
                        elif '-' in date_str and len(date_str.split('-')[0]) == 4:
                            return datetime.strptime(date_str, "%Y-%m-%d")
                        elif '-' in date_str and len(date_str.split('-')[0]) == 2:
                            return datetime.strptime(date_str, "%m-%d-%Y")
                        else:
                            # Try to parse month names
                            return self._parse_date_with_month(date_str)
                    except:
                        continue
            
            return None
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting expiration date: {str(e)}")
            return None
    
    def _extract_maturity_date(self, text: str, start_pos: int) -> Optional[datetime]:
        """Extract maturity date near convertible mention with enhanced format support."""
        try:
            # Use larger context window for better date detection
            context = text[max(0, start_pos-500):start_pos+500]
            
            date_patterns = [
                # Standard formats
                r"matur(?:es?|ity)\s+(?:on\s+)?(\d{1,2}/\d{1,2}/\d{4})",
                r"matur(?:es?|ity)\s+(?:on\s+)?(\d{4}-\d{2}-\d{2})",
                r"(\d{1,2}/\d{1,2}/\d{4})\s*(?:maturity|matures)",
                r"(\d{4}-\d{2}-\d{2})\s*(?:maturity|matures)",
                
                # Enhanced formats
                r"matur(?:es?|ity)\s+(?:on\s+)?(\w+\s+\d{1,2},?\s+\d{4})",  # "December 31, 2025"
                r"matur(?:es?|ity)\s+(?:on\s+)?(\d{1,2}\s+\w+\s+\d{4})",    # "31 December 2025"
                r"(\w+\s+\d{1,2},?\s+\d{4})\s*(?:maturity|matures)",
                r"(\d{1,2}\s+\w+\s+\d{4})\s*(?:maturity|matures)",
                
                # Alternative phrasings
                r"matur(?:es?|ity)\s+(?:date\s+of\s+)?(\d{1,2}/\d{1,2}/\d{4})",
                r"matur(?:es?|ity)\s+(?:date\s+of\s+)?(\d{4}-\d{2}-\d{2})",
                r"matur(?:es?|ity)\s+(?:date\s+of\s+)?(\w+\s+\d{1,2},?\s+\d{4})",
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, context, re.IGNORECASE)
                if match:
                    date_str = match.group(1)
                    try:
                        # Try different date formats
                        if '/' in date_str:
                            return datetime.strptime(date_str, "%m/%d/%Y")
                        elif '-' in date_str and len(date_str.split('-')[0]) == 4:
                            return datetime.strptime(date_str, "%Y-%m-%d")
                        elif '-' in date_str and len(date_str.split('-')[0]) == 2:
                            return datetime.strptime(date_str, "%m-%d-%Y")
                        else:
                            # Try to parse month names
                            return self._parse_date_with_month(date_str)
                    except:
                        continue
            
            return None
            
        except Exception as e:
            logging.error(f"[DILUTION] Error extracting maturity date: {str(e)}")
            return None
    
    def _filter_active_instruments(self, all_dilution_data: List[dict], target_dt: datetime) -> dict:
        """Filter for active instruments only (not expired/matured)."""
        active_warrants = []
        active_convertibles = []
        
        logging.info(f"[DILUTION] Filtering instruments active as of {target_dt.strftime('%Y-%m-%d')}")
        
        for filing_data in all_dilution_data:
            # Filter warrants
            for warrant in filing_data.get('warrants', []):
                expiration = warrant.get('expiration_date')
                if expiration and expiration > target_dt:
                    active_warrants.append(warrant)
                    logging.info(f"[DILUTION] Active warrant: {warrant['shares']:,} shares, expires {expiration.strftime('%Y-%m-%d')}")
                else:
                    logging.info(f"[DILUTION] Expired warrant: {warrant['shares']:,} shares")
            
            # Filter convertibles
            for convertible in filing_data.get('convertibles', []):
                maturity = convertible.get('maturity_date')
                if maturity and maturity > target_dt:
                    active_convertibles.append(convertible)
                    logging.info(f"[DILUTION] Active convertible: {convertible['shares']:,} shares, matures {maturity.strftime('%Y-%m-%d')}")
                else:
                    logging.info(f"[DILUTION] Matured convertible: {convertible['shares']:,} shares")
        
        return {
            'active_warrants': active_warrants,
            'active_convertibles': active_convertibles
        }
    
    def _calculate_dilution_metrics(self, active_dilution: dict) -> dict:
        """Calculate final dilution metrics."""
        active_warrant_shares = sum(w['shares'] for w in active_dilution['active_warrants'])
        active_convertible_shares = sum(c['shares'] for c in active_dilution['active_convertibles'])
        total_active_dilution = active_warrant_shares + active_convertible_shares
        
        return {
            'active_warrant_shares': active_warrant_shares,
            'active_convertible_shares': active_convertible_shares,
            'total_active_dilution': total_active_dilution,
            'warrant_count': len(active_dilution['active_warrants']),
            'convertible_count': len(active_dilution['active_convertibles'])
        }
    
    def _parse_date_with_month(self, date_str: str) -> Optional[datetime]:
        """Parse date string containing month names."""
        try:
            # Common month abbreviations and full names
            month_map = {
                'jan': 1, 'january': 1,
                'feb': 2, 'february': 2,
                'mar': 3, 'march': 3,
                'apr': 4, 'april': 4,
                'may': 5,
                'jun': 6, 'june': 6,
                'jul': 7, 'july': 7,
                'aug': 8, 'august': 8,
                'sep': 9, 'september': 9,
                'oct': 10, 'october': 10,
                'nov': 11, 'november': 11,
                'dec': 12, 'december': 12
            }
            
            # Clean the date string
            date_str = date_str.strip().lower()
            
            # Try different patterns
            patterns = [
                r"(\w+)\s+(\d{1,2}),?\s+(\d{4})",  # "December 31, 2025"
                r"(\d{1,2})\s+(\w+)\s+(\d{4})",    # "31 December 2025"
            ]
            
            for pattern in patterns:
                match = re.match(pattern, date_str)
                if match:
                    if len(match.groups()) == 3:
                        if match.group(1).isalpha():
                            # "December 31, 2025" format
                            month_name = match.group(1)
                            day = int(match.group(2))
                            year = int(match.group(3))
                        else:
                            # "31 December 2025" format
                            day = int(match.group(1))
                            month_name = match.group(2)
                            year = int(match.group(3))
                        
                        if month_name in month_map:
                            month = month_map[month_name]
                            return datetime(year, month, day)
            
            return None
            
        except Exception as e:
            logging.debug(f"[DILUTION] Error parsing date with month '{date_str}': {e}")
            return None
    
    def _is_warrant_active(self, text: str, start_pos: int, expiration_date: Optional[datetime]) -> bool:
        """Check if a warrant is still active (not exercised or expired)."""
        try:
            # Use context around the warrant mention
            context = text[max(0, start_pos-200):start_pos+200].lower()
            
            # Check for exercised indicators
            exercise_indicators = [
                'exercised', 'exercise', 'cancelled', 'cancelled', 'terminated', 
                'expired', 'lapsed', 'void', 'surrendered'
            ]
            
            for indicator in exercise_indicators:
                if indicator in context:
                    return False
            
            # Check expiration date if available
            if expiration_date and expiration_date < datetime.now():
                return False
            
            return True
            
        except Exception as e:
            logging.debug(f"[DILUTION] Error checking warrant status: {e}")
            return True  # Default to active if we can't determine
    
    def _is_convertible_active(self, text: str, start_pos: int, maturity_date: Optional[datetime]) -> bool:
        """Check if a convertible is still active (not converted or matured)."""
        try:
            # Use context around the convertible mention
            context = text[max(0, start_pos-200):start_pos+200].lower()
            
            # Check for converted indicators
            conversion_indicators = [
                'converted', 'conversion', 'cancelled', 'cancelled', 'terminated', 
                'matured', 'paid', 'redeemed', 'settled'
            ]
            
            for indicator in conversion_indicators:
                if indicator in context:
                    return False
            
            # Check maturity date if available
            if maturity_date and maturity_date < datetime.now():
                return False
            
            return True
            
        except Exception as e:
            logging.debug(f"[DILUTION] Error checking convertible status: {e}")
            return True  # Default to active if we can't determine
    
    def _get_default_result(self) -> dict:
        """Return default result structure."""
        return {
            'active_warrant_shares': 0,
            'active_convertible_shares': 0,
            'total_active_dilution': 0,
            'warrant_count': 0,
            'convertible_count': 0
        }
def get_dilution_data(ticker: str, target_date: str) -> dict:
    """
    Get dilution data for a ticker as of a specific date.
    
    Args:
        ticker (str): Stock ticker symbol
        target_date (str): Target date in YYYY-MM-DD format
    
    Returns:
        dict: Dilution data with active warrants and convertibles
    """
    try:
        if not SEC_API_KEY or not SEC_WORKING:
            logging.warning("[DILUTION] SEC API not available")
            return {
                'active_warrant_shares': 0,
                'active_convertible_shares': 0,
                'total_active_dilution': 0,
                'warrant_count': 0,
                'convertible_count': 0,
                'dilution_vs_outstanding_pct': 0.0,
                'dilution_vs_float_pct': 0.0,
                'float_dilution_risk': 'Low'
            }
        
        # Initialize dilution extractor
        extractor = DilutionExtractor()
        
        # Extract dilution data
        dilution_data = extractor.extract_active_dilution(ticker, target_date)
        
        return dilution_data
        
    except Exception as e:
        logging.error(f"[DILUTION] Error getting dilution data for {ticker}: {e}")
        return {
            'active_warrant_shares': 0,
            'active_convertible_shares': 0,
            'total_active_dilution': 0,
            'warrant_count': 0,
            'convertible_count': 0,
            'dilution_vs_outstanding_pct': 0.0,
            'dilution_vs_float_pct': 0.0,
            'float_dilution_risk': 'Low'
        }

# SEC API functions re-enabled for fallback system
def get_sec_float_data(ticker, date=None):
    """
    Get float and shares outstanding data from SEC API with historical date support.
    
    Args:
        ticker (str): Stock ticker symbol
        date (str): Date in YYYY-MM-DD format (optional)
    
    Returns:
        dict: Float and shares data from SEC API
    """
    try:
        logging.info(f"[SEC] Fetching data for {ticker} (date: {date})")
        
        if not SEC_API_KEY or not SEC_WORKING:
            logging.warning("[SEC] API not available or not working")
            return {"shares_outstanding": None, "float_shares": None, "timing": None}
        
        # Initialize SEC API with error handling
        try:
            sec_api = SECAdvanced()
        except ValueError as e:
            logging.warning(f"[SEC] API key invalid or missing: {e}")
            return {"shares_outstanding": None, "float_shares": None, "timing": None}
        
        # Get comprehensive data
        data = sec_api.get_float_and_shares_data(ticker, date)
        
        if data:
            shares_outstanding = data.get("shares_outstanding")
            float_shares = data.get("float_shares")
            timing = data.get("timing", "unknown")
            
            # Format values safely, handling None values
            shares_str = f"{shares_outstanding:,}" if shares_outstanding is not None else "None"
            float_str = f"{float_shares:,}" if float_shares is not None else "None"
            
            logging.info(f"[SEC] SUCCESS: {ticker} - {shares_str} shares outstanding, {float_str} float shares (timing: {timing})")
            
            # Return formatted result with timing
            return {
                "shares_outstanding": shares_outstanding,
                "float_shares": float_shares,
                "timing": timing
            }
        else:
            logging.warning(f"[SEC] No data found for {ticker}")
            return {
                "shares_outstanding": None,
                "float_shares": None,
                "timing": None
            }
            
    except Exception as e:
        logging.error(f"[SEC] Error fetching data for {ticker}: {e}")
        return {
            "shares_outstanding": None,
            "float_shares": None,
            "timing": None
        }

def get_sec_fallback_data(ticker, date, primary_data=None):
    """
    Get float and shares data using fallback methods when primary endpoint is incomplete.
    
    Args:
        ticker (str): Stock ticker symbol
        date (str): Date in YYYY-MM-DD format
        primary_data (dict): Data from primary endpoint (optional)
    
    Returns:
        dict: Float and shares data from fallback methods
    """
    try:
        if not SEC_API_KEY or not SEC_WORKING:
            return None
        
        # Initialize SEC API with error handling
        try:
            sec_api = SECAdvanced()
        except ValueError as e:
            logging.warning(f"[SEC] API key invalid or missing in fallback: {e}")
            return None
        
        # Determine what data we already have
        existing_shares = primary_data.get("shares_outstanding") if primary_data else None
        existing_float = primary_data.get("float_shares") if primary_data else None
        
        logging.info(f"[SEC] Fallback: {ticker} - Existing shares: {existing_shares}, Existing float: {existing_float}")
        
        # Try to get missing data using fallback methods
        fallback_shares = None
        fallback_float = None
        
        # If we need shares, try XBRL fallback
        if not existing_shares:
            logging.info(f"[SEC] Trying XBRL fallback for outstanding shares: {ticker}")
            fallback_shares = sec_api.get_shares_from_xbrl_fallback(ticker, date)
            if fallback_shares:
                logging.info(f"[SEC] XBRL fallback successful: {fallback_shares:,} shares")
        
        # If we need float, try Extractor API fallback
        if not existing_float:
            logging.info(f"[SEC] Trying Extractor API fallback for float: {ticker}")
            fallback_float = sec_api.get_float_from_extractor(ticker, date)
            if fallback_float:
                logging.info(f"[SEC] Extractor API fallback successful: {fallback_float:,} float shares")
        
        # If we have shares but no float, try DEF 14A estimation
        if (existing_shares or fallback_shares) and not existing_float and not fallback_float:
            total_shares = existing_shares or fallback_shares
            logging.info(f"[SEC] Trying DEF 14A fallback for float estimation: {ticker}")
            estimated_float = sec_api._estimate_float_from_def14a(ticker, date, total_shares)
            if estimated_float:
                fallback_float = estimated_float
                logging.info(f"[SEC] DEF 14A estimation successful: {fallback_float:,} estimated float shares")
        
        # Return data if we found anything new or have existing data
        if fallback_shares or fallback_float or existing_shares or existing_float:
            return {
                "shares_outstanding": fallback_shares or existing_shares,  # Preserve existing if fallback is None
                "float_shares": fallback_float or existing_float,  # Preserve existing if fallback is None
                "timing": "current"
            }
        
        return None
        
    except Exception as e:
        logging.error(f"[SEC] Error in get_sec_fallback_data for {ticker}: {e}")
        return None

class SECAdvanced:
    """Advanced SEC API integration for float and shares data using sec-api.io."""
    
    def __init__(self):
        self.api_key = SEC_API_KEY
        if not self.api_key:
            raise ValueError("SEC_API_KEY not found in environment variables")
        
        self.base_url = "https://api.sec-api.io"
        self.ticker = None  # Will be set when processing a specific ticker
        logging.info("SEC API initialized successfully")
    
    def get_float_and_shares_data(self, ticker, date=None):
        """
        Get comprehensive float and shares outstanding data using SEC API.
        
        Args:
            ticker (str): Stock ticker symbol
            date (str): Date in YYYY-MM-DD format (optional)
        
        Returns:
            dict: Comprehensive float and shares data with timing info
        """
        try:
            # Set ticker for historical price lookups
            self.ticker = ticker
            
            # Get float data from SEC API with timing info
            float_data, timing = self._get_float_data(ticker, date)
            
            if float_data:
                # Sum all share classes for outstanding shares
                shares_outstanding = self._sum_outstanding_shares(float_data)
                
                # Get public float (if available)
                float_shares = self._get_public_float(float_data)
                
                result = {
                    "ticker": ticker,
                    "shares_outstanding": shares_outstanding,
                    "float_shares": float_shares,
                    "date": date,
                    "period_of_report": float_data.get("periodOfReport"),
                    "reported_at": float_data.get("reportedAt"),
                    "timing": timing  # "before", "after", or "current"
                }
                
                return result
            else:
                return None
            
        except Exception as e:
            logging.error(f"Error in get_float_and_shares_data for {ticker}: {e}")
            return None
    
    def _get_cik(self, ticker):
        """Get CIK for a ticker, handling ticker name changes."""
        try:
            # Known ticker name changes - add more as discovered
            ticker_cik_map = {
                "SHOT": "1760903",  # JUPW/SHOT (without leading zeros)
                # Add more ticker changes here as discovered
            }
            
            if ticker in ticker_cik_map:
                logging.info(f"[SEC] Using known CIK for {ticker}: {ticker_cik_map[ticker]}")
                return ticker_cik_map[ticker]
            
            # Try to get CIK from SEC API
            query_str = f"ticker:{ticker}"
            payload = {
                "query": query_str, 
                "from": "0", 
                "size": "1", 
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            headers = {"Authorization": self.api_key, "Content-Type": "application/json"}
            
            response = requests.post(self.base_url, json=payload, headers=headers, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            filings = data.get("filings", [])
            if filings:
                cik = filings[0].get("cik")
                if cik:
                    logging.info(f"[SEC] Found CIK for {ticker}: {cik}")
                    return cik
            
            logging.warning(f"[SEC] Could not find CIK for {ticker}")
            return None
            
        except Exception as e:
            logging.error(f"[SEC] Error getting CIK for {ticker}: {e}")
            return None

    def _get_float_data(self, ticker, date=None):
        """Get float data from SEC API using CIK when possible."""
        try:
            # Try to get CIK for more reliable data retrieval
            cik = self._get_cik(ticker)
            
            if cik:
                url = f"{self.base_url}/float?cik={cik}"
                logging.info(f"[SEC] Calling float API for {ticker} using CIK: {cik}")
            else:
                url = f"{self.base_url}/float?ticker={ticker}"
                logging.info(f"[SEC] Calling float API for {ticker} using ticker symbol")
            
            headers = {"Authorization": self.api_key}
            
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            if not data or "data" not in data or not data["data"]:
                logging.warning(f"[SEC] No float data found for {ticker}")
                return None, None  # Return tuple (data, timing)
            
            float_records = data["data"]
            
            if date:
                target_date = datetime.strptime(date, "%Y-%m-%d")
                
                # First try to get data before or on the target date
                filtered_records = [
                    record for record in float_records 
                    if datetime.strptime(record.get("periodOfReport", "1900-01-01"), "%Y-%m-%d") <= target_date
                ]
                
                if filtered_records:
                    # Sort by periodOfReport and get the most recent
                    filtered_records.sort(key=lambda x: x.get("periodOfReport", ""), reverse=True)
                    best_before_record = filtered_records[0]
                    
                    # Check if the "before" record has float data
                    if self._has_float_data(best_before_record):
                        logging.info(f"[SEC] Using BEFORE data for {ticker}: {best_before_record.get('periodOfReport')} (has float data)")
                        return best_before_record, "before"
                    else:
                        logging.info(f"[SEC] BEFORE data for {ticker} has no float data, checking AFTER data")
                
                # No "before" data with float, or no "before" data at all - find closest AFTER
                after_records = [
                    record for record in float_records 
                    if datetime.strptime(record.get("periodOfReport", "1900-01-01"), "%Y-%m-%d") > target_date
                ]
                
                if after_records:
                    # Sort by date and get the closest (earliest after target)
                    after_records.sort(key=lambda x: x.get("periodOfReport", ""))
                    best_after_record = after_records[0]
                    
                    # Check if the "after" record has float data
                    if self._has_float_data(best_after_record):
                        logging.info(f"[SEC] Using AFTER data for {ticker}: {best_after_record.get('periodOfReport')} (has float data)")
                        return best_after_record, "after"
                    else:
                        logging.info(f"[SEC] AFTER data for {ticker} also has no float data")
                
                # If we get here, we found records but none have float data
                # Return the best "before" record if available, otherwise best "after"
                if filtered_records:
                    logging.info(f"[SEC] Returning BEFORE data for {ticker} (no float data available)")
                    return filtered_records[0], "before"
                elif after_records:
                    logging.info(f"[SEC] Returning AFTER data for {ticker} (no float data available)")
                    return after_records[0], "after"
                else:
                    logging.info(f"[SEC] No SEC filing data found for {ticker}")
                    return None, None
            else:
                # Get the most recent data
                float_records.sort(key=lambda x: x.get("periodOfReport", ""), reverse=True)
                return (float_records[0], "current") if float_records else (None, None)
            
        except Exception as e:
            logging.error(f"Error getting float data for {ticker}: {e}")
            return None, None
    
    def _has_float_data(self, record):
        """Check if a record has float data available."""
        try:
            # Check if publicFloat array has data
            public_float = record.get("publicFloat", [])
            if public_float and len(public_float) > 0:
                # Check if any publicFloat entry has a non-zero value
                for float_entry in public_float:
                    if float_entry.get("value") and float(float_entry.get("value", 0)) > 0:
                        return True
            return False
        except Exception as e:
            logging.error(f"Error checking float data availability: {e}")
            return False
    
    def _get_historical_price(self, period_date):
        """Get historical closing price for a specific date with multiple fallbacks."""
        try:
            if not period_date:
                return None
            
            # Method 1: Try Polygon API first
            polygon_price = self._get_polygon_historical_price(period_date)
            if polygon_price:
                # Sanity check for obviously wrong prices
                if polygon_price > 1000:
                    logging.warning(f"Suspiciously high Polygon price for {self.ticker} on {period_date}: ${polygon_price:.2f}")
                    polygon_price = None
                else:
                    return polygon_price
            
            # Method 2: Try yfinance as fallback
            yf_price = self._get_yfinance_historical_price(period_date)
            if yf_price:
                # Sanity check for obviously wrong prices
                if yf_price > 1000:
                    logging.warning(f"Suspiciously high yfinance price for {self.ticker} on {period_date}: ${yf_price:.2f}")
                    yf_price = None
                else:
                    logging.info(f"Using yfinance historical price for {self.ticker}: ${yf_price:.2f}")
                    return yf_price
            
            # Method 3: Try Finnhub as second fallback
            finnhub_price = self._get_finnhub_historical_price(period_date)
            if finnhub_price:
                # Sanity check for obviously wrong prices
                if finnhub_price > 1000:
                    logging.warning(f"Suspiciously high Finnhub price for {self.ticker} on {period_date}: ${finnhub_price:.2f}")
                    finnhub_price = None
                else:
                    logging.info(f"Using Finnhub historical price for {self.ticker}: ${finnhub_price:.2f}")
                    return finnhub_price
            
            logging.warning(f"No valid historical price found from any source for {self.ticker} on {period_date}")
            return None
                
        except Exception as e:
            logging.error(f"Error in _get_historical_price: {e}")
            return None
    
    def _get_polygon_historical_price(self, period_date):
        """Get historical closing price from Polygon API."""
        try:
            if not period_date or not self.ticker:
                return None
            
            # Use the global Polygon client and limiter with thread safety
            client = get_polygon_client()
            polygon_limiter.acquire_with_lock()
            
            # Get daily aggregates for the period date
            aggs = client.get_aggs(self.ticker, multiplier=1, timespan="day", from_=period_date, to=period_date)
            
            if aggs and len(aggs) > 0:
                close_price = aggs[0].close
                if close_price and close_price > 0:
                    logging.debug(f"Polygon historical price for {self.ticker} on {period_date}: ${close_price:.2f}")
                    return close_price
            
            return None
            
        except Exception as e:
            logging.debug(f"Error getting Polygon historical price for {self.ticker} on {period_date}: {e}")
            return None
    
    def _get_yfinance_historical_price(self, period_date):
        """Get historical closing price from YFinance."""
        try:
            if not period_date or not self.ticker:
                return None
            
            import yfinance as yf
            
            # Get historical data from YFinance
            yf_ticker = yf.Ticker(self.ticker)
            hist = yf_ticker.history(start=period_date, end=period_date)
            
            if not hist.empty:
                close_price = hist['Close'].iloc[0]
                if close_price and close_price > 0:
                    logging.debug(f"YFinance historical price for {self.ticker} on {period_date}: ${close_price:.2f}")
                    return close_price
            
            return None
            
        except Exception as e:
            logging.debug(f"Error getting YFinance historical price for {self.ticker} on {period_date}: {e}")
            return None
    
    def _get_finnhub_historical_price(self, period_date):
        """Get historical closing price from Finnhub (placeholder - not implemented)."""
        try:
            if not period_date or not self.ticker:
                return None
            
            # Finnhub implementation would go here if API key was available
            # For now, return None to skip this fallback
            logging.debug(f"Finnhub historical price not implemented for {self.ticker} on {period_date}")
            return None
            
        except Exception as e:
            logging.debug(f"Error getting Finnhub historical price for {self.ticker} on {period_date}: {e}")
            return None

    def query_filings(self, ticker, date, form_types=["10-K", "10-Q", "DEF 14A"]):
        """
        Query SEC filings using the Query API with "before then closest after" logic.
        
        Args:
            ticker (str): Stock ticker symbol
            date (str): Target date in YYYY-MM-DD format
            form_types (list): List of form types to search for
            
        Returns:
            dict: Most recent filing data or None if not found
        """
        try:
            if not self.api_key:
                logging.warning("[SEC] API key not available for filing query")
                return None
                
            # Build query string with Lucene syntax
            form_type_query = " OR ".join([f'formType:"{form_type}"' for form_type in form_types])
            
            # FIRST: Try to find filings BEFORE or ON the specified date
            before_query = f'ticker:{ticker} AND ({form_type_query}) AND filedAt:[1993-01-01 TO {date}]'
            
            payload = {
                "query": before_query,
                "from": "0",
                "size": "50",
                "sort": [{"filedAt": {"order": "desc"}}]
            }
            
            headers = {
                "Authorization": self.api_key,
                "Content-Type": "application/json"
            }
            
            logging.info(f"[SEC] Querying filings for {ticker} before {date}")
            response = requests.post("https://api.sec-api.io", json=payload, headers=headers, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            filings = data.get("filings", [])
            total = data.get("total", 0)
            
            if filings:
                logging.info(f"[SEC] Found {total} filings BEFORE {date} for {ticker}, returning most recent")
                latest_filing = filings[0]
                logging.info(f"[SEC] Latest filing: {latest_filing.get('formType')} - {latest_filing.get('periodOfReport')}")
                return latest_filing
            
            # SECOND: If no filings before, look for the CLOSEST filing AFTER the date
            logging.info(f"[SEC] No filings found BEFORE {date} for {ticker}, looking for closest AFTER")
            
            after_query = f'ticker:{ticker} AND ({form_type_query}) AND filedAt:[{date} TO 2030-12-31]'
            
            payload = {
                "query": after_query,
                "from": "0",
                "size": "50",
                "sort": [{"filedAt": {"order": "asc"}}]  # Sort ascending to get earliest after
            }
            
            response = requests.post("https://api.sec-api.io", json=payload, headers=headers, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            filings = data.get("filings", [])
            total = data.get("total", 0)
            
            if filings:
                logging.info(f"[SEC] Found {total} filings AFTER {date} for {ticker}, returning closest")
                closest_filing = filings[0]  # Earliest after the date
                logging.info(f"[SEC] Closest filing: {closest_filing.get('formType')} - {closest_filing.get('periodOfReport')}")
                return closest_filing
            
            logging.warning(f"[SEC] No filings found for {ticker} before or after {date}")
            return None
            
        except requests.exceptions.RequestException as e:
            logging.error(f"[SEC] Network error querying filings for {ticker}: {e}")
            return None
        except Exception as e:
            logging.error(f"[SEC] Error querying filings for {ticker}: {e}")
            return None

    def extract_section(self, filing_url, section="coverPage", response_type="text"):
        """
        Extract specific sections from SEC filings using the Extractor API.
        
        Args:
            filing_url (str): URL of the filing to extract from
            section (str): Section to extract (e.g., "coverPage", "5")
            response_type (str): Response type ("text" or "html")
            
        Returns:
            str: Extracted text content or None if failed
        """
        try:
            if not self.api_key:
                logging.warning("[SEC] API key not available for section extraction")
                return None
                
            # Use the correct endpoint and parameters according to documentation
            extractor_url = "https://api.sec-api.io/extractor"
            json_data = {
                "url": filing_url,
                "section": section,  # Confirmed working parameter name
                "type": response_type
            }
            
            headers = {"Authorization": self.api_key, "Content-Type": "application/json"}
            
            logging.info(f"[SEC] Extracting {section} section from filing")
            response = requests.post(extractor_url, json=json_data, headers=headers, timeout=15)
            response.raise_for_status()
            
            text = response.text
            logging.info(f"[SEC] Successfully extracted {len(text)} characters from {section}")
            
            return text
            
        except requests.exceptions.RequestException as e:
            logging.error(f"[SEC] Network error extracting section: {e}")
            return None
        except Exception as e:
            logging.error(f"[SEC] Error extracting section: {e}")
            return None

    def parse_float_from_text(self, text):
        """
        Parse float data from extracted text using multiple regex patterns.
        
        Args:
            text (str): Text content to parse
            
        Returns:
            int: Float dollar value or None if not found
        """
        try:
            if not text:
                return None
                
            # Multiple regex patterns to catch different ways float is reported
            patterns = [
                r"aggregate market value.*?\$([\d,]+)",
                r"market value.*?held by non-affiliates.*?\$([\d,]+)",
                r"public float.*?\$([\d,]+)",
                r"float.*?\$([\d,]+)",
                r"held by non-affiliates.*?\$([\d,]+)",
                r"non-affiliates.*?\$([\d,]+)"
            ]
            
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                if match:
                    dollar_value_str = match.group(1).replace(",", "")
                    try:
                        dollar_value = int(dollar_value_str)
                        logging.info(f"[SEC] Found float value: ${dollar_value:,} using pattern: {pattern}")
                        return dollar_value
                    except ValueError:
                        continue
                        
            logging.warning("[SEC] No float value found in extracted text")
            return None
            
        except Exception as e:
            logging.error(f"[SEC] Error parsing float from text: {e}")
            return None

    def get_float_from_extractor(self, ticker, date):
        """
        Get float data using the Extractor API as a fallback method.
        
        Args:
            ticker (str): Stock ticker symbol
            date (str): Target date in YYYY-MM-DD format
            
        Returns:
            int: Float shares or None if not found
        """
        try:
            # Query for relevant filings - prioritize supported filing types
            # Extractor API supports: 10-K, 10-Q, 8-K, S-1, etc.
            # Does NOT support: DEF 14A, 13F, etc.
            supported_form_types = ["10-K", "10-Q", "8-K", "S-1", "S-3", "424B2", "424B3", "424B4"]
            
            filing = self.query_filings(ticker, date, supported_form_types)
            if not filing:
                logging.info(f"[SEC] No supported filing types found for {ticker}, trying all types")
                # Fallback to all types if no supported types found
                filing = self.query_filings(ticker, date)
                if not filing:
                    return None
                
            filing_url = filing.get("linkToFilingDetails")
            period_date = filing.get("periodOfReport")
            form_type = filing.get("formType")
            
            if not filing_url or not period_date:
                logging.warning(f"[SEC] Missing filing URL or period date for {ticker}")
                return None
                
            logging.info(f"[SEC] Processing {form_type} filing for {ticker} with period {period_date}")
            
            # Check if this filing type is supported by Extractor API
            if form_type not in supported_form_types:
                logging.info(f"[SEC] {form_type} not supported by Extractor API, skipping extraction")
                return None
            
            # Try cover page first (most common location for float data)
            text = self.extract_section(filing_url, "coverPage")
            if text:
                dollar_value = self.parse_float_from_text(text)
                if dollar_value:
                    return self._convert_dollar_to_shares(dollar_value, period_date)
            
            # Try Item 5 if cover page didn't work
            text = self.extract_section(filing_url, "5")
            if text:
                dollar_value = self.parse_float_from_text(text)
                if dollar_value:
                    return self._convert_dollar_to_shares(dollar_value, period_date)
            
            # Try Item 1 if Item 5 didn't work
            text = self.extract_section(filing_url, "1")
            if text:
                dollar_value = self.parse_float_from_text(text)
                if dollar_value:
                    return self._convert_dollar_to_shares(dollar_value, period_date)
            
            logging.warning(f"[SEC] No float data found in any section for {ticker}")
            return None
            
        except Exception as e:
            logging.error(f"[SEC] Error in get_float_from_extractor for {ticker}: {e}")
            return None

    def _convert_dollar_to_shares(self, dollar_value, period_date):
        """
        Convert dollar value to share count using historical price.
        
        Args:
            dollar_value (int): Dollar value of float
            period_date (str): Period date for price lookup
            
        Returns:
            int: Number of shares or None if conversion failed
        """
        try:
            historical_price = self._get_historical_price(period_date)
            if historical_price and historical_price > 0:
                float_shares = int(dollar_value / historical_price)
                
                # Sanity check: if float shares is unreasonably small, the price might be wrong
                if float_shares < 1000 and dollar_value > 1000000:  # Less than 1K shares for $1M+ float
                    logging.warning(f"[SEC] Suspiciously small float calculation for {self.ticker}: ${dollar_value:,} / ${historical_price:.2f} = {float_shares:,} shares")
                    logging.warning(f"[SEC] This suggests the historical price of ${historical_price:.2f} might be incorrect")
                    return None
                
                logging.info(f"[SEC] Converted float: ${dollar_value:,} / ${historical_price:.2f} = {float_shares:,} shares")
                return float_shares
            else:
                logging.warning(f"[SEC] Could not get historical price for {period_date} to convert float")
                return None
        except Exception as e:
            logging.error(f"[SEC] Error converting dollar to shares: {e}")
            return None

    def get_shares_from_xbrl(self, accession_number):
        """
        Get outstanding shares from XBRL data using the XBRL-to-JSON API.
        
        Args:
            accession_number (str): Filing accession number
            
        Returns:
            int: Outstanding shares or None if not found
        """
        try:
            if not self.api_key:
                logging.warning("[SEC] API key not available for XBRL extraction")
                return None
                
            # Use the correct endpoint and parameter name according to documentation
            xbrl_url = "https://api.sec-api.io/xbrl-to-json"
            # Correct parameter format: accession-no (with hyphen)
            params = {"accession-no": accession_number}
            headers = {"Authorization": self.api_key}
            
            logging.info(f"[SEC] Fetching XBRL data for accession {accession_number}")
            response = requests.get(xbrl_url, params=params, headers=headers, timeout=15)
            response.raise_for_status()
            
            xbrl_data = response.json()
            
            # The XBRL API returns structured data, not raw Facts/us-gaap
            # Look for share information in various sections
            sections_to_check = [
                "BalanceSheets",
                "BalanceSheetsParenthetical", 
                "ShareholdersEquity",
                "ShareholdersEquityTables",
                "CoverPage"
            ]
            
            for section_name in sections_to_check:
                if section_name in xbrl_data:
                    section_data = xbrl_data[section_name]
                    logging.info(f"[SEC] Checking {section_name} section for share data")
                    
                    # Look for common share-related keys
                    share_patterns = [
                        "CommonStockSharesOutstanding",
                        "CommonStockSharesIssued", 
                        "SharesOutstanding",
                        "CommonStock",
                        "EntityCommonStockSharesOutstanding"
                    ]
                    
                    for pattern in share_patterns:
                        if isinstance(section_data, dict):
                            for key, value in section_data.items():
                                if pattern.lower() in key.lower():
                                    try:
                                        # Handle different value formats
                                        if isinstance(value, dict) and "value" in value:
                                            shares = int(float(value["value"]))
                                        elif isinstance(value, (int, float)):
                                            shares = int(float(value))
                                        elif isinstance(value, str) and value.replace(",", "").replace(".", "").isdigit():
                                            shares = int(float(value.replace(",", "")))
                                        else:
                                            continue
                                            
                                        if shares > 0:
                                            logging.info(f"[SEC] Found {shares:,} shares using XBRL field: {key}")
                                            return shares
                                    except (ValueError, TypeError):
                                        continue
            
            logging.warning(f"[SEC] No outstanding shares found in XBRL data for {accession_number}")
            return None
            
        except requests.exceptions.RequestException as e:
            logging.error(f"[SEC] Network error fetching XBRL data: {e}")
            return None
        except Exception as e:
            logging.error(f"[SEC] Error fetching XBRL data: {e}")
            return None

    def get_shares_from_xbrl_fallback(self, ticker, date):
        """
        Get outstanding shares using XBRL data as a fallback method.
        
        Args:
            ticker (str): Stock ticker symbol
            date (str): Target date in YYYY-MM-DD format
            
        Returns:
            int: Outstanding shares or None if not found
        """
        try:
            # Query for relevant filings
            filing = self.query_filings(ticker, date)
            if not filing:
                return None
                
            accession_number = filing.get("accessionNo")
            if not accession_number:
                logging.warning(f"[SEC] No accession number found for {ticker}")
                return None
                
            return self.get_shares_from_xbrl(accession_number)
            
        except Exception as e:
            logging.error(f"[SEC] Error in get_shares_from_xbrl_fallback for {ticker}: {e}")
            return None
    def _estimate_float_from_def14a(self, ticker, date, total_shares):
        """
        Estimate float from DEF 14A proxy statements by finding insider ownership.
        
        Args:
            ticker (str): Stock ticker symbol
            date (str): Target date in YYYY-MM-DD format
            total_shares (int): Total outstanding shares
            
        Returns:
            int: Estimated float shares or None if not found
        """
        try:
            # Query specifically for DEF 14A filings
            filing = self.query_filings(ticker, date, form_types=["DEF 14A"])
            if not filing:
                return None
                
            filing_url = filing.get("linkToFilingDetails")
            if not filing_url:
                return None
                
            # NOTE: DEF 14A filings are NOT supported by the Extractor API
            # This method will likely fail, but we'll try anyway for completeness
            logging.info(f"[SEC] Attempting DEF 14A extraction (likely to fail due to API limitations)")
            
            # Extract sections that typically contain ownership information
            sections_to_try = ["5", "1", "coverPage"]
            
            for section in sections_to_try:
                text = self.extract_section(filing_url, section)
                if text:
                    # Look for insider ownership patterns
                    insider_patterns = [
                        r"beneficially owned.*?(\d{1,3}(?:,\d{3})*)",
                        r"owned by.*?(\d{1,3}(?:,\d{3})*)",
                        r"insider.*?(\d{1,3}(?:,\d{3})*)",
                        r"executive.*?(\d{1,3}(?:,\d{3})*)"
                    ]
                    
                    for pattern in insider_patterns:
                        matches = re.findall(pattern, text, re.IGNORECASE)
                        if matches:
                            # Take the largest number found (likely total insider ownership)
                            try:
                                insider_shares = max([int(match.replace(",", "")) for match in matches])
                                if insider_shares < total_shares:
                                    float_shares = total_shares - insider_shares
                                    logging.info(f"[SEC] Estimated float: {total_shares:,} - {insider_shares:,} = {float_shares:,}")
                                    return float_shares
                            except (ValueError, TypeError):
                                continue
            
            logging.info(f"[SEC] DEF 14A extraction failed (expected due to API limitations)")
            return None
            
        except Exception as e:
            logging.error(f"[SEC] Error estimating float from DEF 14A for {ticker}: {e}")
            return None

    def _sum_outstanding_shares(self, float_data):
        """Sum all outstanding shares across all share classes."""
        try:
            if not float_data or "float" not in float_data or "outstandingShares" not in float_data["float"]:
                return None
            
            outstanding_shares = float_data["float"]["outstandingShares"]
            total_shares = 0
            
            for share_class in outstanding_shares:
                if "value" in share_class and share_class["value"] is not None:
                    total_shares += int(share_class["value"])
            
            return total_shares if total_shares > 0 else None
            
        except Exception as e:
            logging.error(f"Error summing outstanding shares: {e}")
            return None
    
    def _get_public_float(self, float_data):
        """Get public float value in shares by converting from dollar value."""
        try:
            if not float_data or "float" not in float_data:
                return None
            
            float_obj = float_data["float"]
            
            # Only get publicFloat - NO fallback to outstandingShares
            if "publicFloat" in float_obj:
                public_float = float_obj["publicFloat"]
                
                # Get the most recent public float
                if public_float and len(public_float) > 0:
                    # Sort by period and get the most recent
                    try:
                        public_float.sort(key=lambda x: x.get("period", ""), reverse=True)
                        latest_float = public_float[0]
                        
                        if "value" in latest_float and latest_float["value"] is not None:
                            # publicFloat value is in DOLLARS, not shares
                            public_float_dollars = float(latest_float["value"])
                            period_date = latest_float.get("period")
                            
                            # Get historical price for the period date
                            historical_price = self._get_historical_price(period_date)
                            
                            if historical_price and historical_price > 0:
                                # Calculate actual float shares: dollars / price per share
                                float_shares = int(public_float_dollars / historical_price)
                                
                                # Sanity check: if float shares is unreasonably small, the price might be wrong
                                if float_shares < 1000 and public_float_dollars > 1000000:  # Less than 1K shares for $1M+ float
                                    logging.warning(f"Suspiciously small float calculation for {self.ticker}: ${public_float_dollars:,.0f} / ${historical_price:.2f} = {float_shares:,} shares")
                                    logging.warning(f"This suggests the historical price of ${historical_price:.2f} might be incorrect")
                                    return None
                                
                                logging.info(f"Converted public float: ${public_float_dollars:,.0f} / ${historical_price:.2f} = {float_shares:,} shares")
                                return float_shares
                            else:
                                logging.warning(f"Could not get historical price for {period_date} to convert public float")
                                return None
                    except Exception as sort_error:
                        logging.warning(f"Error sorting public float data: {sort_error}")
                        # If sorting fails, just get the first item
                        if public_float and len(public_float) > 0:
                            latest_float = public_float[0]
                            if "value" in latest_float and latest_float["value"] is not None:
                                public_float_dollars = float(latest_float["value"])
                                period_date = latest_float.get("period")
                                historical_price = self._get_historical_price(period_date)
                                
                                if historical_price and historical_price > 0:
                                    float_shares = int(public_float_dollars / historical_price)
                                    
                                    # Sanity check: if float shares is unreasonably small, the price might be wrong
                                    if float_shares < 1000 and public_float_dollars > 1000000:  # Less than 1K shares for $1M+ float
                                        logging.warning(f"Suspiciously small float calculation for {self.ticker}: ${public_float_dollars:,.0f} / ${historical_price:.2f} = {float_shares:,} shares")
                                        logging.warning(f"This suggests the historical price of ${historical_price:.2f} might be incorrect")
                                        return None
                                    
                                    logging.info(f"Converted public float: ${public_float_dollars:,.0f} / ${historical_price:.2f} = {float_shares:,} shares")
                                    return float_shares
                                else:
                                    logging.warning(f"Could not get historical price for {period_date} to convert public float")
                                    return None
            
            # NO FALLBACK - just return None if no publicFloat data
            return None
            
        except Exception as e:
            logging.error(f"Error getting public float: {e}")
            return None



def safe_percentage_change(new_value, old_value, default="Not Available"):
    """Safely calculate percentage change with error handling."""
    try:
        if old_value is None or old_value == 0 or old_value == "Not Available":
            return default
        if isinstance(old_value, str) or isinstance(new_value, str):
            return default
        return round((new_value - old_value) / old_value * 100, 1)
    except (TypeError, ZeroDivisionError, ValueError):
        return default

def safe_percentage_ratio(numerator, denominator, default="Not Available"):
    """Safely calculate percentage ratio with error handling."""
    try:
        if denominator is None or denominator == 0 or denominator == "Not Available":
            return default
        if isinstance(denominator, str) or isinstance(numerator, str):
            return default
        return round((numerator / denominator) * 100, 2)
    except (TypeError, ZeroDivisionError, ValueError):
        return default

def is_trading_day(date_str):
    """Check if the date is a weekday and not a US public holiday using dynamic library.
    Note: Veterans Day and Columbus Day are treated as trading days since the market is open on those days."""
    try:
        date = datetime.strptime(date_str, "%Y-%m-%d")
        if date.weekday() >= 5:  # Weekend
            return False
        
        current_year = date.year
        us_holidays = holidays.US(years=range(current_year - 1, current_year + 2))  # Covers +/-1 year for flexibility
        if date in us_holidays:
            holiday_name = us_holidays[date]
            # Veterans Day and Columbus Day are trading days - market is open
            if "Veterans Day" in holiday_name or "Veterans" in holiday_name:
                logging.info(f"{date_str} is Veterans Day - market is open, treating as trading day")
                return True
            if "Columbus Day" in holiday_name or "Columbus" in holiday_name:
                logging.info(f"{date_str} is Columbus Day - market is open, treating as trading day")
                return True
            logging.info(f"{date_str} is a US holiday ({holiday_name}), skipping")
            return False
        
        return True
    except ValueError:
        logging.error(f"Invalid date format: {date_str}")
        return False

def calculate_vwap_reclaim(data, session_name, session_high, session_high_time):
    """Enhanced VWAP reclaim logic to handle starting below VWAP and include times for breakdowns, reclaims, and losses."""
    if data.empty or len(data) < 5 or data['volume'].sum() < 10000:
        return {"reclaim": "Insufficient data to determine VWAP behavior.", "vwap": None}

    try:
        data_copy = data.copy()
        
        # Validate session_high
        actual_session_high = data_copy['high'].max()
        if pd.isna(actual_session_high) or actual_session_high <= 0:
            return {"reclaim": "Invalid session high data.", "vwap": None}
        
        if pd.isna(session_high) or session_high <= 0:
            session_high = actual_session_high
            logging.warning(f"Using actual session high for {session_name}: {session_high}")
        
        data_copy['typical_price'] = (data_copy['high'] + data_copy['low'] + data_copy['close']) / 3
        data_copy['cum_vol'] = data_copy['volume'].cumsum()
        data_copy['cum_typical_vol'] = (data_copy['typical_price'] * data_copy['volume']).cumsum()
        data_copy['vwap'] = data_copy['cum_typical_vol'] / data_copy['cum_vol']
        
        if data_copy['cum_vol'].iloc[-1] == 0:
            return {"reclaim": "No volume data for VWAP calculation.", "vwap": None}
        
        data_copy['above_vwap'] = data_copy['close'] > data_copy['vwap']
        min_volume_threshold = 1000
        
        # Track all transitions
        transitions = []
        prev_above = None
        breakdown_time = None
        first_reclaim_time = None
        last_loss_time = None
        reclaim_high = 0
        reclaim_high_price = 0
        reclaim_high_time = None
        started_below = not data_copy['above_vwap'].iloc[0] if not data_copy.empty else False
        
        for idx, row in data_copy.iterrows():
            current_above = row['above_vwap']
            
            # Update reclaim high if currently above VWAP
            if current_above and row['high'] > reclaim_high:
                reclaim_high = row['high']
                reclaim_high_price = row['high']
                reclaim_high_time = idx
            
            if prev_above is not None and prev_above != current_above:
                current_vol = row['volume']
                prev_idx = data_copy.index[data_copy.index.get_loc(idx) - 1] if data_copy.index.get_loc(idx) > 0 else None
                prev_vol = data_copy.loc[prev_idx, 'volume'] if prev_idx else 0
                
                if current_vol >= min_volume_threshold and prev_vol >= min_volume_threshold:
                    if current_above:  # Reclaim (below to above)
                        transitions.append(('reclaim', idx))
                        if first_reclaim_time is None:
                            first_reclaim_time = idx
                    else:  # Loss (above to below)
                        transitions.append(('lost', idx))
                        last_loss_time = idx  # Update last loss time
                        if breakdown_time is None:
                            breakdown_time = idx  # First loss is the breakdown
            prev_above = current_above
        
        reclaims = [t for t in transitions if t[0] == 'reclaim']
        losses = [t for t in transitions if t[0] == 'lost']
        reclaim_count = len(reclaims)
        final_above = data_copy['above_vwap'].iloc[-1] if not data_copy.empty else False
        
        # Determine high type
        high_type = "lower high"
        if reclaim_count > 0 and reclaim_high > 0:
            if reclaim_high > session_high:
                high_type = "higher high"
            elif abs(reclaim_high - session_high) < 0.01:
                high_type = "equal high"
        
        # Format times
        breakdown_time_str = breakdown_time.strftime("%H:%M:%S") if breakdown_time else "Not Available"
        reclaim_time_str = first_reclaim_time.strftime("%H:%M:%S") if first_reclaim_time else "Not Available"
        loss_time_str = last_loss_time.strftime("%H:%M:%S") if last_loss_time else "Not Available"
        high_time_str = reclaim_high_time.strftime("%H:%M:%S") if reclaim_high_time else "Not Available"
        high_price_str = f"${reclaim_high_price:.2f}" if reclaim_high_price > 0 else "N/A"
        
        reclaim_text = "once" if reclaim_count == 1 else "multiple times"
        held_text = "and held it" if final_above else f"but lost it at {loss_time_str}"
        
        # Get the final VWAP value for deviation calculation
        final_vwap = data_copy['vwap'].iloc[-1] if not data_copy.empty else None
        
        if reclaim_count == 0:
            if final_above:
                if started_below:
                    # Implicit reclaim without transitions (e.g., gradual cross)
                    vwap_behavior = f"Started below VWAP but reclaimed implicitly and held it."
                else:
                    vwap_behavior = f"Always above VWAP with no breakdown."
                return {"reclaim": vwap_behavior, "vwap": final_vwap, "breakdown_time": "Not Available"}
            else:
                return {"reclaim": f"Always below VWAP with no breakdown.", "vwap": final_vwap, "breakdown_time": "Not Available"}
        
        if started_below:
            vwap_behavior = (f"Started below VWAP but reclaimed {reclaim_text} "
                             f"with a {high_type} at {reclaim_time_str}, "
                             f"high of {high_price_str} at {high_time_str} {held_text}.")
            # For started below scenarios, there was no breakdown, so return "Not Available"
            return {"reclaim": vwap_behavior, "vwap": final_vwap, "breakdown_time": "Not Available"}
        else:
            vwap_behavior = (f"VWAP breakdown at {breakdown_time_str}, reclaimed {reclaim_text} "
                             f"with a {high_type} at {reclaim_time_str}, "
                             f"high of {high_price_str} at {high_time_str} {held_text}.")
            # For breakdown scenarios, return the actual breakdown time
            return {"reclaim": vwap_behavior, "vwap": final_vwap, "breakdown_time": breakdown_time_str}
        
    except Exception as e:
        logging.warning(f"Error calculating VWAP reclaim for {session_name}: {str(e)}")
        return {"reclaim": "Insufficient data to determine VWAP behavior.", "vwap": None, "breakdown_time": "Not Available"}

def calculate_vwap_deviation_percentage(price, vwap):
    if vwap == 0:
        return 0
    return ((price - vwap) / vwap) * 100

def fetch_news_playwright(ticker, session, target_date):
    """Fetch news from Yahoo Finance using Playwright."""
    try:
        url = f"https://finance.yahoo.com/quote/{ticker}/news/"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_extra_http_headers(headers)
            
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=8000)
                page.wait_for_timeout(2000)
                html_content = page.content()
                soup = BeautifulSoup(html_content, 'html.parser')
                logging.info(f"Yahoo Finance response (Playwright) for {ticker}")
            except Exception as e:
                logging.warning(f"Playwright navigation failed for {ticker}: {e}")
                try:
                    page.goto(url, timeout=5000)
                    page.wait_for_timeout(1000)
                    html_content = page.content()
                    soup = BeautifulSoup(html_content, 'html.parser')
                    logging.info(f"Yahoo Finance response (Playwright fallback) for {ticker}")
                except Exception as e2:
                    logging.warning(f"Playwright fallback also failed for {ticker}: {e2}")
                    soup = None
            finally:
                browser.close()
        
        if soup:
            return parse_yahoo_news(soup, ticker, target_date)
        return []
    except Exception as e:
        logging.warning(f"Playwright failed for {ticker}: {e}")
        return []

def fetch_news_finviz(ticker, session, target_date):
    """Fetch news from Finviz."""
    try:
        finviz_url = f"https://finviz.com/quote.ashx?t={ticker}"
        finviz_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }
        
        response = session.get(finviz_url, headers=finviz_headers, timeout=8)
        response.raise_for_status()
        finviz_soup = BeautifulSoup(response.text, 'html.parser')
        
        finviz_news = []
        news_rows = finviz_soup.find_all('tr', class_='cursor-pointer')
        
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").replace(tzinfo=EASTERN_TZ)
        start_dt = target_dt - timedelta(days=2)
        end_dt = target_dt
        
        for row in news_rows:
            try:
                news_link = row.find('a', class_='tab-link-news')
                if news_link:
                    title = news_link.text.strip()
                    if title and len(title) > 10:
                        date_cell = row.find('td', class_='news-link-container')
                        if date_cell:
                            date_text = date_cell.get_text(strip=True)
                            if 'today' in date_text.lower():
                                pub_date = target_dt
                            elif 'yesterday' in date_text.lower():
                                pub_date = target_dt - timedelta(days=1)
                            else:
                                time_match = re.search(r'(\d+)\s+(hour|day|minute)s?\s+ago', date_text.lower())
                                if time_match:
                                    num, unit = time_match.groups()
                                    if unit == 'hour':
                                        pub_date = target_dt - timedelta(hours=int(num))
                                    elif unit == 'day':
                                        pub_date = target_dt - timedelta(days=int(num))
                                    elif unit == 'minute':
                                        pub_date = target_dt - timedelta(minutes=int(num))
                                else:
                                    pub_date = target_dt
                            
                            if start_dt.date() <= pub_date.date() <= end_dt.date():
                                finviz_news.append(f"{pub_date.strftime('%Y-%m-%d')}: {title}")
            except Exception as e:
                logging.warning(f"Error processing Finviz news row for {ticker}: {e}")
                continue
        
        if finviz_news:
            logging.info(f"Added {len(finviz_news)} Finviz news items for {ticker}")
        return finviz_news
    except Exception as e:
        logging.warning(f"Finviz scraping failed for {ticker}: {e}")
        return []

def fetch_news_yahoo_traditional(ticker, session, target_date):
    """Fetch news from Yahoo Finance using traditional requests."""
    try:
        url = f"https://finance.yahoo.com/quote/{ticker}/news/"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
        
        for attempt in range(3):
            try:
                response = session.get(url, headers=headers, timeout=10)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                logging.info(f"Yahoo Finance response (traditional) for {ticker}")
                return parse_yahoo_news(soup, ticker, target_date)
            except HTTPError as http_e:
                if "429" in str(http_e):
                    logging.warning(f"Yahoo Finance rate limit hit for {ticker}")
                else:
                    logging.warning(f"Yahoo Finance HTTP error for {ticker}: {http_e}")
                if attempt == 2:
                    break
            except Exception as e:
                logging.warning(f"Error fetching Yahoo news for {ticker}: {e}")
                if attempt == 2:
                    break
            time.sleep(0.5)
        return []
    except Exception as e:
        logging.warning(f"Yahoo traditional failed for {ticker}: {e}")
        return []

def fetch_news_finnhub(ticker, session, target_date):
    """Fetch news from Finnhub API."""
    try:
        finnhub_limiter.acquire()
        quote = finnhub_client.quote(ticker)
        if not isinstance(quote, dict) or quote.get('c', 0) == 0:
            logging.warning(f"Invalid or delisted ticker {ticker} or no quote data")
            return []
    except Exception as e:
        logging.warning(f"Finnhub validation error for {ticker}: {e}")
        return []

    try:
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").replace(tzinfo=EASTERN_TZ)
        start_dt = target_dt - timedelta(days=2)
        end_dt = target_dt
        
        finnhub_limiter.acquire()
        news = finnhub_client.company_news(ticker, _from=start_dt.strftime("%Y-%m-%d"), to=end_dt.strftime("%Y-%m-%d"))
        
        news_items = []
        for item in news:
            try:
                pub_date = datetime.fromtimestamp(item["datetime"], tz=pytz.UTC).astimezone(EASTERN_TZ)
                if start_dt.date() <= pub_date.date() <= end_dt.date():
                    headline = item.get("headline", "No headline")
                    news_items.append(f"{pub_date.strftime('%Y-%m-%d')}: {headline}")
            except (ValueError, TypeError) as e:
                logging.warning(f"Invalid timestamp in Finnhub news for {ticker}: {item.get('datetime')}")
                continue
        return news_items
    except Exception as e:
        logging.error(f"Error fetching Finnhub news for {ticker}: {e}")
        return []

def fetch_news_polygon(ticker, session, target_date):
    """Fetch news from Polygon.io API using the correct API parameters."""
    try:
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").replace(tzinfo=EASTERN_TZ)
        start_dt = target_dt - timedelta(days=2)
        end_dt = target_dt
        
        client = get_polygon_client()
        polygon_limiter.acquire()
        
        # Use correct API parameters based on documentation
        news = client.list_ticker_news(
            ticker=ticker,
            published_utc_gte=start_dt.strftime("%Y-%m-%d"),
            published_utc_lte=end_dt.strftime("%Y-%m-%d"),
            order="desc",
            limit=10,
            sort="published_utc"
        )
        
        news_items = []
        for item in news:
            try:
                # Verify this is a TickerNews object as per sample code
                if isinstance(item, TickerNews):
                    pub_date = datetime.strptime(item.published_utc, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=pytz.UTC).astimezone(EASTERN_TZ)
                    if start_dt.date() <= pub_date.date() <= end_dt.date():
                        news_items.append(f"{pub_date.strftime('%Y-%m-%d')}: {item.title}")
            except Exception as e:
                logging.warning(f"Error processing Polygon.io news for {ticker}: {e}")
        return news_items
    except Exception as e:
        logging.error(f"Error fetching Polygon.io news for {ticker}: {e}")
        return []

def parse_yahoo_news(soup, ticker, target_date):
    """Parse Yahoo Finance news with optimized timestamp parsing."""
    news_items = []
    
    try:
        # Find all article containers - look for actual news articles
        articles = soup.find_all('section', {'data-testid': 'storyitem'})
        
        if not articles:
            # Fallback: look for other news containers
            articles = soup.find_all(['article', 'div'], class_=re.compile(r'story|item'))
        
        if not articles:
            # Another fallback
            articles = soup.find_all('li', class_=re.compile(r'story'))
        
        logging.info(f"[YAHOO NEWS] Found {len(articles)} potential articles for {ticker}")
        
        # Date range for filtering - target date and 2 days prior
        target_dt = datetime.strptime(target_date, "%Y-%m-%d").replace(tzinfo=EASTERN_TZ)
        start_dt = target_dt - timedelta(days=2)  # 2 days before target date
        end_dt = target_dt  # Up to target date (not after)
        logging.info(f"Looking for news from {start_dt.strftime('%Y-%m-%d')} to {end_dt.strftime('%Y-%m-%d')} (target date and 2 days prior)")
        
        for article in articles:
            try:
                # Extract title
                title_elem = article.find(['h1', 'h2', 'h3', 'h4', 'a'], class_=re.compile(r'title|headline|subject'))
                if not title_elem:
                    title_elem = article.find(['h1', 'h2', 'h3', 'h4'])
                
                if not title_elem:
                    continue
                    
                title = title_elem.get_text(strip=True)
                if not title or len(title) < 10:
                    continue
            
                # Find timestamp element with optimized selectors
                time_elem = None
                
                # OPTIMIZED SELECTORS - Ordered by efficiency and Yahoo Finance prevalence
                time_selectors = [
                    # YAHOO FINANCE SPECIFIC (current structure) - Most common patterns first
                    ('div', 'publishing'),  # Most common Yahoo pattern - publishing info
                    ('div', lambda x: x and hasattr(x, 'get') and x.get('class') and 'publishing' in str(x.get('class', []))),
                    ('div', 'footer'),  # Footer often contains timestamp
                    ('div', lambda x: x and hasattr(x, 'get') and x.get('class') and 'footer' in str(x.get('class', []))),
                    
                    # YAHOO FINANCE LEGACY PATTERNS
                    ('span', 'caas-time'),  # Legacy Yahoo pattern
                    ('time', 'caas-time'),  # HTML5 time with caas-time class
                    ('*', lambda x: x and hasattr(x, 'get') and x.get('class') and 'caas-time' in str(x.get('class', []))),
                    
                    # DATA ATTRIBUTES (modern approach)
                    ('*', lambda x: hasattr(x, 'get') and x.get('data-time')),
                    ('*', lambda x: hasattr(x, 'get') and x.get('data-timestamp')),
                    ('*', lambda x: hasattr(x, 'get') and x.get('data-published')),
                    ('*', lambda x: hasattr(x, 'get') and x.get('datetime')),
                    
                    # HTML5 TIME ELEMENTS (high reliability)
                    ('time', None),  # Any time element
                    ('time', lambda x: True),  # Any time element with any class
                    
                    # COMMON NEWS PATTERNS (keyword-based) - Medium efficiency
                    ('span', lambda x: x and hasattr(x, 'get') and x.get('class') and any(term in str(x.get('class', [])).lower() for term in ['timestamp', 'time', 'date', 'published', 'ago', 'byline', 'meta'])),
                    ('div', lambda x: x and hasattr(x, 'get') and x.get('class') and any(term in str(x.get('class', [])).lower() for term in ['timestamp', 'time', 'date', 'published', 'ago', 'byline', 'meta'])),
                    ('p', lambda x: x and hasattr(x, 'get') and x.get('class') and any(term in str(x.get('class', [])).lower() for term in ['timestamp', 'time', 'date', 'published', 'ago', 'byline', 'meta'])),
                    ('small', lambda x: x and hasattr(x, 'get') and x.get('class') and any(term in str(x.get('class', [])).lower() for term in ['timestamp', 'time', 'date', 'published', 'ago'])),
                    
                    # FALLBACK CLASS-BASED SELECTORS
                    ('span', 'time'),
                    ('div', 'time'),
                    ('span', 'date'),
                    ('div', 'date'),
                    
                    # REGEX WILDCARDS (last resort) - Only when needed
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'\d+\s+(hour|day|minute|second)s?\s+ago', x.text.lower())),
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'(today|yesterday|just now|\d{1,2}:\d{2})', x.text.lower())),
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{1,2}', x.text.lower())),
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', x.text)),
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'\d{4}-\d{2}-\d{2}', x.text)),
                    
                    # Last resort: look for any element containing time-like text
                    ('*', lambda x: x and hasattr(x, 'text') and re.search(r'\d{1,2}:\d{2}', x.text))
                ]
            
                # Try each selector until we find a timestamp
                for tag, selector in time_selectors:
                    try:
                        if callable(selector):
                            elements = article.find_all(tag)
                            for elem in elements:
                                if elem and selector(elem):  # Add None check
                                    time_elem = elem
                                    break
                        else:
                            time_elem = article.find(tag, class_=selector)
                        
                        if time_elem:
                            break
                    except Exception as e:
                        logging.debug(f"Selector failed: {tag}, {selector}: {e}")
                        continue
                
                # Extract timestamp text with enhanced extraction
                pub_time = None
                if time_elem:
                    # Try data attributes first (most reliable)
                    for attr in ['data-time', 'data-timestamp', 'data-published', 'datetime']:
                        if time_elem.get(attr):
                            pub_time = time_elem.get(attr)
                            break
                    
                    # Fall back to text content
                    if not pub_time:
                        pub_time = time_elem.text.strip()
                
                if not pub_time:
                    logging.debug(f"[YAHOO NEWS] No timestamp found for article: {title[:50]}...")
                    continue
                
                logging.debug(f"[YAHOO NEWS] Found timestamp '{pub_time}' for article: {title[:50]}...")
                
                # Parse timestamp using four-tier strategy
                pub_date = None
                try:
                    # Use current time as fetch context for relative timestamps
                    fetch_time = datetime.now(pytz.UTC).astimezone(EASTERN_TZ)
                    
                    # STEP 1: Try relative timestamp parsing first (for "1mo ago" format)
                    pub_date = try_relative_timestamp_parsing(pub_time, fetch_time)
                    
                    # STEP 2: Try direct parsing (fast path for absolute dates)
                    if not pub_date:
                        pub_date = try_direct_timestamp_parsing(pub_time)
                    
                    # STEP 3: Use dateutil.parser as fallback (robust)
                    if not pub_date:
                        pub_date = try_dateutil_timestamp_parsing(pub_time, fetch_time)
                    
                    # STEP 4: Fallback to original logic for edge cases
                    if not pub_date:
                        try:
                            pub_date = datetime.strptime(pub_time, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=pytz.UTC).astimezone(EASTERN_TZ)
                        except ValueError:
                            try:
                                pub_date = datetime.strptime(pub_time, "%Y-%m-%d %H:%M:%S").replace(tzinfo=EASTERN_TZ)
                            except ValueError:
                                try:
                                    pub_date = datetime.strptime(pub_time, "%m/%d/%Y %I:%M %p").replace(tzinfo=EASTERN_TZ)
                                except ValueError:
                                    try:
                                        pub_date = datetime.strptime(pub_time, "%B %d, %Y").replace(tzinfo=EASTERN_TZ)
                                    except ValueError:
                                        pass
                    
                    if not pub_date:
                        logging.debug(f"[YAHOO NEWS] Could not parse timestamp '{pub_time}' for article: {title[:50]}...")
                        continue
                    
                    logging.debug(f"[YAHOO NEWS] Parsed timestamp '{pub_time}' -> {pub_date.strftime('%Y-%m-%d')} for article: {title[:50]}...")
                    
                    # Check if article is within our date range
                    if start_dt.date() <= pub_date.date() <= end_dt.date():
                        # Clean title for logging to avoid Unicode issues
                        clean_title = title.encode('utf-8', errors='replace').decode('utf-8')
                        news_items.append(f"{pub_date.strftime('%Y-%m-%d')}: {title}")
                        logging.info(f"[YAHOO NEWS] Added news: {pub_date.strftime('%Y-%m-%d')}: {clean_title[:50]}...")
                    else:
                        clean_title = title.encode('utf-8', errors='replace').decode('utf-8')
                        logging.debug(f"[YAHOO NEWS] Article outside date range: {pub_date.strftime('%Y-%m-%d')} - {clean_title[:50]}...")
                        
                except Exception as e:
                    logging.debug(f"[YAHOO NEWS] Error parsing timestamp '{pub_time}': {e}")
                    continue
                    
            except Exception as e:
                logging.debug(f"[YAHOO NEWS] Error processing article: {e}")
                continue
        
        logging.info(f"[YAHOO NEWS] Successfully parsed {len(news_items)} news items for {ticker}")
        return news_items
        
    except Exception as e:
        logging.warning(f"Error parsing Yahoo news for {ticker}: {e}")
        return []

def try_relative_timestamp_parsing(raw_timestamp, fetch_time):
    """Parse relative timestamps like '1mo ago', '2mo ago', etc."""
    try:
        clean_timestamp = raw_timestamp.strip().lower()
        
        # Handle relative time formats
        if 'ago' in clean_timestamp:
            # Extract number and unit
            import re
            match = re.search(r'(\d+)\s*(mo|month|day|d|hour|h|minute|min|m|second|sec|s)\s*ago', clean_timestamp)
            if match:
                number = int(match.group(1))
                unit = match.group(2)
                
                # Calculate the actual date
                if unit in ['mo', 'month']:
                    # Approximate months as 30 days
                    actual_date = fetch_time - timedelta(days=number * 30)
                elif unit in ['day', 'd']:
                    actual_date = fetch_time - timedelta(days=number)
                elif unit in ['hour', 'h']:
                    actual_date = fetch_time - timedelta(hours=number)
                elif unit in ['minute', 'min', 'm']:
                    actual_date = fetch_time - timedelta(minutes=number)
                elif unit in ['second', 'sec', 's']:
                    actual_date = fetch_time - timedelta(seconds=number)
                else:
                    return None
                
                return actual_date
        
        return None
        
    except Exception as e:
        logging.debug(f"Relative timestamp parsing failed for '{raw_timestamp}': {e}")
        return None

def try_direct_timestamp_parsing(raw_timestamp):
    """Quick parsing for known formats (fast path)."""
    if not raw_timestamp:
        return None
    
    try:
        from datetime import datetime
        
        # Handle common formats directly - these are the most frequent patterns
        formats_to_try = [
            '%m/%d/%Y',           # 12/25/2023
            '%Y-%m-%d',           # 2023-12-25
            '%B %d, %Y',          # December 25, 2023
            '%b %d, %Y',          # Dec 25, 2023
            '%d %B %Y',           # 25 December 2023
            '%d %b %Y',           # 25 Dec 2023
            '%I:%M %p',           # 3:45 PM
            '%H:%M',              # 15:45
            '%m/%d/%y',           # 12/25/23
            '%Y-%m-%dT%H:%M:%S',  # 2023-12-25T15:45:00
            '%Y-%m-%dT%H:%M:%SZ', # 2023-12-25T15:45:00Z
        ]
        
        clean_timestamp = raw_timestamp.strip()
        
        for fmt in formats_to_try:
            try:
                parsed = datetime.strptime(clean_timestamp, fmt)
                # Handle timezone - if no timezone info, assume Eastern
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=EASTERN_TZ)
                else:
                    parsed = parsed.astimezone(EASTERN_TZ)
                return parsed
            except ValueError:
                continue
                
    except Exception as e:
        logging.debug(f"Direct parsing failed for '{raw_timestamp}': {e}")
    
    return None

def try_dateutil_timestamp_parsing(raw_timestamp, fetch_time):
    """Robust parsing with dateutil.parser (fallback)."""
    try:
        from dateutil import parser
        
        # Clean the timestamp
        clean_timestamp = raw_timestamp.strip()
        
        # Parse with fetch time as context for relative timestamps
        parsed = parser.parse(clean_timestamp, default=fetch_time, fuzzy=True)
        
        # Convert to Eastern Time
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=EASTERN_TZ)
        else:
            parsed = parsed.astimezone(EASTERN_TZ)
        
        return parsed
        
    except Exception as e:
        logging.debug(f"Dateutil parsing failed for '{raw_timestamp}': {e}")
        return None

def fetch_news(ticker, session, target_date):
    """Fetch news from all sources with Polygon as first priority and report source."""
    try:
        # Create a separate session for this thread's parallel operations to avoid thread safety issues
        with requests.Session() as thread_session:
            # Configure the thread session
            retries = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
            adapter = HTTPAdapter(
                pool_connections=10,
                pool_maxsize=10,
                max_retries=retries
            )
            thread_session.mount("https://", adapter)
            
            # Run all news sources in parallel with the thread-safe session (reduced workers for optimal performance)
            with ThreadPoolExecutor(max_workers=5) as executor:
                future_playwright = executor.submit(fetch_news_playwright, ticker, thread_session, target_date)
                future_finviz = executor.submit(fetch_news_finviz, ticker, thread_session, target_date)
                future_yahoo = executor.submit(fetch_news_yahoo_traditional, ticker, thread_session, target_date)
                future_finnhub = executor.submit(fetch_news_finnhub, ticker, thread_session, target_date)
                future_polygon = executor.submit(fetch_news_polygon, ticker, thread_session, target_date)

                # Gather results by source - MOVED INSIDE the with block
                polygon_news = future_polygon.result()
                playwright_news = future_playwright.result()
                finviz_news = future_finviz.result()
                yahoo_news = future_yahoo.result()
                finnhub_news = future_finnhub.result()

        # Combine with Polygon first
        all_news = []
        all_news.extend(polygon_news)
        all_news.extend(playwright_news)
        all_news.extend(finviz_news)
        all_news.extend(yahoo_news)
        all_news.extend(finnhub_news)

        # Deduplicate by URL/title similarity
        unique_news = []
        seen_titles = set()

        for news_item in all_news:
            # Extract title from "date: title" format
            if ': ' in news_item:
                title = news_item.split(': ', 1)[1]
            else:
                title = news_item

            # Simple deduplication by title similarity
            title_lower = title.lower()
            is_duplicate = False

            for seen_title in seen_titles:
                if title_lower == seen_title.lower():
                    is_duplicate = True
                    break
                # Check for similar titles (80% similarity)
                if len(title_lower) > 10 and len(seen_title.lower()) > 10:
                    similarity = sum(1 for a, b in zip(title_lower, seen_title.lower()) if a == b) / max(len(title_lower), len(seen_title.lower()))
                    if similarity > 0.8:
                        is_duplicate = True
                        break

            if not is_duplicate:
                seen_titles.add(title_lower)
                unique_news.append(news_item)

        # Sort by date (newest first)
        unique_news = sorted(unique_news, key=lambda x: x.split(":")[0], reverse=True)

        # Determine primary source (first non-empty, with Polygon priority)
        primary_source = "Unknown"
        for source_name, lst in (
            ("Polygon", polygon_news),
            ("Playwright", playwright_news),
            ("Finviz", finviz_news),
            ("Yahoo", yahoo_news),
            ("Finnhub", finnhub_news),
        ):
            if lst:
                primary_source = source_name
                break

        return ticker, (", ".join(unique_news) if unique_news else "No news"), primary_source

    except Exception as e:
        logging.error(f"Error in parallel news fetching for {ticker}: {e}")
        return ticker, "No news", "Unknown"

def get_all_news(ticker_date_pairs):
    """Fetch news for multiple ticker-date pairs with proper session management."""
    with requests.Session() as session:  # Use context manager for session
        retries = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(
            pool_connections=100,  # Increased to match global session for 20 workers
            pool_maxsize=100,      # Increased to match global session for 20 workers
            max_retries=retries
        )
        session.mount("https://", adapter)
        
        news_dict = {}
        max_workers = 25  # Increased to 25 workers for better performance
        batch_size = max_workers
        for i in range(0, len(ticker_date_pairs), batch_size):
            batch = ticker_date_pairs[i:i + batch_size]
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_pair = {executor.submit(fetch_news, pair[0], session, pair[1]): pair for pair in batch}
                for i, future in enumerate(as_completed(future_to_pair), 1):
                    try:
                        ticker, news, source = future.result()
                        news_dict[(ticker, future_to_pair[future][1])] = (news, source)
                    except Exception as e:
                        ticker, date = future_to_pair[future]
                        logging.error(f"Error fetching news for {ticker} on {date}: {e}")
                        news_dict[(ticker, date)] = ("No news", "Unknown")
                    if i % 10 == 0:
                        logging.info(f"Processed {i} of {len(ticker_date_pairs)} pairs")
            if i + batch_size < len(ticker_date_pairs):
                time.sleep(0.1)
    return news_dict

def calculate_low_after_high(session_data, high_time, session_name):
    try:
        if session_data.empty or pd.isna(high_time):
            return "Not Available", "Not Available"
        after_high_data = session_data.loc[high_time:]
        if after_high_data.empty:
            return "Not Available", "Not Available"
        if len(after_high_data) > 1:
            after_high_data = after_high_data.iloc[1:]
        if after_high_data.empty:
            return "Not Available", "Not Available"
        low_after = after_high_data["low"].min()
        low_after_time = after_high_data["low"].idxmin()
        return round(low_after, 2), low_after_time.strftime("%H:%M:%S") if not pd.isna(low_after_time) else "Not Available"
    except Exception as e:
        logging.error(f"Error calculating low after high for {session_name}: {e}")
        return "Not Available", "Not Available"

def calculate_low_1h_after_high(session_data, high_time, session_name):
    try:
        if session_data.empty or pd.isna(high_time):
            return "Not Available", "Not Available"
        one_hour_later = high_time + timedelta(hours=1)
        after_high_1h = session_data.loc[high_time:one_hour_later]
        if after_high_1h.empty:
            return "Not Available", "Not Available"
        elif len(after_high_1h) > 1:
            after_high_1h = after_high_1h.iloc[1:]
        if after_high_1h.empty:
            return "Not Available", "Not Available"
        low_1h_after = after_high_1h["low"].min()
        low_1h_after_time = after_high_1h["low"].idxmin()
        return (
            round(low_1h_after, 2),
            low_1h_after_time.strftime("%H:%M:%S") if not pd.isna(low_1h_after_time) else "Not Available"
        )
    except Exception as e:
        logging.warning(f"Error calculating 1h low after high for {session_name}: {e}")
        return "Not Available", "Not Available"

def calculate_low_30m_after_high(session_data, high_time, session_name):
    try:
        if session_data.empty or pd.isna(high_time):
            return "Not Available", "Not Available"
        thirty_min_later = high_time + timedelta(minutes=30)
        after_high_30m = session_data.loc[high_time:thirty_min_later]
        if after_high_30m.empty:
            return "Not Available", "Not Available"
        elif len(after_high_30m) > 1:
            after_high_30m = after_high_30m.iloc[1:]
        if after_high_30m.empty:
            return "Not Available", "Not Available"
        low_30m_after = after_high_30m["low"].min()
        low_30m_after_time = after_high_30m["low"].idxmin()
        return (
            round(low_30m_after, 2),
            low_30m_after_time.strftime("%H:%M:%S") if not pd.isna(low_30m_after_time) else "Not Available"
        )
    except Exception as e:
        logging.warning(f"Error calculating 30m low after high for {session_name}: {e}")
        return "Not Available", "Not Available"

def calculate_low_15m_after_high(session_data, high_time, session_name):
    try:
        if session_data.empty or pd.isna(high_time):
            return "Not Available", "Not Available"
        fifteen_min_later = high_time + timedelta(minutes=15)
        after_high_15m = session_data.loc[high_time:fifteen_min_later]
        if after_high_15m.empty:
            return "Not Available", "Not Available"
        elif len(after_high_15m) > 1:
            after_high_15m = after_high_15m.iloc[1:]
        if after_high_15m.empty:
            return "Not Available", "Not Available"
        low_15m_after = after_high_15m["low"].min()
        low_15m_after_time = after_high_15m["low"].idxmin()
        return (
            round(low_15m_after, 2),
            low_15m_after_time.strftime("%H:%M:%S") if not pd.isna(low_15m_after_time) else "Not Available"
        )
    except Exception as e:
        logging.warning(f"Error calculating 15m low after high for {session_name}: {e}")
        return "Not Available", "Not Available"
def calculate_rsi_at_high(data, high_time, timeframe_minutes=1, period=14, hist_data=None):
    """Calculate RSI at a specific high time for a given timeframe using data imputation like TradeZero Pro."""
    try:
        # Use hist_data for rolling calculation if available, but ensure final value is from target date
        if hist_data is not None and not hist_data.empty:
            # Calculate RSI using historical data for rolling window (30 days)
            if timeframe_minutes == 1:
                working_data = hist_data.copy()
            else:
                try:
                    working_data = hist_data.resample(f'{timeframe_minutes}min').agg({
                        'open': 'first',
                        'high': 'max',
                        'low': 'min',
                        'close': 'last',
                        'volume': 'sum'
                    }).dropna()
                    
                    if working_data.empty:
                        logging.debug(f"RSI calculation failed: resampled DataFrame is empty. Original={len(hist_data)}, Timeframe={timeframe_minutes}min")
                        return "Not Available"
                        
                except Exception as resample_error:
                    logging.debug(f"RSI resampling failed for {timeframe_minutes}min: {resample_error}")
                    return "Not Available"
            
            # Data imputation like TradeZero Pro - forward-fill missing closing prices
            working_data['close'] = working_data['close'].ffill()
            working_data['close'] = working_data['close'].bfill()
            
            # Ensure we have enough data for RSI calculation (standard 14-period RSI)
            if len(working_data) < period + 1:
                logging.debug(f"RSI calculation failed: insufficient data even after imputation. Available={len(working_data)}, Required={period + 1}, Timeframe={timeframe_minutes}min")
                return "Not Available"
            
            # Calculate RSI using standard 14-period window (like TradeZero Pro)
            rsi = ta.momentum.RSIIndicator(close=working_data['close'], window=period)
            rsi_values = rsi.rsi()
            
            # Find the RSI value at the specific high_time from resampled historical data
            # For longer timeframes (like 60min), use the RSI from historical data since target date data may not have enough points
            if high_time in data.index:
                # For all timeframes, use RSI from historical data for proper rolling calculation (like brokers)
                # This ensures consistent calculation across all timeframes
                if high_time in working_data.index:
                    # Exact match found in resampled historical data
                    rsi_at_high = rsi_values.loc[high_time]
                else:
                    # Find the bar that contains high_time
                    # For 1min: exact minute match
                    # For 2min/5min/60min: find the resampled bar that contains high_time
                    if timeframe_minutes == 1:
                        # For 1-minute, find exact minute or closest minute in historical data
                        time_diff = abs(working_data.index - high_time)
                        closest_idx = time_diff.argmin()
                        rsi_at_high = rsi_values.iloc[closest_idx]
                    else:
                        # For longer timeframes, find the bar that contains high_time
                        # Get all bars up to and including the one containing high_time
                        bars_before_high = working_data[working_data.index <= high_time]
                        if not bars_before_high.empty:
                            # Use the last bar that ends at or before high_time
                            rsi_at_high = rsi_values.loc[bars_before_high.index[-1]]
                        else:
                            # If no bars before high_time, find closest bar (shouldn't happen in practice)
                            time_diff = abs(working_data.index - high_time)
                            closest_idx = time_diff.argmin()
                            rsi_at_high = rsi_values.iloc[closest_idx]
            else:
                # High time not in target date data - return Not Available
                logging.debug(f"High time {high_time} not found in target date data")
                return "Not Available"
        else:
            # No historical data available - use target date data only
            use_data = data
            if use_data.empty or pd.isna(high_time):
                logging.debug(f"RSI calculation failed: use_data empty={use_data.empty}, high_time={high_time}")
                return "Not Available"
            
            # For 1-minute data, calculate RSI directly without resampling (like brokers do)
            if timeframe_minutes == 1:
                working_data = use_data.copy()
            else:
                try:
                    working_data = use_data.resample(f'{timeframe_minutes}min').agg({
                        'open': 'first',
                        'high': 'max',
                        'low': 'min',
                        'close': 'last',
                        'volume': 'sum'
                    }).dropna()
                    
                    if working_data.empty:
                        logging.debug(f"RSI calculation failed: resampled DataFrame is empty. Original={len(use_data)}, Timeframe={timeframe_minutes}min")
                        return "Not Available"
                        
                except Exception as resample_error:
                    logging.debug(f"RSI resampling failed for {timeframe_minutes}min: {resample_error}")
                    return "Not Available"
            
            # Data imputation like TradeZero Pro - forward-fill missing closing prices
            working_data['close'] = working_data['close'].ffill()
            working_data['close'] = working_data['close'].bfill()
            
            # Ensure we have enough data for RSI calculation (standard 14-period RSI)
            if len(working_data) < period + 1:
                logging.debug(f"RSI calculation failed: insufficient data even after imputation. Available={len(working_data)}, Required={period + 1}, Timeframe={timeframe_minutes}min")
                return "Not Available"
            
            # Calculate RSI using standard 14-period window (like TradeZero Pro)
            rsi = ta.momentum.RSIIndicator(close=working_data['close'], window=period)
            rsi_values = rsi.rsi()
            
            # Find the RSI value at or closest to the high time
            if high_time in working_data.index:
                rsi_at_high = rsi_values.loc[high_time]
            else:
                # Find the closest time within a reasonable range
                time_diff = abs(working_data.index - high_time)
                closest_idx = time_diff.argmin()
                rsi_at_high = rsi_values.iloc[closest_idx]
        
        if pd.isna(rsi_at_high):
            return "Not Available"
        
        return round(rsi_at_high, 2)
        
    except Exception as e:
        logging.warning(f"Error calculating RSI at high for {timeframe_minutes}min: {e}")
        return "Not Available"

def calculate_bb_percent_b_at_high(data, high_time, period=20, std_dev=2, hist_data=None):
    """Calculate Bollinger Bands %B at a specific high time."""
    try:
        # Use hist_data for rolling calculation if available, but ensure final value is from target date
        if hist_data is not None and not hist_data.empty:
            # Calculate Bollinger Bands using historical data for rolling window (30 days)
            if len(hist_data) < period + 1:
                return "Not Available"
            
            bb = ta.volatility.BollingerBands(close=hist_data['close'], window=period, window_dev=std_dev)
            bb_upper = bb.bollinger_hband()
            bb_lower = bb.bollinger_lband()
            bb_middle = bb.bollinger_mavg()
            
            # Find the %B value at the specific high_time from target date data
            if high_time in data.index:
                # Use target date data for the actual time point
                if len(data) < period + 1:
                    return "Not Available"
                
                bb_target = ta.volatility.BollingerBands(close=data['close'], window=period, window_dev=std_dev)
                bb_upper_target = bb_target.bollinger_hband()
                bb_lower_target = bb_target.bollinger_lband()
                bb_middle_target = bb_target.bollinger_mavg()
                
                if high_time in data.index:
                    upper = bb_upper_target.loc[high_time]
                    lower = bb_lower_target.loc[high_time]
                    middle = bb_middle_target.loc[high_time]
                    close = data['close'].loc[high_time]
                else:
                    # Find closest time in target date data
                    time_diff = abs(data.index - high_time)
                    closest_idx = time_diff.argmin()
                    upper = bb_upper_target.iloc[closest_idx]
                    lower = bb_lower_target.iloc[closest_idx]
                    middle = bb_middle_target.iloc[closest_idx]
                    close = data['close'].iloc[closest_idx]
            else:
                # High time not in target date data - return Not Available
                logging.debug(f"High time {high_time} not found in target date data")
                return "Not Available"
        else:
            # No historical data available - use target date data only
            use_data = data
            if use_data.empty or pd.isna(high_time):
                return "Not Available"
            
            if len(use_data) < period + 1:
                return "Not Available"
            
            # Calculate Bollinger Bands
            bb = ta.volatility.BollingerBands(close=use_data['close'], window=period, window_dev=std_dev)
            bb_upper = bb.bollinger_hband()
            bb_lower = bb.bollinger_lband()
            bb_middle = bb.bollinger_mavg()
            
            # Find the %B value at or closest to the high time
            if high_time in use_data.index:
                upper = bb_upper.loc[high_time]
                lower = bb_lower.loc[high_time]
                middle = bb_middle.loc[high_time]
                close = use_data['close'].loc[high_time]
            else:
                # Find the closest time
                time_diff = abs(use_data.index - high_time)
                closest_idx = time_diff.argmin()
                upper = bb_upper.iloc[closest_idx]
                lower = bb_lower.iloc[closest_idx]
                middle = bb_middle.iloc[closest_idx]
                close = use_data['close'].iloc[closest_idx]
        
        if pd.isna(upper) or pd.isna(lower) or pd.isna(close):
            return "Not Available"
        
        # Calculate %B: (close - lower) / (upper - lower)
        if upper == lower:
            return 0.5  # Middle of the bands
        else:
            percent_b = (close - lower) / (upper - lower)
            return round(percent_b, 3)
        
    except Exception as e:
        logging.warning(f"Error calculating BB %B at high: {e}")
        return "Not Available"

def get_short_interest_data(ticker, date, shares_outstanding=None, float_shares=None):
    """Fetch short interest data from Polygon and calculate percentages."""
    try:
        client = get_polygon_client()
        polygon_limiter.acquire()
        
        # Only fetch ticker details if shares data is not provided (to avoid duplicate API calls)
        if shares_outstanding is None or float_shares is None:
            try:
                # Fetch ticker details for shares data
                details = client.get_ticker_details(ticker)
                shares_outstanding = details.weighted_shares_outstanding if hasattr(details, 'weighted_shares_outstanding') else None
                float_shares = details.share_class_shares_outstanding if hasattr(details, 'share_class_shares_outstanding') else None  # Approximation for float
            except Exception as e:
                logging.warning(f"Could not fetch ticker details for {ticker}: {e}")
                shares_outstanding = None
                float_shares = None
        
        # Fetch most recent short interest up to the given date (bi-monthly)
        response = []
        for item in client.list_short_interest(ticker=ticker, settlement_date_lte=date, sort='settlement_date.desc', limit=1):
            response.append(item)
        
        if not response:
            logging.warning(f"No short interest data for {ticker} up to {date}")
            return {"short_float_pct": "Not Available", "short_interest_pct": "Not Available", "days_to_cover": "Not Available"}
        
        data = response[0]
        short_interest = data.short_interest if hasattr(data, 'short_interest') else "Not Available"
        days_to_cover = data.days_to_cover if hasattr(data, 'days_to_cover') else "Not Available"
        
        # Calculate percentages if we have shares data
        short_float_pct = "Not Available"
        short_interest_pct = "Not Available"
        if isinstance(short_interest, (int, float)):
            if float_shares and float_shares > 0:
                short_float_pct = round((short_interest / float_shares) * 100, 2)
            if shares_outstanding and shares_outstanding > 0:
                short_interest_pct = round((short_interest / shares_outstanding) * 100, 2)
        
        return {
            "short_float_pct": short_float_pct,
            "short_interest_pct": short_interest_pct,
            "days_to_cover": round(days_to_cover, 2) if isinstance(days_to_cover, (int, float)) else days_to_cover
        }
    
    except Exception as e:
        logging.warning(f"Error fetching short interest from Polygon for {ticker}: {e}")
        return {"short_float_pct": "Not Available", "short_interest_pct": "Not Available", "days_to_cover": "Not Available"}

def get_short_volume_data(ticker, date):
    """Fetch daily short volume data from Polygon for the given date using REST API."""
    try:
        polygon_limiter.acquire()
        
        # Use direct REST API endpoint instead of Python SDK
        url = f"https://api.polygon.io/stocks/v1/short-volume"
        params = {
            'ticker': ticker,
            'date': date,
            'apikey': POLYGON_API_KEY
        }
        
        response = requests.get(url, params=params, timeout=20)
        response.raise_for_status()
        js = response.json()
        results = js.get('results') or []
        
        if not results:
            logging.warning(f"No short volume data for {ticker} on {date}")
            return {"short_volume": "Not Available", "short_volume_ratio": "Not Available"}
        
        entry = results[0]
        sv = entry.get('short_volume') or entry.get('shortVolume')
        svr = entry.get('short_volume_ratio') or entry.get('shortVolumeRatio')
        
        return {
            "short_volume": sv if isinstance(sv, (int, float)) else (sv if sv is not None else "Not Available"),
            "short_volume_ratio": round(float(svr), 2) if isinstance(svr, (int, float)) else (svr if svr is not None else "Not Available")
        }
    
    except Exception as e:
        logging.warning(f"Error fetching short volume for {ticker} on {date}: {e}")
        return {"short_volume": "Not Available", "short_volume_ratio": "Not Available"}

def get_short_interest_raw_data(ticker, date):
    """Fetch raw short interest data from Polygon using REST API."""
    try:
        polygon_limiter.acquire()
        
        # Use direct REST API endpoint for short interest
        url = f"https://api.polygon.io/stocks/v1/short-interest"
        params = {
            'ticker': ticker,
            'settlement_date.lte': date,
            'apikey': POLYGON_API_KEY,
            'limit': 1,
            'sort': 'settlement_date.desc'
        }
        
        response = requests.get(url, params=params, timeout=20)
        response.raise_for_status()
        js = response.json()
        results = js.get('results') or []
        
        if not results:
            logging.warning(f"No short interest data for {ticker} up to {date}")
            return {"short_interest": "Not Available"}
        
        entry = results[0]
        short_interest = entry.get('short_interest')
        
        return {
            "short_interest": short_interest if isinstance(short_interest, (int, float)) else (short_interest if short_interest is not None else "Not Available")
        }
    
    except Exception as e:
        logging.warning(f"Error fetching short interest for {ticker} on {date}: {e}")
        return {"short_interest": "Not Available"}

def _create_minute_data(data, result, previous_close):
    """Create minute-by-minute data with forward fill from 04:00 to 20:00"""
    
    # Create time range from 04:00 to 20:00 (PM to AH close)
    start_time = datetime.strptime('04:00', '%H:%M').time()
    end_time = datetime.strptime('20:00', '%H:%M').time()
    
    # Get the date from the data index
    base_date = data.index[0].date()
    
    # Create minute data dictionary
    minute_data = {}
    last_price = previous_close
    
    # Fill minute data
    current_time = datetime.combine(base_date, start_time)
    end_datetime = datetime.combine(base_date, end_time)
    
    while current_time <= end_datetime:
        time_key = current_time.strftime('%H:%M')
        
        # Find bar for this minute in the DataFrame
        matching_bars = data[data.index.time == current_time.time()]
        
        if not matching_bars.empty:
            minute_data[time_key] = matching_bars['close'].iloc[-1]  # Use last bar if multiple
            last_price = matching_bars['close'].iloc[-1]
        else:
            # No trade this minute, use last price (forward fill)
            minute_data[time_key] = last_price
        
        current_time += timedelta(minutes=1)
    
    result["raw_minute_data"] = minute_data

    # Convenience: set specific clock time prices if available
    def set_price(label: str, hhmm: str):
        if hhmm in minute_data and minute_data[hhmm] is not None:
            result[label] = round(float(minute_data[hhmm]), 4)

    set_price("price_10am", "10:00")
    set_price("price_1030am", "10:30")
    set_price("price_11am", "11:00")
    set_price("price_1130am", "11:30")
    set_price("price_12pm", "12:00")

def _create_vwap_deviation_minute_columns(data, result, date, ticker):
    """Create per-minute VWAP % Deviation columns from 04:00 to 20:00.

    Deviation = ((price - vwap) / vwap) * 100. Uses cumulative VWAP since 04:00.
    Forward fills price data for minutes with no trades.
    """
    try:
        if not isinstance(data, pd.DataFrame) or data.empty:
            logging.warning(f"[VWAP DEV] No data provided for VWAP calculation")
            return

        # Restrict to 04:00–20:00 ET for the target date
        start_ts = pd.Timestamp(f"{date} 04:00:00").tz_localize(EASTERN_TZ)
        end_ts = pd.Timestamp(f"{date} 20:00:00").tz_localize(EASTERN_TZ)
        
        # Filter data to the target date and time range
        day_data = data.loc[(data.index >= start_ts) & (data.index <= end_ts)].copy()
        if day_data.empty:
            logging.warning(f"[VWAP DEV] No data in 04:00-20:00 range for {date}")
            return

        # Sort data by timestamp to ensure proper ordering
        day_data = day_data.sort_index()
        
        # Get the first available price for forward filling
        first_price = day_data['close'].iloc[0] if not day_data.empty else None
        if first_price is None or pd.isna(first_price):
            logging.warning(f"[VWAP DEV] No valid first price found")
            return
        
        # Get previous trading day's closing price and VWAP for continuous calculation
        try:
            from datetime import datetime, timedelta
            target_date_dt = datetime.strptime(date, "%Y-%m-%d")
            
            client = get_polygon_client()
            prev_close = None
            prev_vwap = None
            
            # Look back up to 7 days to find the last trading day
            for days_back in range(1, 8):
                prev_date = target_date_dt - timedelta(days=days_back)
                prev_date_str = prev_date.strftime("%Y-%m-%d")
                
                # Skip weekends
                if prev_date.weekday() >= 5:  # Saturday=5, Sunday=6
                    continue
                
                try:
                    polygon_limiter.acquire()
                    prev_aggs = client.get_aggs(
                        ticker=ticker,
                        multiplier=1,
                        timespan="day",
                        from_=prev_date_str,
                        to=prev_date_str
                    )
                    
                    if prev_aggs and len(prev_aggs) > 0:
                        prev_close = prev_aggs[0].close
                        prev_vwap = getattr(prev_aggs[0], 'vwap', prev_close)
                        logging.info(f"[VWAP DEV] Found previous trading day ({prev_date_str}) close: {prev_close}, VWAP: {prev_vwap}")
                        break
                except Exception as e:
                    logging.debug(f"[VWAP DEV] No data for {prev_date_str}: {e}")
                    continue
            
            if prev_close is None:
                logging.warning(f"[VWAP DEV] No previous trading day data found, using first price")
                prev_close = first_price
                prev_vwap = first_price
                
        except Exception as e:
            logging.warning(f"[VWAP DEV] Error fetching previous data: {e}, using first price")
            prev_close = first_price
            prev_vwap = first_price

        # Create a complete minute-by-minute time series from 04:00 to 20:00
        minute_range = pd.date_range(start=start_ts, end=end_ts, freq='1min', tz=EASTERN_TZ)
        minute_df = pd.DataFrame(index=minute_range)
        minute_df['close'] = np.nan
        minute_df['volume'] = 0.0

        # Fill in actual trade data where available
        for ts, row in day_data.iterrows():
            if ts in minute_df.index:
                minute_df.loc[ts, 'close'] = row['close']
                minute_df.loc[ts, 'volume'] = row['volume']

        # Special handling for 4:00 AM - use yesterday's close if no trades at 4:00 AM
        first_minute_ts = minute_range[0]  # 4:00 AM timestamp
        if pd.isna(minute_df.loc[first_minute_ts, 'close']):
            # No trade at 4:00 AM, use yesterday's closing price
            minute_df.loc[first_minute_ts, 'close'] = prev_close
            logging.info(f"[VWAP DEV] No trade at 4:00 AM, using previous close: {prev_close}")
        else:
            logging.info(f"[VWAP DEV] Found actual trade at 4:00 AM: {minute_df.loc[first_minute_ts, 'close']}")
        
        # Forward fill prices for minutes with no trades
        minute_df['close'] = minute_df['close'].ffill()
        minute_df['close'] = minute_df['close'].fillna(first_price)  # Fill any remaining NaN with first price
        minute_df['volume'] = minute_df['volume'].fillna(0)

        # Compute cumulative VWAP using forward-filled close prices
        minute_df['pv'] = (minute_df['close'].astype(float)) * (minute_df['volume'].astype(float))
        minute_df['cum_vol'] = minute_df['volume'].cumsum()
        minute_df['cum_pv'] = minute_df['pv'].cumsum()
        
        # Calculate VWAP: use cumulative VWAP, but handle zero volume periods properly
        minute_df['vwap_calc'] = np.where(minute_df['cum_vol'] > 0,
                                         minute_df['cum_pv'] / minute_df['cum_vol'],
                                         np.nan)
        
        # Forward-fill VWAP values for minutes with no trades
        # This ensures VWAP continues from the last calculated value
        minute_df['vwap_calc'] = minute_df['vwap_calc'].ffill()
        
        # For the very first minute, if no VWAP calculated yet, use the previous day's VWAP
        minute_df['vwap_calc'] = minute_df['vwap_calc'].fillna(prev_vwap)

        # Compute deviation vectorized
        with np.errstate(divide='ignore', invalid='ignore'):
            dev_series = ((minute_df['close'] - minute_df['vwap_calc']) / minute_df['vwap_calc']) * 100.0
        dev_series = dev_series.replace([np.inf, -np.inf], np.nan)

        # Build columns in chronological order (4:00 AM to 8:00 PM)
        for ts in minute_range:
            dev = dev_series.loc[ts]
            price = minute_df['close'].loc[ts]
            vwap = minute_df['vwap_calc'].loc[ts]
            price_diff = price - vwap if pd.notna(price) and pd.notna(vwap) else np.nan
            
            # Use Windows-safe hour formatting: %I with lstrip('0') or '0'
            hh = ts.strftime('%I').lstrip('0') or '0'
            mm = ts.strftime('%M')
            ampm = ts.strftime('%p')
            
            # VWAP % Deviation column with % symbol
            pct_label = f"VWAP % Deviation {hh}:{mm}{ampm}" if mm != '00' else f"VWAP % Deviation {hh}{ampm}"
            if pd.notna(dev):
                result[pct_label] = f"{round(float(dev), 2)}%"
            else:
                result[pct_label] = "Not Available"
            
            # VWAP Deviation Price column (actual price difference)
            price_label = f"VWAP Deviation Price {hh}:{mm}{ampm}" if mm != '00' else f"VWAP Deviation Price {hh}{ampm}"
            if pd.notna(price_diff):
                result[price_label] = round(float(price_diff), 4)
            else:
                result[price_label] = "Not Available"
                
        logging.info(f"[VWAP DEV] Created {len([k for k in result.keys() if k.startswith('VWAP % Deviation')])} VWAP deviation columns")
        
    except Exception as e:
        logging.warning(f"[MINUTE VWAP DEV] Error creating VWAP deviation minute columns: {e}")


def get_intraday_data(ticker, date, previous_close=None, shares_outstanding=None, float_shares=None):
    """Fetch intraday data from Polygon.io, relying on API for split adjustments."""
    polygon_limiter.acquire()
    
    client = get_polygon_client()
    try:
        start_date = datetime.strptime(date, "%Y-%m-%d")
        end_date = start_date + timedelta(days=1)

        hist_start = (start_date - timedelta(days=30)).strftime("%Y-%m-%d")
        hist_data = pd.DataFrame()
        try:
            polygon_limiter.acquire()
            hist_aggs = client.get_aggs(ticker, 1, "minute", hist_start, end_date.strftime("%Y-%m-%d"), adjusted=True)
            if hist_aggs:
                hist_data = pd.DataFrame([{
                    'timestamp': agg.timestamp,
                    'open': agg.open,
                    'high': agg.high,
                    'low': agg.low,
                    'close': agg.close,
                    'volume': agg.volume
                } for agg in hist_aggs])
                hist_data['datetime'] = pd.to_datetime(hist_data['timestamp'], unit='ms').dt.tz_localize('UTC').dt.tz_convert(EASTERN_TZ)
                hist_data.set_index('datetime', inplace=True)
                hist_data = hist_data[~hist_data.index.duplicated(keep='first')]
                logging.info(f"Fetched {len(hist_data)} historical bars for {ticker} from {hist_start} to {end_date.strftime('%Y-%m-%d')}")
            else:
                logging.warning(f"No historical data returned for {ticker}")
        except Exception as e:
            logging.warning(f"Historical data fetch failed for {ticker}: {e}")

        current_date = datetime.now(EASTERN_TZ).date()
        if start_date.date() > current_date:
            logging.warning(f"Target date {date} is in the future for {ticker}")
            return None, False

        pre_market_start = pd.Timestamp(f"{date} 04:00:00").tz_localize(EASTERN_TZ)
        pre_market_end = pd.Timestamp(f"{date} 09:29:59").tz_localize(EASTERN_TZ)
        regular_start = pd.Timestamp(f"{date} 09:30:00").tz_localize(EASTERN_TZ)
        regular_end = pd.Timestamp(f"{date} 16:00:00").tz_localize(EASTERN_TZ)
        after_hours_end = pd.Timestamp(f"{date} 20:00:00").tz_localize(EASTERN_TZ)

        pre_market_data = None
        for attempt in range(3):
            try:
                polygon_limiter.acquire()
                aggs = client.get_aggs(
                    ticker,
                    multiplier=1,
                    timespan="minute",
                    from_=start_date.strftime("%Y-%m-%d"),
                    to=start_date.strftime("%Y-%m-%d"),
                    adjusted=True
                )
                # Ensure aggs is a list of proper objects, not bytes
                if not isinstance(aggs, list):
                    logging.warning(f"Invalid aggs type for {ticker}: {type(aggs)}")
                    continue
                    
                pre_market_aggs = pd.DataFrame([
                    {
                        "timestamp": getattr(agg, 'timestamp', None),
                        "open": getattr(agg, 'open', None),
                        "high": getattr(agg, 'high', None),
                        "low": getattr(agg, 'low', None),
                        "close": getattr(agg, 'close', None),
                        "volume": getattr(agg, 'volume', None)
                    }
                    for agg in aggs if hasattr(agg, 'timestamp')
                ])
                logging.info(f"Fetched {len(pre_market_aggs)} pre-market records for {ticker} on {date}")
                if not pre_market_aggs.empty:
                    pre_market_aggs["timestamp"] = pd.to_datetime(pre_market_aggs["timestamp"], unit="ms", errors="coerce")
                    # Check for null timestamps using pandas method
                    if pre_market_aggs["timestamp"].isnull().any():
                        logging.warning(f"Invalid timestamps in pre-market data for {ticker}")
                        pre_market_aggs = pre_market_aggs.dropna(subset=["timestamp"])
                    pre_market_aggs["datetime"] = pre_market_aggs["timestamp"].dt.tz_localize("UTC").dt.tz_convert(EASTERN_TZ)
                    pre_market_aggs.set_index("datetime", inplace=True)
                    pre_market_data = pre_market_aggs[pre_market_start:pre_market_end].copy()
                    if not pre_market_data.empty:
                        logging.info(f"Pre-market data range for {ticker}: {pre_market_data.index[0]} to {pre_market_data.index[-1]}")
                break
            except Exception as e:
                if "429" in str(e):
                    logging.warning(f"Rate limit hit for {ticker} pre-market, attempt {attempt + 1}")
                    time.sleep(0.1)  # Add delay for rate limit
                    continue
                logging.error(f"Error fetching pre-market data for {ticker}: {e}")
                pre_market_data = None
                break

        full_data = None
        for attempt in range(3):
            try:
                polygon_limiter.acquire()
                aggs = client.get_aggs(
                    ticker,
                    multiplier=1,
                    timespan="minute",
                    from_=start_date.strftime("%Y-%m-%d"),
                    to=end_date.strftime("%Y-%m-%d"),
                    adjusted=True
                )
                # Ensure aggs is a list of proper objects, not bytes
                if not isinstance(aggs, list):
                    logging.warning(f"Invalid aggs type for {ticker}: {type(aggs)}")
                    continue
                    
                full_data = pd.DataFrame([
                    {
                        "timestamp": getattr(agg, 'timestamp', None),
                        "open": getattr(agg, 'open', None),
                        "high": getattr(agg, 'high', None),
                        "low": getattr(agg, 'low', None),
                        "close": getattr(agg, 'close', None),
                        "volume": getattr(agg, 'volume', None)
                    }
                    for agg in aggs if hasattr(agg, 'timestamp')
                ])
                logging.info(f"Fetched {len(full_data)} aggregate records (including after-hours) for {ticker} on {date}")
                break
            except Exception as e:
                if "429" in str(e):
                    logging.warning(f"Rate limit hit for {ticker}, attempt {attempt + 1}")
                    time.sleep(0.1)  # Add delay for rate limit
                    continue
                logging.error(f"Error fetching aggregate data for {ticker}: {e}")
                full_data = None
                break

        if pre_market_data is None and (full_data is None or full_data.empty):
            logging.warning(f"No Polygon.io data for {ticker} on {date}, falling back to yfinance which may not be split-adjusted")
            used_yfinance = True
            try:
                yf_ticker = yf.Ticker(ticker)
                polygon_limiter.acquire()
                # Fix: Remove conflicting parameters - use only start date for intraday data with split adjustment
                hist = yf_ticker.history(interval="1m", start=date, auto_adjust=True)
                if not hist.empty:
                    data = hist[["Open", "High", "Low", "Close", "Volume"]].reset_index()
                    data.columns = ["timestamp", "open", "high", "low", "close", "volume"]
                    data["timestamp"] = pd.to_datetime(data["timestamp"], errors="coerce")
                    if data["timestamp"].dt.tz:
                        data["datetime"] = data["timestamp"].dt.tz_convert(EASTERN_TZ)
                    else:
                        data["datetime"] = data["timestamp"].dt.tz_localize(EASTERN_TZ)
                    data.set_index("datetime", inplace=True)
                    logging.info(f"Fetched {len(data)} yfinance records for {ticker} on {date}")
                else:
                    logging.warning(f"No yfinance intraday data for {ticker} on {date}")
                    return None, used_yfinance
            except Exception as e:
                logging.error(f"Error fetching yfinance intraday data for {ticker}: {e}")
                return None, used_yfinance
        else:
            used_yfinance = False

        # Store filtered data sections before concatenation to avoid re-slicing issues
        regular_market_filtered = None
        after_hours_filtered = None

        if pre_market_data is not None and full_data is not None:
            full_data["timestamp"] = pd.to_datetime(full_data["timestamp"], unit="ms", errors="coerce")
            if full_data["timestamp"].isna().any():
                logging.warning(f"Invalid timestamps in Polygon.io data for {ticker}")
                return None, used_yfinance
            full_data["datetime"] = full_data["timestamp"].dt.tz_localize("UTC").dt.tz_convert(EASTERN_TZ)
            full_data.set_index("datetime", inplace=True)
            regular_after_hours = full_data[regular_start:after_hours_end].copy()
            # Extract regular_market and after_hours from regular_after_hours before concatenation
            regular_market_filtered = regular_after_hours[regular_start:regular_end].copy()
            after_hours_filtered = regular_after_hours[regular_end:after_hours_end].copy()
            data = pd.concat([pre_market_data, regular_after_hours])
        elif full_data is not None:
            full_data["timestamp"] = pd.to_datetime(full_data["timestamp"], unit="ms", errors="coerce")
            if full_data["timestamp"].isna().any():
                logging.warning(f"Invalid timestamps in Polygon.io data for {ticker}")
                return None, used_yfinance
            full_data["datetime"] = full_data["timestamp"].dt.tz_localize("UTC").dt.tz_convert(EASTERN_TZ)
            full_data.set_index("datetime", inplace=True)
            # Extract regular_market and after_hours from full_data before assignment
            regular_market_filtered = full_data[regular_start:regular_end].copy()
            after_hours_filtered = full_data[regular_end:after_hours_end].copy()
            data = full_data
        elif pre_market_data is not None:
            data = pre_market_data
        else:
            logging.warning(f"No intraday data for {ticker} on {date}")
            return None, used_yfinance

        if data.empty:
            logging.warning(f"No intraday data for {ticker} on {date}")
            return None, used_yfinance

        data = data[~data.index.duplicated(keep='first')]
        
        # Enhanced data validation - ensure data is a DataFrame
        if not isinstance(data, pd.DataFrame):
            logging.error(f"Data is not a DataFrame for {ticker} on {date}: {type(data)}")
            return None, used_yfinance
            
        required_columns = ["open", "high", "low", "close", "volume"]
        if not all(col in data.columns for col in required_columns):
            logging.error(f"Missing required columns for {ticker} on {date}")
            return None, used_yfinance
            
        # Check for NaN values
        if data[required_columns].isna().any().any():
            logging.warning(f"NaN values in required columns for {ticker} on {date}")
            return None, used_yfinance

        max_price = data[["open", "high", "low", "close"]].max().max()
        min_price = data[["open", "high", "low", "close"]].min().min()
        
        if max_price > 1000:
            logging.error(f"Suspicious high prices for {ticker} on {date}: max=${max_price}")
            return None, used_yfinance
        
        if min_price <= 0:
            logging.warning(f"Invalid zero/negative prices for {ticker} on {date}")
            data = data[(data["open"] > 0) & (data["high"] > 0) & (data["low"] > 0) & (data["close"] > 0)]
            if data.empty:
                logging.error(f"No valid price data for {ticker} on {date}")
                return None, used_yfinance

        # Use pre_market_data directly if available (already filtered correctly), otherwise slice from data
        if pre_market_data is not None and not pre_market_data.empty:
            pre_market = pre_market_data.copy()
            logging.debug(f"[DATA SLICE DEBUG] Using pre_market_data directly for {ticker}: {len(pre_market)} rows, high max={pre_market['high'].max()}")
        else:
            pre_market = data[pre_market_start:pre_market_end].copy()
            logging.debug(f"[DATA SLICE DEBUG] Slicing pre_market from data for {ticker}: {len(pre_market)} rows, high max={pre_market['high'].max()}")
        
        # Use filtered regular_market and after_hours if available (already filtered correctly), otherwise slice from data
        if regular_market_filtered is not None and not regular_market_filtered.empty:
            regular_market = regular_market_filtered.copy()
            logging.debug(f"[DATA SLICE DEBUG] Using regular_market_filtered directly for {ticker}: {len(regular_market)} rows, high max={regular_market['high'].max()}")
        else:
            regular_market = data[regular_start:regular_end].copy()
            logging.debug(f"[DATA SLICE DEBUG] Slicing regular_market from data for {ticker}: {len(regular_market)} rows, high max={regular_market['high'].max()}")
        
        if after_hours_filtered is not None and not after_hours_filtered.empty:
            after_hours = after_hours_filtered.copy()
            logging.debug(f"[DATA SLICE DEBUG] Using after_hours_filtered directly for {ticker}: {len(after_hours)} rows, high max={after_hours['high'].max()}")
        else:
            after_hours = data[regular_end:after_hours_end].copy()
            logging.debug(f"[DATA SLICE DEBUG] Slicing after_hours from data for {ticker}: {len(after_hours)} rows, high max={after_hours['high'].max()}")

        # DEBUG PATCH: Log after-hours diagnostics with type checks
        if isinstance(after_hours, pd.DataFrame):
            logging.info(f"[DEBUG] {ticker} {date} after_hours rows: {after_hours.shape[0]}")
            logging.info(f"[DEBUG] after_hours head: {after_hours.head()} ")
        else:
            logging.warning(f"[DEBUG] after_hours is not a DataFrame: {type(after_hours)}")
        if isinstance(data, pd.DataFrame) and isinstance(data.index, pd.DatetimeIndex):
            logging.info(f"[DEBUG] data index min: {data.index.min()}, max: {data.index.max()}")
            logging.info(f"[DEBUG] data index type: {type(data.index)}, tz: {getattr(data.index, 'tz', None)}")
            logging.info(f"[DEBUG] data index sample: {list(data.index[:5])}")
        else:
            logging.warning(f"[DEBUG] data is not a DataFrame or invalid index: {type(data)}")

        result = {
            "total_day_volume": format_number(float(data["volume"].sum())) if isinstance(data, pd.DataFrame) else "Not Available",
            "pm_open": "No PM Data",
            "pm_high": "Not Available",
            "pm_high_time": "Not Available",
            "pm_low_after_high": "Not Available",
            "pm_low_after_high_time": "Not Available",
            "pm_low_15m_after_high": "Not Available",
            "pm_low_15m_after_high_time": "Not Available",
            "pm_low_30m_after_high": "Not Available",
            "pm_low_30m_after_high_time": "Not Available",
            "pm_low_1h_after_high": "Not Available",
            "pm_low_1h_after_high_time": "Not Available",
            "pm_low": "Not Available",
            "pm_low_time": "Not Available",
            "pm_close": "Not Available",
            "reclaim_vwap_pm": "Not Available",
            "vwap_breakdown_time_pm": "Not Available",
            "vol_pm_high": "Not Available",
            "vol_pm_low": "Not Available",
            "open": "No Regular Data",
            "open_high": "Not Available",
            "open_high_time": "Not Available",
            "open_low_after_high": "Not Available",
            "open_low_after_high_time": "Not Available",
            "open_low_15m_after_high": "Not Available",
            "open_low_15m_after_high_time": "Not Available",
            "open_low_30m_after_high": "Not Available",
            "open_low_30m_after_high_time": "Not Available",
            "open_low_1h_after_high": "Not Available",
            "open_low_1h_after_high_time": "Not Available",
            "open_low": "Not Available",
            "open_low_time": "Not Available",
            "reclaim_vwap_open": "Not Available",
            "vwap_breakdown_time_open": "Not Available",
            "open_close": "Not Available",
            "vol_open_high": "Not Available",
            "vol_open_low": "Not Available",
            "ah_open": "No AH Data",
            "ah_high": "Not Available",
            "ah_high_time": "Not Available",
            "ah_low": "Not Available",
            "ah_low_time": "Not Available",
            "ah_low_after_high": "Not Available",
            "ah_low_after_high_time": "Not Available",
            "ah_low_15m_after_high": "Not Available",
            "ah_low_15m_after_high_time": "Not Available",
            "ah_low_30m_after_high": "Not Available",
            "ah_low_30m_after_high_time": "Not Available",
            "ah_low_1h_after_high": "Not Available",
            "ah_low_1h_after_high_time": "Not Available",
            "ah_close": "No AH Data",
            "reclaim_vwap_ah": "Not Available",
            "vwap_breakdown_time_ah": "Not Available",
            "vol_ah_high": "Not Available",
            "Vol @ 6AM": "Not Available",
            "Vol @ 7AM": "Not Available",
            "Vol @ 8AM": "Not Available",
            "Vol @ 9AM": "Not Available",
            "Vol @ 10AM": "Not Available",
            "Vol @ 11AM": "Not Available",
            "Vol @ 12PM": "Not Available",
            "percent_gain_pm_high": "Not Available",
            "percent_gain_open_high": "Not Available",
            "percent_gain_ah_high": "Not Available",
            "market_cap_pm_high": "Not Available",
            "market_cap_open_high": "Not Available",
            "market_cap_10am": "Not Available",
            "market_cap_11am": "Not Available",
            "market_cap_12pm": "Not Available",
            "market_cap_ah_high": "Not Available",
            # NEW: Only truly new metrics for short strategy
            "pm_short_pnl": "Not Available",    # Profit/Loss from PM high to open
            "pm_short_pnl_pct": "Not Available", # P&L as percentage
            "open_short_pnl": "Not Available",   # Profit/Loss from open high to 12pm
            "open_short_pnl_pct": "Not Available", # P&L as percentage
            "pm_to_open_gap": "Not Available",   # Gap from PM close to open
            "open_to_12pm_range": "Not Available", # Range from open to 12pm
            "pm_volume_spike": "Not Available",  # Volume spike in pre-market
            "open_volume_spike": "Not Available", # Volume spike at open
            "pm_momentum": "Not Available",      # Pre-market momentum indicator
            "open_momentum": "Not Available",    # Open momentum indicator
            "pm_reversal_signal": "Not Available", # Pre-market reversal signals
            "open_reversal_signal": "Not Available", # Open reversal signals
            # RSI columns for PM and Open sessions
                        "PM RSI at High 1min": "Not Available",
            "PM RSI at High 2min": "Not Available",
            "PM RSI at High 5min": "Not Available",
            "PM RSI at High 1hour": "Not Available",
            "MACD_Line_PM_High": "Not Available",
            "MACD_Signal_PM_High": "Not Available",
            "MACD_Histogram_PM_High": "Not Available",
            "Open RSI at High 1min": "Not Available",
            "Open RSI at High 2min": "Not Available",
            "Open RSI at High 5min": "Not Available",
            "Open RSI at High 1hour": "Not Available",
            "MACD_Line_Open_High": "Not Available",
            "MACD_Signal_Open_High": "Not Available",
            "MACD_Histogram_Open_High": "Not Available",
            # Short interest and short float columns
            "Short Float %": "Not Available",
            "Short Interest % of Outstanding": "Not Available",
            "Days to Cover": "Not Available",
            "Short Interest": "Not Available",
            # Technical indicators
            "BB %B at PM High": "Not Available",
            "BB %B at Open High": "Not Available",
            "Short Volume": "Not Available",
            "Short Volume Ratio": "Not Available",
            # Time-based prices
            "price_10am": "Not Available",
            "price_1030am": "Not Available",
            "price_11am": "Not Available",
            "price_1130am": "Not Available",
            "price_12pm": "Not Available",
            # VWAP Deviation columns
            "vwap_deviation_pm_high": "Not Available",
            "vwap_deviation_open_high": "Not Available",
            "vwap_deviation_ah_high": "Not Available",
            "raw_minute_data": {}
        }

        # Process regular market with type check
        if isinstance(regular_market, pd.DataFrame) and not regular_market.empty:
            try:
                regular_market = regular_market[~regular_market.index.duplicated(keep='first')]
                if regular_market.empty or len(regular_market) < 2:
                    logging.warning(f"Insufficient regular market data for {ticker}: {len(regular_market)} rows")
                    raise ValueError("Too few data points")
                
                if regular_market[["open", "high", "low", "close", "volume"]].isna().all().any():
                    logging.warning(f"Regular market data for {ticker} contains all NaN values")
                    raise ValueError("Invalid data: all NaN")
                
                result["open"] = round(regular_market.iloc[0]["open"], 2)
                open_high = regular_market["high"].max()
                if pd.isna(open_high):
                    logging.warning(f"No valid high price in regular market data for {ticker}")
                    raise ValueError("No valid high price")
                result["open_high"] = round(float(open_high), 2) if not pd.isna(open_high) else "Not Available"
                
                open_high_idx = regular_market["high"].idxmax()
                open_high_time = open_high_idx if not pd.isna(open_high_idx) else None
                result["open_high_time"] = open_high_time.strftime("%H:%M:%S") if open_high_time else "Not Available"
                
                # Calculate RSI, Bollinger Bands, and MACD at Open high
                if open_high_time is not None:
                    # RSI calculations for different timeframes
                    result["Open RSI at High 1min"] = calculate_rsi_at_high(regular_market, open_high_time, 1, hist_data=hist_data)
                    result["Open RSI at High 2min"] = calculate_rsi_at_high(regular_market, open_high_time, 2, hist_data=hist_data)
                    result["Open RSI at High 5min"] = calculate_rsi_at_high(regular_market, open_high_time, 5, hist_data=hist_data)
                    result["Open RSI at High 1hour"] = calculate_rsi_at_high(regular_market, open_high_time, 60, hist_data=hist_data)
                    
                    # Bollinger Bands %B calculation
                    result["BB %B at Open High"] = calculate_bb_percent_b_at_high(regular_market, open_high_time, hist_data=hist_data)
                    
                    # MACD calculation at Open high
                    try:
                        macd_result = calculate_macd_at_high(regular_market, open_high_time, hist_data=hist_data)
                        if macd_result:
                            result["MACD_Line_Open_High"] = macd_result['macd_line']
                            result["MACD_Signal_Open_High"] = macd_result['signal_line']
                            result["MACD_Histogram_Open_High"] = macd_result['histogram']
                        else:
                            result["MACD_Line_Open_High"] = "Not Available"
                            result["MACD_Signal_Open_High"] = "Not Available"
                            result["MACD_Histogram_Open_High"] = "Not Available"
                    except Exception as e:
                        logging.warning(f"Error calculating MACD at Open high for {ticker}: {e}")
                        result["MACD_Line_Open_High"] = "Not Available"
                        result["MACD_Signal_Open_High"] = "Not Available"
                        result["MACD_Histogram_Open_High"] = "Not Available"
                    
                    # Note: RVOL calculation moved to later in the function where volume_data is defined
                
                result["open_low_after_high"], result["open_low_after_high_time"] = calculate_low_after_high(
                    regular_market, open_high_time, "Regular Market"
                )
                result["open_low_15m_after_high"], result["open_low_15m_after_high_time"] = calculate_low_15m_after_high(
                    regular_market, open_high_time, "Regular Market"
                )
                result["open_low_30m_after_high"], result["open_low_30m_after_high_time"] = calculate_low_30m_after_high(
                    regular_market, open_high_time, "Regular Market"
                )
                result["open_low_1h_after_high"], result["open_low_1h_after_high_time"] = calculate_low_1h_after_high(
                    regular_market, open_high_time, "Regular Market"
                )
                
                open_low = regular_market["low"].min()
                result["open_low"] = round(open_low, 2) if not pd.isna(open_low) else "Not Available"
                open_low_idx = regular_market["low"].idxmin()
                result["open_low_time"] = open_low_idx.strftime("%H:%M:%S") if not pd.isna(open_low_idx) else "Not Available"
                result["open_close"] = round(regular_market.iloc[-1]["close"], 2)
                
                if not regular_market["high"].isna().all() and not pd.isna(open_high_idx):
                    first_high_data = regular_market.loc[open_high_idx]
                    result["vol_open_high"] = format_number(float(first_high_data["volume"]))
                
                if not regular_market["low"].isna().all() and not pd.isna(open_low_idx):
                    first_low_data = regular_market.loc[open_low_idx]
                    result["vol_open_low"] = format_number(float(first_low_data["volume"]))
                
                vwap_result = calculate_vwap_reclaim(regular_market, "Regular Market", result["open_high"], open_high_time)
                result["reclaim_vwap_open"] = vwap_result["reclaim"]
                result["vwap_breakdown_time_open"] = vwap_result.get("breakdown_time", "Not Available")
                
                # Calculate VWAP deviation for Open high
                if vwap_result.get("vwap") is not None and result["open_high"] != "Not Available":
                    vwap_deviation = calculate_vwap_deviation_percentage(result["open_high"], vwap_result["vwap"])
                    if vwap_deviation is not None:
                        result["vwap_deviation_open_high"] = f"{round(vwap_deviation, 2)}%"
                
                if result["open"] != "No Regular Data" and isinstance(result["open"], (int, float)) and result["open"] > 0:
                    if result['open'] > 0:  # Additional safety check
                        percent_gain = safe_percentage_change(result['open_high'], result['open'])
                        result["percent_gain_open_high"] = f"{percent_gain}%" if percent_gain != "Not Available" else "Not Available"

                if shares_outstanding and result["open_high"] != "No Regular Data":
                    result["market_cap_open_high"] = format_number(shares_outstanding * result["open_high"])
                    
                    # Calculate market cap at specific times (10AM, 11AM, 12PM)
                    try:
                        # 10AM market cap
                        ten_am_start = pd.Timestamp(f"{date} 10:00:00").tz_localize(EASTERN_TZ)
                        ten_am_end = pd.Timestamp(f"{date} 10:05:00").tz_localize(EASTERN_TZ)
                        ten_am_data = regular_market[ten_am_start:ten_am_end]
                        if not ten_am_data.empty:
                            ten_am_price = ten_am_data["close"].iloc[-1]
                            result["market_cap_10am"] = format_number(shares_outstanding * ten_am_price)
                        else:
                            result["market_cap_10am"] = "Not Available"
                        
                        # 11AM market cap
                        eleven_am_start = pd.Timestamp(f"{date} 11:00:00").tz_localize(EASTERN_TZ)
                        eleven_am_end = pd.Timestamp(f"{date} 11:05:00").tz_localize(EASTERN_TZ)
                        eleven_am_data = regular_market[eleven_am_start:eleven_am_end]
                        if not eleven_am_data.empty:
                            eleven_am_price = eleven_am_data["close"].iloc[-1]
                            result["market_cap_11am"] = format_number(shares_outstanding * eleven_am_price)
                        else:
                            result["market_cap_11am"] = "Not Available"
                        
                        # 12PM market cap
                        twelve_pm_start = pd.Timestamp(f"{date} 12:00:00").tz_localize(EASTERN_TZ)
                        twelve_pm_end = pd.Timestamp(f"{date} 12:05:00").tz_localize(EASTERN_TZ)
                        twelve_pm_data = regular_market[twelve_pm_start:twelve_pm_end]
                        if not twelve_pm_data.empty:
                            twelve_pm_price = twelve_pm_data["close"].iloc[-1]
                            result["market_cap_12pm"] = format_number(shares_outstanding * twelve_pm_price)
                        else:
                            result["market_cap_12pm"] = "Not Available"
                            
                    except Exception as e:
                        logging.warning(f"Error calculating time-based market caps for {ticker}: {e}")
                        result["market_cap_10am"] = "Not Available"
                        result["market_cap_11am"] = "Not Available"
                        result["market_cap_12pm"] = "Not Available"
                        
                    # Calculate prices at specific times
                    try:
                        # 10AM price
                        ten_am_start = pd.Timestamp(f"{date} 10:00:00").tz_localize(EASTERN_TZ)
                        ten_am_end = pd.Timestamp(f"{date} 10:05:00").tz_localize(EASTERN_TZ)
                        ten_am_data = regular_market[ten_am_start:ten_am_end]
                        if not ten_am_data.empty:
                            result["price_10am"] = round(ten_am_data["close"].iloc[-1], 2)
                        else:
                            result["price_10am"] = "Not Available"
                        
                        # 10:30AM price
                        ten_thirty_am_start = pd.Timestamp(f"{date} 10:30:00").tz_localize(EASTERN_TZ)
                        ten_thirty_am_end = pd.Timestamp(f"{date} 10:35:00").tz_localize(EASTERN_TZ)
                        ten_thirty_am_data = regular_market[ten_thirty_am_start:ten_thirty_am_end]
                        if not ten_thirty_am_data.empty:
                            result["price_1030am"] = round(ten_thirty_am_data["close"].iloc[-1], 2)
                        else:
                            result["price_1030am"] = "Not Available"
                        
                        # 11AM price
                        eleven_am_start = pd.Timestamp(f"{date} 11:00:00").tz_localize(EASTERN_TZ)
                        eleven_am_end = pd.Timestamp(f"{date} 11:05:00").tz_localize(EASTERN_TZ)
                        eleven_am_data = regular_market[eleven_am_start:eleven_am_end]
                        if not eleven_am_data.empty:
                            result["price_11am"] = round(eleven_am_data["close"].iloc[-1], 2)
                        else:
                            result["price_11am"] = "Not Available"
                        
                        # 11:30AM price
                        eleven_thirty_am_start = pd.Timestamp(f"{date} 11:30:00").tz_localize(EASTERN_TZ)
                        eleven_thirty_am_end = pd.Timestamp(f"{date} 11:35:00").tz_localize(EASTERN_TZ)
                        eleven_thirty_am_data = regular_market[eleven_thirty_am_start:eleven_thirty_am_end]
                        if not eleven_thirty_am_data.empty:
                            result["price_1130am"] = round(eleven_thirty_am_data["close"].iloc[-1], 2)
                        else:
                            result["price_1130am"] = "Not Available"
                        
                        # 12PM price
                        twelve_pm_start = pd.Timestamp(f"{date} 12:00:00").tz_localize(EASTERN_TZ)
                        twelve_pm_end = pd.Timestamp(f"{date} 12:05:00").tz_localize(EASTERN_TZ)
                        twelve_pm_data = regular_market[twelve_pm_start:twelve_pm_end]
                        if not twelve_pm_data.empty:
                            result["price_12pm"] = round(twelve_pm_data["close"].iloc[-1], 2)
                        else:
                            result["price_12pm"] = "Not Available"
                            
                    except Exception as e:
                        logging.warning(f"Error calculating time-based prices for {ticker}: {e}")
                        result["price_10am"] = "Not Available"
                        result["price_1030am"] = "Not Available"
                        result["price_11am"] = "Not Available"
                        result["price_1130am"] = "Not Available"
                        result["price_12pm"] = "Not Available"
                    
                # NEW: Open to 12pm short strategy calculations
                if result["open_high"] != "Not Available" and result["open"] != "No Regular Data":
                    try:
                        open_high_price = float(result["open_high"])
                        open_price = float(result["open"])
                        
                        # Get 12pm price (approximate - using data around 12pm)
                        twelve_pm_start = pd.Timestamp(f"{date} 12:00:00").tz_localize(EASTERN_TZ)
                        twelve_pm_end = pd.Timestamp(f"{date} 12:30:00").tz_localize(EASTERN_TZ)
                        twelve_pm_data = regular_market[twelve_pm_start:twelve_pm_end]
                        
                        if not twelve_pm_data.empty:
                            twelve_pm_price = twelve_pm_data["close"].iloc[-1]  # Last price in 12pm window
                            
                            # Open to 12pm short strategy P&L
                            result["open_short_pnl"] = open_high_price - twelve_pm_price
                            open_short_pnl_pct = safe_percentage_ratio(open_high_price - twelve_pm_price, open_high_price)
                            result["open_short_pnl_pct"] = f"{open_short_pnl_pct}%" if open_short_pnl_pct != "Not Available" else "Not Available"
                            
                            # Range analysis - Open price to 12pm price
                            result["open_to_12pm_range"] = f"${open_price:.2f} - ${twelve_pm_price:.2f}"
                            
                            # Volume spike detection at open
                            open_avg_volume = regular_market["volume"].mean()
                            open_max_volume = regular_market["volume"].max()
                            if open_max_volume > open_avg_volume * 3:
                                result["open_volume_spike"] = f"{(open_max_volume/open_avg_volume):.1f}x"
                            
                            # Momentum indicator
                            open_price_change = safe_percentage_change(twelve_pm_price, open_high_price)
                            if open_price_change == "Not Available":
                                result["open_momentum"] = "Not Available"
                            elif open_price_change < -20:
                                result["open_momentum"] = "Strong Down"
                            elif open_price_change < -2:
                                result["open_momentum"] = "Down"
                            elif open_price_change > 20:
                                result["open_momentum"] = "Strong Up"
                            elif open_price_change > 2:
                                result["open_momentum"] = "Up"
                            else:
                                result["open_momentum"] = "Neutral"
                            
                            # Reversal signal
                            if open_price_change == "Not Available":
                                result["open_reversal_signal"] = "Not Available"
                            elif open_price_change < -50 and result["open_volume_spike"] != "Not Available":
                                result["open_reversal_signal"] = "Strong Short"
                            elif open_price_change < -30:
                                result["open_reversal_signal"] = "Short"
                            elif open_price_change < -10 and open_price_change >= -29.99:
                                result["open_reversal_signal"] = "Weak Short"
                            else:
                                result["open_reversal_signal"] = "None"
                            
                    except Exception as e:
                        logging.warning(f"Error calculating open short metrics for {ticker}: {e}")
                    
            except Exception as e:
                logging.warning(f"Error processing regular market data for {ticker}: {e}")
        # Now process pre-market with type check
        if isinstance(pre_market, pd.DataFrame) and not pre_market.empty:
            try:
                pre_market = pre_market[~pre_market.index.duplicated(keep='first')]
                if pre_market.empty or len(pre_market) < 2:
                    logging.warning(f"Insufficient pre-market data for {ticker}: {len(pre_market)} rows")
                    raise ValueError("Too few data points")
                
                # Check for NaN values in required columns
                required_cols = ["open", "high", "low", "close", "volume"]
                if all(col in pre_market.columns for col in required_cols):
                    if pre_market[required_cols].isna().all().any():
                        logging.warning(f"Pre-market data for {ticker} contains all NaN values")
                        raise ValueError("Invalid data: all NaN")
                
                result["pm_open"] = round(pre_market.iloc[0]["open"], 2)
                pm_high = pre_market["high"].max()
                if pd.isna(pm_high):
                    logging.warning(f"No valid high price for {ticker} pre-market data")
                    raise ValueError("No valid high price")
                result["pm_high"] = round(float(pm_high), 2) if not pd.isna(pm_high) else "Not Available"
                
                # Debug: Compare PM High with minute data to detect scaling issues
                # Check if we can access the original data DataFrame for comparison
                # Note: 'data' variable should be in scope from earlier in the function
                try:
                    # Try to get pre-market data from the original data source for comparison
                    pre_market_from_data = pre_market.copy()  # Use the pre_market we already have
                    # Also check raw minute data if available (created later in the function)
                    # This validation will help detect if there's a data source mismatch
                    if not pre_market_from_data.empty:
                        pre_market_high_check = pre_market_from_data["high"].max()
                        if not pd.isna(pre_market_high_check) and not pd.isna(pm_high):
                            # Log if there's any discrepancy (should be 0 if using same source)
                            if abs(pm_high - pre_market_high_check) > 0.01:  # More than 1 cent difference
                                logging.warning(f"[PM HIGH DEBUG] {ticker} Internal PM High check: calculated=${pm_high:.2f}, DataFrame max=${pre_market_high_check:.2f}")
                except Exception as debug_error:
                    logging.debug(f"[PM HIGH DEBUG] Could not perform internal validation for {ticker}: {debug_error}")
                
                # Validate PM High against minute data to detect scaling issues
                if isinstance(data, pd.DataFrame) and not data.empty:
                    # Use the original full 'data' DataFrame for minute-by-minute comparison
                    pre_market_minute_data = data[pre_market_start:pre_market_end].copy()
                    if not pre_market_minute_data.empty:
                        minute_high = pre_market_minute_data["high"].max()
                        if not pd.isna(minute_high) and not pd.isna(pm_high):
                            if abs(pm_high - minute_high) > 1.0:  # More than $1 difference
                                logging.warning(f"[PM HIGH DEBUG] {ticker} PM High mismatch: pre_market high=${pm_high:.2f}, minute data high=${minute_high:.2f}, ratio={pm_high/minute_high:.4f}")
                                # If there's a significant discrepancy, use the minute data high instead
                                if abs(pm_high - minute_high) > pm_high * 0.5:  # More than 50% difference
                                    logging.error(f"[PM HIGH FIX] {ticker} Correcting PM High from ${pm_high:.2f} to ${minute_high:.2f}")
                                    pm_high = minute_high
                                    result["pm_high"] = round(float(pm_high), 2)
                
                pm_high_idx = pre_market["high"].idxmax()
                pm_high_time = pm_high_idx if not pd.isna(pm_high_idx) else None
                result["pm_high_time"] = pm_high_time.strftime("%H:%M:%S") if pm_high_time else "Not Available"
                
                # Calculate RSI, Bollinger Bands, and MACD at PM high
                if pm_high_time is not None:
                    result["PM RSI at High 1min"] = calculate_rsi_at_high(pre_market, pm_high_time, 1, hist_data=hist_data)
                    result["PM RSI at High 2min"] = calculate_rsi_at_high(pre_market, pm_high_time, 2, hist_data=hist_data)
                    result["PM RSI at High 5min"] = calculate_rsi_at_high(pre_market, pm_high_time, 5, hist_data=hist_data)
                    result["PM RSI at High 1hour"] = calculate_rsi_at_high(pre_market, pm_high_time, 60, hist_data=hist_data)
                    
                    result["BB %B at PM High"] = calculate_bb_percent_b_at_high(pre_market, pm_high_time, hist_data=hist_data)
                    
                    # MACD calculation at PM high
                    try:
                        macd_result = calculate_macd_at_high(pre_market, pm_high_time, hist_data=hist_data)
                        if macd_result:
                            result["MACD_Line_PM_High"] = macd_result['macd_line']
                            result["MACD_Signal_PM_High"] = macd_result['signal_line']
                            result["MACD_Histogram_PM_High"] = macd_result['histogram']
                        else:
                            result["MACD_Line_PM_High"] = "Not Available"
                            result["MACD_Signal_PM_High"] = "Not Available"
                            result["MACD_Histogram_PM_High"] = "Not Available"
                    except Exception as e:
                        logging.warning(f"Error calculating MACD at PM high for {ticker}: {e}")
                        result["MACD_Line_PM_High"] = "Not Available"
                        result["MACD_Signal_PM_High"] = "Not Available"
                        result["MACD_Histogram_PM_High"] = "Not Available"
                
                result["pm_low_after_high"], result["pm_low_after_high_time"] = calculate_low_after_high(
                    pre_market, pm_high_time, "Pre-Market"
                )
                result["pm_low_15m_after_high"], result["pm_low_15m_after_high_time"] = calculate_low_15m_after_high(
                    pre_market, pm_high_time, "Pre-Market"
                )
                result["pm_low_30m_after_high"], result["pm_low_30m_after_high_time"] = calculate_low_30m_after_high(
                    pre_market, pm_high_time, "Pre-Market"
                )
                result["pm_low_1h_after_high"], result["pm_low_1h_after_high_time"] = calculate_low_1h_after_high(
                    pre_market, pm_high_time, "Pre-Market"
                )
                
                pm_low = pre_market["low"].min()
                result["pm_low"] = round(pm_low, 2) if not pd.isna(pm_low) else "Not Available"
                pm_low_idx = pre_market["low"].idxmin()
                result["pm_low_time"] = pm_low_idx.strftime("%H:%M:%S") if not pd.isna(pm_low_idx) else "Not Available"
                result["pm_close"] = round(pre_market.iloc[-1]["close"], 2)
                
                if not pre_market["high"].isna().all() and not pd.isna(pm_high_idx):
                    first_high_data = pre_market.loc[pm_high_idx]
                    result["vol_pm_high"] = format_number(float(first_high_data["volume"]))
                
                if not pre_market["low"].isna().all() and not pd.isna(pm_low_idx):
                    first_low_data = pre_market.loc[pm_low_idx]
                    result["vol_pm_low"] = format_number(float(first_low_data["volume"]))
                
                vwap_result = calculate_vwap_reclaim(pre_market, "Pre-Market", result["pm_high"], pm_high_time)
                result["reclaim_vwap_pm"] = vwap_result["reclaim"]
                result["vwap_breakdown_time_pm"] = vwap_result.get("breakdown_time", "Not Available")
                
                # Calculate VWAP deviation for PM high
                if vwap_result.get("vwap") is not None and result["pm_high"] != "Not Available":
                    vwap_deviation = calculate_vwap_deviation_percentage(result["pm_high"], vwap_result["vwap"])
                    if vwap_deviation is not None:
                        result["vwap_deviation_pm_high"] = f"{round(vwap_deviation, 2)}%"
                
                if result["pm_open"] != "No PM Data" and isinstance(result["pm_open"], (int, float)) and result["pm_open"] > 0:
                    if result['pm_open'] > 0:  # Additional safety check
                        percent_gain = safe_percentage_change(result['pm_high'], result['pm_open'])
                        result["percent_gain_pm_high"] = f"{percent_gain}%" if percent_gain != "Not Available" else "Not Available"

                if shares_outstanding and result["pm_high"] != "No PM Data":
                    result["market_cap_pm_high"] = format_number(shares_outstanding * result["pm_high"])
                    
                # NEW: Pre-market short strategy calculations (now with access to result["open"])
                if result["pm_high"] != "Not Available" and result["pm_close"] != "Not Available" and result["open"] != "No Regular Data":
                    try:
                        pm_high_price = float(result["pm_high"])
                        pm_close_price = float(result["pm_close"])
                        open_price = float(result["open"])
                        
                        # Pre-market short strategy P&L
                        result["pm_short_pnl"] = round(pm_high_price - open_price, 2)
                        pm_short_pnl_pct = safe_percentage_ratio(pm_high_price - open_price, pm_high_price)
                        result["pm_short_pnl_pct"] = f"{pm_short_pnl_pct}%" if pm_short_pnl_pct != "Not Available" else "Not Available"
                        
                        # Gap analysis
                        pm_to_open_gap = safe_percentage_change(open_price, pm_close_price)
                        result["pm_to_open_gap"] = f"{pm_to_open_gap}%" if pm_to_open_gap != "Not Available" else "Not Available"
                        
                        # Volume spike detection
                        pm_avg_volume = pre_market["volume"].mean()
                        pm_max_volume = pre_market["volume"].max()
                        if pm_max_volume > pm_avg_volume * 3:
                            result["pm_volume_spike"] = f"{(pm_max_volume/pm_avg_volume):.1f}x"
                        
                        # Momentum indicator
                        pm_price_change = safe_percentage_change(pm_close_price, pm_high_price)
                        if pm_price_change == "Not Available":
                            result["pm_momentum"] = "Not Available"
                        elif pm_price_change < -20:
                            result["pm_momentum"] = "Strong Down"
                        elif pm_price_change < -2:
                            result["pm_momentum"] = "Down"
                        elif pm_price_change > 20:
                            result["pm_momentum"] = "Strong Up"
                        elif pm_price_change > 2:
                            result["pm_momentum"] = "Up"
                        else:
                            result["pm_momentum"] = "Neutral"
                        
                        # Reversal signal
                        if pm_price_change == "Not Available":
                            result["pm_reversal_signal"] = "Not Available"
                        elif pm_price_change < -50 and result["pm_volume_spike"] != "Not Available":
                            result["pm_reversal_signal"] = "Strong Short"
                        elif pm_price_change < -30:
                            result["pm_reversal_signal"] = "Short"
                        elif pm_price_change < -10 and pm_price_change >= -29.99:
                            result["pm_reversal_signal"] = "Weak Short"
                        else:
                            result["pm_reversal_signal"] = "None"
                        
                    except Exception as e:
                        logging.warning(f"Error calculating PM short metrics for {ticker}: {e}")
                    
            except Exception as e:
                logging.warning(f"Error processing pre-market data for {ticker}: {e}")
                
        if isinstance(after_hours, pd.DataFrame) and not after_hours.empty:
            try:
                after_hours = after_hours[~after_hours.index.duplicated(keep='first')]
                if after_hours.empty or len(after_hours) < 2:
                    logging.warning(f"Insufficient after-hours data for {ticker}: {len(after_hours)} rows")
                    raise ValueError("Too few data points")
                
                if after_hours[["open", "high", "low", "close", "volume"]].isna().all().any():
                    logging.warning(f"After-hours data for {ticker} contains all NaN values")
                    raise ValueError("Invalid data: all NaN")
                
                result["ah_open"] = round(after_hours.iloc[0]["open"], 2)
                ah_high = after_hours["high"].max()
                if pd.isna(ah_high):
                    logging.warning(f"No valid high price in after-hours data for {ticker}")
                    raise ValueError("No valid high price")
                result["ah_high"] = round(float(ah_high), 2) if not pd.isna(ah_high) else "Not Available"
                
                ah_high_idx = after_hours["high"].idxmax()
                ah_high_time = ah_high_idx if not pd.isna(ah_high_idx) else None
                result["ah_high_time"] = ah_high_time.strftime("%H:%M:%S") if ah_high_time else "Not Available"
                
                # Add low after high to mimic PM/Open
                result["ah_low_after_high"], result["ah_low_after_high_time"] = calculate_low_after_high(
                    after_hours, ah_high_time, "After-Hours"
                )
                result["ah_low_15m_after_high"], result["ah_low_15m_after_high_time"] = calculate_low_15m_after_high(
                    after_hours, ah_high_time, "After-Hours"
                )
                result["ah_low_30m_after_high"], result["ah_low_30m_after_high_time"] = calculate_low_30m_after_high(
                    after_hours, ah_high_time, "After-Hours"
                )
                result["ah_low_1h_after_high"], result["ah_low_1h_after_high_time"] = calculate_low_1h_after_high(
                    after_hours, ah_high_time, "After-Hours"
                )
                
                ah_low = after_hours["low"].min()
                result["ah_low"] = round(ah_low, 2) if not pd.isna(ah_low) else "Not Available"
                result["ah_low_time"] = after_hours["low"].idxmin().strftime("%H:%M:%S") if not pd.isna(after_hours["low"].idxmin()) else "Not Available"
                result["ah_close"] = round(after_hours.iloc[-1]["close"], 2)
                
                if not after_hours["high"].isna().all() and not pd.isna(ah_high_idx):
                    first_high_data = after_hours.loc[ah_high_idx]
                    result["vol_ah_high"] = format_number(float(first_high_data["volume"]))
                
                vwap_result = calculate_vwap_reclaim(after_hours, "After-Hours", result["ah_high"], ah_high_time)
                result["reclaim_vwap_ah"] = vwap_result["reclaim"]
                result["vwap_breakdown_time_ah"] = vwap_result.get("breakdown_time", "Not Available")
                
                # Calculate VWAP deviation for AH high
                if vwap_result.get("vwap") is not None and result["ah_high"] != "Not Available":
                    vwap_deviation = calculate_vwap_deviation_percentage(result["ah_high"], vwap_result["vwap"])
                    if vwap_deviation is not None:
                        result["vwap_deviation_ah_high"] = f"{round(vwap_deviation, 2)}%"
                
                if result["ah_open"] != "No AH Data" and isinstance(result["ah_open"], (int, float)) and result["ah_open"] > 0:
                        percent_gain = safe_percentage_change(result['ah_high'], result['ah_open'])
                        result["percent_gain_ah_high"] = f"{percent_gain}%" if percent_gain != "Not Available" else "Not Available"

                if shares_outstanding and result["ah_high"] != "No AH Data":
                    result["market_cap_ah_high"] = format_number(shares_outstanding * result["ah_high"])
                    
            except Exception as e:
                logging.warning(f"Error processing after-hours data for {ticker}: {e}")

        # Fetch short interest data from Polygon
        try:
            # Always try to get short interest data, even if float/shares data failed
            short_data = get_short_interest_data(ticker, date, shares_outstanding=result.get("shares_outstanding_raw"), float_shares=result.get("float_shares_raw"))
            result["Short Float %"] = short_data["short_float_pct"]
            result["Short Interest % of Outstanding"] = short_data["short_interest_pct"]
            result["Days to Cover"] = short_data["days_to_cover"]
        except Exception as e:
            logging.warning(f"Error fetching short interest data for {ticker}: {e}")
            result["Short Float %"] = "Not Available"
            result["Short Interest % of Outstanding"] = "Not Available"
            result["Days to Cover"] = "Not Available"

        # Fetch and add short volume data
        short_volume_data = get_short_volume_data(ticker, date)
        result["Short Volume"] = short_volume_data["short_volume"]
        result["Short Volume Ratio"] = short_volume_data["short_volume_ratio"]

        # Fetch and add raw short interest data
        short_interest_data = get_short_interest_raw_data(ticker, date)
        result["Short Interest"] = short_interest_data["short_interest"]

        # Create minute-by-minute data
        try:
            # Use data (target date only) instead of hist_data (30 days) for minute data creation
            if isinstance(data, pd.DataFrame) and not data.empty and previous_close:
                _create_minute_data(data, result, previous_close)
                logging.info(f"[MINUTE DATA] Created minute data for {ticker} from 04:00 to 20:00")
                
                # Validate PM High against minute data to detect scaling issues
                if result.get("pm_high") != "Not Available" and isinstance(result.get("pm_high"), (int, float)):
                    pm_high_value = float(result["pm_high"])
                    # Check pre-market minute data (04:00 to 09:29)
                    pre_market_start_check = pd.Timestamp(f"{date} 04:00:00").tz_localize(EASTERN_TZ)
                    pre_market_end_check = pd.Timestamp(f"{date} 09:29:59").tz_localize(EASTERN_TZ)
                    pre_market_minute_check = data[pre_market_start_check:pre_market_end_check].copy()
                    if not pre_market_minute_check.empty:
                        minute_high_check = pre_market_minute_check["high"].max()
                        if not pd.isna(minute_high_check):
                            ratio = pm_high_value / minute_high_check if minute_high_check > 0 else 0
                            if abs(pm_high_value - minute_high_check) > 1.0:  # More than $1 difference
                                logging.warning(f"[PM HIGH VALIDATION] {ticker} PM High=${pm_high_value:.2f}, Minute Data High=${minute_high_check:.2f}, Ratio={ratio:.4f}")
                                # If ratio suggests a 100x scaling issue (like $5 vs $500), correct it
                                if 0.009 < ratio < 0.011:  # Approximately 1/100 (scaling down issue)
                                    logging.error(f"[PM HIGH FIX] {ticker} Detected 100x scaling issue! Correcting PM High from ${pm_high_value:.2f} to ${minute_high_check:.2f}")
                                    result["pm_high"] = round(float(minute_high_check), 2)
                                elif 90 < ratio < 110:  # Approximately 100x (scaling up issue)
                                    logging.error(f"[PM HIGH FIX] {ticker} Detected 100x scaling issue! Correcting PM High from ${pm_high_value:.2f} to ${minute_high_check:.2f}")
                                    result["pm_high"] = round(float(minute_high_check), 2)
                
                # Also create per-minute VWAP deviation columns
                _create_vwap_deviation_minute_columns(data, result, date, ticker)
        except Exception as e:
            logging.warning(f"[MINUTE DATA] Error creating minute data for {ticker}: {e}")

        # Calculate float rotation and RVOL if we have float shares data
        logging.info(f"[DEBUG] Float shares value for {ticker}: {float_shares} (type: {type(float_shares)})")
        # Check if float_shares is a valid number (not None, not "Not Available", not 0)
        if float_shares and float_shares != "Not Available" and float_shares != "None" and str(float_shares).lower() not in ['nan', 'none', 'not available']:
            try:
                # Ensure float_shares is a number
                try:
                    if float_shares is None or float_shares == "" or str(float_shares).lower() in ['nan', 'none', 'not available']:
                        float_shares_num = None
                    elif isinstance(float_shares, str):
                        float_shares_num = float(float_shares)
                    else:
                        float_shares_num = float(float_shares)
                except (ValueError, TypeError):
                    logging.warning(f"[FLOAT ROTATION/RVOL] Invalid float_shares value for {ticker}: {float_shares}")
                    float_shares_num = None
                
                if float_shares_num and float_shares_num > 0:
                    logging.info(f"[FLOAT ROTATION] Calculating float rotation for {ticker} on {date}")
                    volume_data = get_intraday_cumulative_volume(ticker, datetime.strptime(date, "%Y-%m-%d"), POLYGON_API_KEY)
                    
                    if volume_data:
                        # Calculate float rotation using float shares
                        float_rotation = calculate_float_rotation(volume_data, float_shares_num)
                    
                    if float_rotation:
                        # Add float rotation data to result with proper time point mapping
                        time_point_mapping = {
                            '5AM': '5AM',
                            '6AM': '6AM', 
                            '7AM': '7AM',
                            '8AM': '8AM',
                            '9AM': '9AM',
                            '9:30AM': '9:30AM',
                            '10AM': '10AM',
                            '11AM': '11AM',
                            '12PM': '12PM',
                            '1PM': '1PM',
                            '4PM': '4PM'
                        }
                        
                        for time_point, rotation in float_rotation.items():
                            mapped_name = time_point_mapping.get(time_point, time_point)
                            result[f"FloatRotation{mapped_name}"] = round(rotation, 4)
                        
                        logging.info(f"[FLOAT ROTATION] Completed for {ticker}: {len(float_rotation)} time points calculated")
                    else:
                        logging.warning(f"[FLOAT ROTATION] Failed to calculate float rotation for {ticker}")
                    
                    # Calculate RVOL and get historical data with exact high times
                    logging.info(f"[RVOL] Calculating RVOL for {ticker} on {date}")
                    pm_high_time_str = result.get("pm_high_time", "Not Available")
                    open_high_time_str = result.get("open_high_time", "Not Available")
                    rvol_data, historical_data = calculate_rvol(volume_data, ticker, date, POLYGON_API_KEY, pm_high_time_str, open_high_time_str)
                    
                    if rvol_data:
                        # Add RVOL data to result with proper time point mapping
                        time_point_mapping = {
                            '5AM': '5AM',
                            '6AM': '6AM', 
                            '7AM': '7AM',
                            '8AM': '8AM',
                            '9AM': '9AM',
                            '9:30AM': '9:30AM',
                            '10AM': '10AM',
                            '11AM': '11AM',
                            '12PM': '12PM',
                            '1PM': '1PM',
                            '4PM': '4PM'
                        }
                        
                        for time_point, rvol in rvol_data.items():
                            mapped_name = time_point_mapping.get(time_point, time_point)
                            if rvol is not None:
                                result[f"RVOL{mapped_name}"] = round(rvol, 2)
                            else:
                                result[f"RVOL{mapped_name}"] = "Not Available"
                        
                        logging.info(f"[RVOL] Completed for {ticker}: {len(rvol_data)} time points calculated")
                        
                        # Add special time points (PM_High and Open_High) from updated RVOL data
                        if 'PM_High' in rvol_data:
                            if rvol_data['PM_High'] is not None:
                                result["RVOLPMHigh"] = round(rvol_data['PM_High'], 2)
                            else:
                                result["RVOLPMHigh"] = "Not Available"
                        else:
                            result["RVOLPMHigh"] = "Not Available"
                            
                        if 'Open_High' in rvol_data:
                            if rvol_data['Open_High'] is not None:
                                result["RVOLOpenHigh"] = round(rvol_data['Open_High'], 2)
                            else:
                                result["RVOLOpenHigh"] = "Not Available"
                        else:
                            result["RVOLOpenHigh"] = "Not Available"
                            
                        logging.info(f"[RVOL] Special time points added: PM_High={result.get('RVOLPMHigh', 'N/A')}, Open_High={result.get('RVOLOpenHigh', 'N/A')}")
                    else:
                        logging.warning(f"[RVOL] Failed to calculate RVOL for {ticker}")
                else:
                    logging.warning(f"[FLOAT ROTATION/RVOL] No volume data available for {ticker}")
            except Exception as e:
                logging.error(f"[FLOAT ROTATION/RVOL] Error calculating for {ticker}: {e}")
        else:
            logging.info(f"[FLOAT ROTATION/RVOL] Skipping for {ticker} - no valid float shares data available (float_shares: {float_shares})")

        # Calculate hourly volumes (6AM to 12PM)
        try:
            logging.info(f"[HOURLY VOLUMES] Calculating hourly volumes for {ticker} on {date}")
            raw_data = get_raw_intraday_data(ticker, date, POLYGON_API_KEY)
            
            if raw_data is not None and not raw_data.empty:
                hourly_volumes = calculate_hourly_volumes(raw_data)
                
                if hourly_volumes:
                    # Add hourly volume data to result
                    for hour, volume in hourly_volumes.items():
                        result[f"Vol @ {hour}"] = format_number(volume) if volume > 0 else "Not Available"
                    
                    logging.info(f"[HOURLY VOLUMES] Completed for {ticker}: {len(hourly_volumes)} hours calculated")
                else:
                    logging.warning(f"[HOURLY VOLUMES] No hourly volume data calculated for {ticker}")
                    # Set default values
                    for hour in ['6AM', '7AM', '8AM', '9AM', '10AM', '11AM', '12PM']:
                        result[f"Vol @ {hour}"] = "Not Available"
            else:
                logging.warning(f"[HOURLY VOLUMES] No raw data available for {ticker}")
                # Set default values
                for hour in ['6AM', '7AM', '8AM', '9AM', '10AM', '11AM', '12PM']:
                    result[f"Vol @ {hour}"] = "Not Available"
                    
        except Exception as e:
            logging.error(f"[HOURLY VOLUMES] Error calculating hourly volumes for {ticker}: {e}")
            # Set default values
            for hour in ['6AM', '7AM', '8AM', '9AM', '10AM', '11AM', '12PM']:
                result[f"Vol @ {hour}"] = "Not Available"

    except Exception as e:
        logging.error(f"Error in get_intraday_data for {ticker}: {e}")
        return None, used_yfinance

    return result, used_yfinance

def get_static_data(ticker, date, polygon_api_key, float_shares=None, outstanding_shares=None):
    """Fetch static data for a ticker on a specific date, relying on Polygon for splits.
    If float_shares or outstanding_shares are provided (from Excel), they will be used with priority over API values."""
    client = get_polygon_client()
    details = None
    for attempt in range(3):
        try:
            polygon_limiter.acquire_with_lock()
            details = client.get_ticker_details(ticker, date=date)
            break
        except Exception as e:
            if "429" in str(e):
                logging.warning(f"Rate limit hit for {ticker}, attempt {attempt + 1}")
                time.sleep(0.1)  # Add delay for rate limit
                continue
            elif "NOT_FOUND" in str(e) or "404" in str(e):
                logging.warning(f"Ticker {ticker} not found in Polygon for date {date}, continuing with other sources")
                details = None
                break
            else:
                logging.error(f"Error fetching ticker details for {ticker}: {e}")
                details = None
                break

    result = {
        "name": ticker,
        "float_shares": "Not Available",
        "float_shares_raw": None,
        "previous_close": "Not Available",
        "total_day_mcap": "Not Available",
        "shares_outstanding": "Not Available",
        "shares_outstanding_raw": None,
        "news": None
    }

    try:
        result["name"] = details.name if hasattr(details, "name") else ticker
    except Exception:
        pass

    try:
        current_date = datetime.now(EASTERN_TZ).date()
        target_dt = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=EASTERN_TZ)
        if target_dt.date() > current_date:
            logging.warning(f"Target date {date} is in the future for {ticker}")
            return result

        polygon_close = None
        for attempt in range(3):
            try:
                for days in range(1, 8):
                    check_date = (target_dt - timedelta(days=days)).strftime("%Y-%m-%d")
                    polygon_limiter.acquire_with_lock()
                    aggs = client.get_aggs(ticker, multiplier=1, timespan="day", from_=check_date, to=check_date)
                    if aggs and len(aggs) > 0:
                        close = aggs[0].close
                        if close > 0 and close < 50000:
                            polygon_close = close
                            logging.info(f"Fetched previous close for {ticker} on {check_date}: {polygon_close}")
                            break
                    else:
                        logging.warning(f"No Polygon data for {ticker} on {check_date}")
                if polygon_close:
                    break
            except Exception as e:
                if "429" in str(e):
                    logging.warning(f"Rate limit hit for {ticker} previous close, attempt {attempt + 1}")
                    time.sleep(0.1)  # Add delay for rate limit
                    continue
                logging.error(f"Error fetching previous close for {ticker}: {e}")
                break

        yf_close = None
        is_delisted = ticker in DELISTED_TICKERS and target_dt >= DELISTED_TICKERS[ticker]
        if is_delisted:
            logging.warning(f"Ticker {ticker} is delisted")
            return result
        if ticker in DELISTED_TICKERS:
            logging.info(f"Processing pre-delisting data for {ticker} on {date}")
        if polygon_close is None and not is_delisted:
            logging.warning(f"Polygon failed for {ticker}, trying yfinance")
            try:
                yf_ticker = yf.Ticker(ticker)
                polygon_limiter.acquire_with_lock()
                hist = yf_ticker.history(start=(target_dt - timedelta(days=5)).strftime("%Y-%m-%d"), end=(target_dt - timedelta(days=1)).strftime("%Y-%m-%d"))
                if not hist.empty:
                    yf_close = hist['Close'].iloc[-1]
                    logging.info(f"Using yfinance previous close for {ticker}: {yf_close}")
                else:
                    logging.warning(f"No yfinance data for {ticker}")
            except Exception as e:
                logging.warning(f"Error fetching yfinance for {ticker}: {e}")

        if polygon_close is not None and yf_close is not None:
            if abs(polygon_close - yf_close) / max(polygon_close, yf_close) > 0.2:
                logging.warning(f"Price discrepancy for {ticker}: Polygon({polygon_close}) vs yfinance({yf_close})")
                result["previous_close"] = polygon_close
            else:
                result["previous_close"] = polygon_close
        elif polygon_close:
            result["previous_close"] = polygon_close
        elif yf_close:
            result["previous_close"] = yf_close
        else:
            result["previous_close"] = "Not Available"
                
    except Exception as e:
        logging.error(f"Error fetching previous close for {ticker}: {e}")
    
    # Get float and shares data - Excel values take priority, then API hierarchy
    # If Excel values provided, use them with "Excel" source
    if outstanding_shares is not None and outstanding_shares > 0:
        shares_outstanding = outstanding_shares
        shares_source = "📊 Excel"
        logging.info(f"[EXCEL] Using outstanding shares from Excel for {ticker}: {shares_outstanding:,.0f}")
    else:
        shares_outstanding = None  # Start with no data, use API hierarchy
    shares_source = None
    
    if float_shares is not None and float_shares > 0:
        # Use Excel value (note: parameter name is same as variable, so we keep it)
        excel_float = float_shares
        float_shares = excel_float
        float_source = "📊 Excel"
        logging.info(f"[EXCEL] Using float shares from Excel for {ticker}: {float_shares:,.0f}")
    else:
        float_shares = None  # Start with no data, use API hierarchy
    float_source = None
    
    # NEW: Outstanding shares hierarchy (current priority, YFinance fallback)
    if shares_outstanding is None:
        try:
            logging.info(f"[POLYGON] Trying current ticker details for {ticker}")
            polygon_data = get_polygon_ticker_details(ticker, date)
            
            if polygon_data and polygon_data.get('outstanding_shares'):
                shares_outstanding = polygon_data['outstanding_shares']
                filing_date = polygon_data.get('filing_date', 'Unknown')
                shares_source = f"🟣 Polygon Current Data ({filing_date})"
                logging.info(f"[POLYGON] Found outstanding shares: {shares_outstanding:,} for {ticker}")
            else:
                logging.warning(f"[POLYGON] No outstanding shares found for {ticker}")
        except Exception as e:
            logging.error(f"[POLYGON] Error fetching ticker details for {ticker}: {e}")
    
        # NEW: Float fallback hierarchy (Phase 1 - Fast APIs only)
    # Only fetch from APIs if Excel value not provided
    if float_shares is None:
        # Phase 1: Skip Wayback Machine - it's handled in Phase 2
        pass
    
    # Commented out - using Excel outstanding shares data only
    # # 2. If outstanding shares not found yet, try SEC API main endpoint
    # if shares_outstanding is None and SEC_WORKING:
    #     try:
    #         logging.info(f"[SEC] Trying main endpoint for outstanding shares: {ticker}")
    #         sec_data = get_sec_float_data(ticker, date)
    #         
    #         if sec_data and sec_data.get("shares_outstanding") is not None:
    #             shares_outstanding = sec_data["shares_outstanding"]
    #             timing = sec_data.get("timing", "unknown")
    #             shares_source = f"📋 SEC API Main ({timing})"
    #             logging.info(f"[SEC] Found outstanding shares from main endpoint: {shares_outstanding:,} for {ticker}")
    #         else:
    #             logging.warning(f"[SEC] No outstanding shares from main endpoint for {ticker}")
    #     except Exception as e:
    #         logging.error(f"[SEC] Error in main endpoint for outstanding shares: {ticker}: {e}")
    
    # Commented out - using Excel float data only
    # 3. If float shares not found yet, try SEC API main float endpoint
    # if float_shares is None and SEC_WORKING:
    #     try:
    #         logging.info(f"[SEC] Trying main float endpoint as fallback for {ticker}")
    #         sec_data = get_sec_float_data(ticker, date)
    #         
    #         if sec_data and sec_data.get("float_shares") is not None:
    #             float_shares = sec_data["float_shares"]
    #             # Set source immediately when data is retrieved
    #             timing = sec_data.get("timing", "unknown")
    #             if timing in ["before", "after", "current"]:
    #                 float_source = f"📋 SEC API Main ({timing.capitalize()})"
    #             else:
    #                 float_source = "📋 SEC API Main"
    #             logging.info(f"[SEC] Found float shares from main endpoint: {float_shares:,} for {ticker}")
    #         else:
    #             logging.warning(f"[SEC] No float data from main endpoint for {ticker}")
    #     except Exception as e:
    #         logging.error(f"[SEC] Error in main endpoint fallback for {ticker}: {e}")
    
    # 4. If SEC main failed, try SEC query/extract API
    # if float_shares is None and SEC_WORKING:
    #     try:
    #         logging.info(f"[SEC] Trying query/extract API as fallback for {ticker}")
    #         fallback_data = get_sec_fallback_data(ticker, date, sec_data)
    #         
    #         if fallback_data and fallback_data.get("float_shares") is not None:
    #             float_shares = fallback_data["float_shares"]
    #             # Set source immediately when data is retrieved
    #             timing = fallback_data.get("timing", "unknown")
    #             if timing in ["before", "after", "current"]:
    #                 float_source = f"🔍 SEC API Extract ({timing.capitalize()})"
    #             else:
    #                 float_source = "🔍 SEC API Extract"
    #             logging.info(f"[SEC] Found float shares from extract API: {float_shares:,} for {ticker}")
    #             
    #             # Also get shares outstanding if not found yet
    #             if shares_outstanding is None and fallback_data.get("shares_outstanding") is not None:
    #                 shares_outstanding = fallback_data["shares_outstanding"]
    #                 shares_source = "🔍 SEC API Extract (fallback)"
    #                 logging.info(f"[SEC] Found outstanding shares from extract API: {shares_outstanding:,} for {ticker}")
    #         else:
    #             logging.warning(f"[SEC] No float data from extract API for {ticker}")
    #     except Exception as e:
    #         logging.error(f"[SEC] Error in extract API fallback for {ticker}: {e}")
    
    # 5. Final fallback: YFinance current float
    if float_shares is None:
        try:
            logging.info(f"[YFINANCE] Trying current float as final fallback for {ticker}")
            yf_ticker = yf.Ticker(ticker)
            yf_info = yf_ticker.info
            
            yf_float = yf_info.get("floatShares")
            if yf_float and yf_float > 0:
                float_shares = yf_float
                # Set source immediately when data is retrieved
                float_source = "📊 YFinance Current"
                logging.info(f"[YFINANCE] Found current float shares: {float_shares:,} for {ticker}")
            else:
                logging.warning(f"[YFINANCE] No float data found for {ticker}")
        except Exception as e:
            logging.error(f"[YFINANCE] Error in YFinance fallback for {ticker}: {e}")
    

    
    # 6. Final fallback: YFinance current outstanding shares
    if shares_outstanding is None:
        try:
            logging.info(f"[YFINANCE] Trying current outstanding shares as final fallback for {ticker}")
            yf_ticker = yf.Ticker(ticker)
            yf_info = yf_ticker.info
            
            yf_outstanding = yf_info.get("sharesOutstanding")
            if yf_outstanding and yf_outstanding > 0:
                shares_outstanding = yf_outstanding
                shares_source = "📊 YFinance Current"
                logging.info(f"[YFINANCE] Found current outstanding shares: {shares_outstanding:,} for {ticker}")
            else:
                logging.warning(f"[YFINANCE] No outstanding shares data found for {ticker}")
        except Exception as e:
            logging.error(f"[YFINANCE] Error in YFinance outstanding shares fallback for {ticker}: {e}")
    
    # Set the results
    if shares_outstanding is not None:
        result["shares_outstanding"] = format_number(shares_outstanding)
        result["shares_outstanding_raw"] = shares_outstanding
    else:
        result["shares_outstanding"] = "Not Available"
        result["shares_outstanding_raw"] = None
    
    # Set float shares results
    if float_shares is not None:
        result["float_shares"] = format_number(float_shares)
        result["float_shares_raw"] = float_shares
    else:
        result["float_shares"] = "Not Available"
        result["float_shares_raw"] = None
    
    # Add float source information
    if float_source:
        result["float_source"] = float_source
    else:
        result["float_source"] = "Unknown"
    
    # Add shares source information
    if shares_source:
        result["shares_source"] = shares_source
    else:
        result["shares_source"] = "Unknown"

    # Get dilution data
    # Commented out - using Excel data only
    # try:
    #     logging.info(f"[DILUTION] Collecting dilution data for {ticker} as of {date}")
    #     dilution_data = get_dilution_data(ticker, date)
    #     
    #     # Calculate dilution percentages against outstanding shares and float
    #     dilution_vs_outstanding_pct = 0.0
    #     dilution_vs_float_pct = 0.0
    #     float_dilution_risk = "Low"
    #     
    #     if shares_outstanding and result["shares_outstanding_raw"] is not None and result["shares_outstanding_raw"] > 0:
    #         if dilution_data['total_active_dilution'] > 0:
    #             dilution_vs_outstanding_pct = (dilution_data['total_active_dilution'] / result["shares_outstanding_raw"]) * 100
    #     
    #     if float_shares and result["float_shares_raw"] is not None and result["float_shares_raw"] > 0:
    #         if dilution_data['total_active_dilution'] > 0:
    #             dilution_vs_float_pct = (dilution_data['total_active_dilution'] / result["float_shares_raw"]) * 100
    #     
    #     # Determine float dilution risk
    #     if dilution_vs_float_pct >= 50:
    #         float_dilution_risk = "Extreme"
    #     elif dilution_vs_float_pct >= 25:
    #         float_dilution_risk = "High"
    #     elif dilution_vs_float_pct >= 10:
    #         float_dilution_risk = "Medium"
    #     elif dilution_vs_float_pct >= 5:
    #         float_dilution_risk = "Moderate"
    #     else:
    #         float_dilution_risk = "Low"
    #     
    #     # Add dilution data to result
    #     result.update({
    #         "active_warrant_shares": dilution_data['active_warrant_shares'],
    #         "active_convertible_shares": dilution_data['active_convertible_shares'],
    #         "total_active_dilution": dilution_data['total_active_dilution'],
    #         "warrant_count": dilution_data['warrant_count'],
    #         "convertible_count": dilution_data['convertible_count'],
    #         "dilution_vs_outstanding_pct": round(dilution_vs_outstanding_pct, 2),
    #         "dilution_vs_float_pct": round(dilution_vs_float_pct, 2),
    #         "float_dilution_risk": float_dilution_risk
    #     })
    #     
    #     logging.info(f"[DILUTION] Completed dilution analysis for {ticker}: {dilution_data['total_active_dilution']:,} total dilution, {dilution_vs_float_pct:.2f}% vs float")
    #     
    # except Exception as e:
    #     logging.error(f"[DILUTION] Error collecting dilution data for {ticker}: {e}")
    #     # Add default dilution values
    #     result.update({
    #         "active_warrant_shares": 0,
    #         "active_convertible_shares": 0,
    #         "total_active_dilution": 0,
    #         "warrant_count": 0,
    #         "convertible_count": 0,
    #         "dilution_vs_outstanding_pct": 0.0,
    #         "dilution_vs_float_pct": 0.0,
    #         "float_dilution_risk": "Low"
    #     })
    
    # Add default dilution values (commented out functionality)
    result.update({
        "active_warrant_shares": 0,
        "active_convertible_shares": 0,
        "total_active_dilution": 0,
        "warrant_count": 0,
        "convertible_count": 0,
        "dilution_vs_outstanding_pct": 0.0,
        "dilution_vs_float_pct": 0.0,
        "float_dilution_risk": "Low"
    })

    if shares_outstanding and result["shares_outstanding_raw"] is not None:
        intraday_data, _ = get_intraday_data(ticker, date, result["previous_close"], result["shares_outstanding_raw"], result["float_shares_raw"])
        if intraday_data and intraday_data.get("open_close") != "Not Available":
            try:
                open_close = float(intraday_data["open_close"])
                result["total_day_mcap"] = format_number(result["shares_outstanding_raw"] * open_close)
            except Exception as e:
                logging.warning(f"Error calculating market cap for {ticker}: {e}")



    return result

def save_to_excel_with_colors(df, output_file):
    """Save DataFrame to Excel with color coding based on data sources."""
    try:
        # Replace .csv extension with .xlsx
        excel_file = output_file.replace('.csv', '.xlsx')
        
        # Create Excel writer
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name='Stock Data', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Stock Data']
            
            # Define color fills
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
            yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Gold
            red_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')  # Light red
            
            # Find column indices
            headers = list(df.columns)
            float_idx = headers.index('Float') + 1 if 'Float' in headers else None
            shares_idx = headers.index('Outstanding Shares') + 1 if 'Outstanding Shares' in headers else None
            timing_idx = headers.index('timing') + 1 if 'timing' in headers else None
            
            # Apply color coding based on timing field
            for row in range(2, len(df) + 2):  # Start from row 2 (after headers)
                # Get timing info from the DataFrame
                row_data = df.iloc[row-2]  # Convert to 0-based index
                timing = row_data.get('timing', 'unknown') if 'timing' in df.columns else 'unknown'
                
                # Color Float column
                if float_idx:
                    float_cell = worksheet.cell(row=row, column=float_idx)
                    float_value = float_cell.value
                    
                    if float_value and float_value != "Not Available":
                        if timing == "before":
                            # Data from before - green
                            float_cell.fill = green_fill
                        elif timing == "after":
                            # Data from after - yellow
                            float_cell.fill = yellow_fill
                        # Current data - no color
                    elif float_value == "Not Available":
                        # No data - red
                        float_cell.fill = red_fill
                
                # Color Outstanding Shares column
                if shares_idx:
                    shares_cell = worksheet.cell(row=row, column=shares_idx)
                    shares_value = shares_cell.value
                    
                    if shares_value and shares_value != "Not Available":
                        if timing == "before":
                            # Data from before - green
                            shares_cell.fill = green_fill
                        elif timing == "after":
                            # Data from after - yellow
                            shares_cell.fill = yellow_fill
                        # Current data - no color
                    elif shares_value == "Not Available":
                        # No data - red
                        shares_cell.fill = red_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Create legend sheet
            legend_df = pd.DataFrame({
                'Color': ['Light Green', 'Yellow/Gold', 'Light Red', 'No Color'],
                'Meaning': ['Historical Data (Before Date)', 'Future Data (After Date)', 'No Data Available', 'Current Data'],
                'Description': [
                    'SEC filing data from before the specified trading date - most accurate for historical analysis',
                    'SEC filing data from after the specified trading date - used when no prior data exists',
                    'No SEC filing data found for this ticker',
                    'Current SEC filing data (no specific date requested)'
                ]
            })
            
            legend_df.to_excel(writer, sheet_name='Legend', index=False)
            
            # Format legend sheet
            legend_sheet = writer.sheets['Legend']
            
            # Apply colors to legend
            legend_sheet.cell(row=2, column=1).fill = green_fill
            legend_sheet.cell(row=3, column=1).fill = yellow_fill
            legend_sheet.cell(row=4, column=1).fill = red_fill
            
            # Auto-adjust legend columns
            for column in legend_sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                legend_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        logging.info(f"Saved Excel file with color coding: {excel_file}")
        return excel_file
        
    except Exception as e:
        logging.error(f"Error saving Excel file: {e}")
        return None

def upload_to_s3(file_path):
    """Upload file to S3 bucket."""
    try:
        s3_client = boto3.client('s3')
        s3_client.upload_file(file_path, S3_BUCKET, os.path.basename(file_path))
        logging.info(f"Uploaded {file_path} to S3 bucket {S3_BUCKET}")
    except ClientError as e:
        logging.error(f"Error uploading {file_path} to S3: {e}")

def get_nasdaq_tickers(test_mode=False):
    """Fetch all NASDAQ tickers using cached file or Polygon.io."""
    if test_mode:
        # Hardcoded test tickers for development/testing
        test_tickers = [
            "AAPL", "MSFT", "TSLA",  # Just 3 tickers for quick testing
        ]
        logging.info(f"Using test mode with {len(test_tickers)} hardcoded tickers")
        return test_tickers
    
    # Try to load from cached file first
    cache_file = "nasdaq_tickers_cache.txt"
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r') as f:
                tickers = [line.strip() for line in f if line.strip()]
            logging.info(f"Loaded {len(tickers)} NASDAQ tickers from cache")
            return tickers
        except Exception as e:
            logging.warning(f"Error reading cache file: {e}")
    
    # Fallback to API call if cache doesn't exist
    client = get_polygon_client()
    tickers = []
    try:
        logging.info("Fetching NASDAQ tickers from Polygon.io (this may take 30+ seconds)...")
        polygon_limiter.acquire()
        for ticker in client.list_tickers(market="stocks", exchange="XNAS"):
            tickers.append(ticker.ticker)
        
        # Cache the results for future use
        try:
            with open(cache_file, 'w') as f:
                for ticker in tickers:
                    f.write(f"{ticker}\n")
            logging.info(f"Cached {len(tickers)} NASDAQ tickers to {cache_file}")
        except Exception as e:
            logging.warning(f"Error caching tickers: {e}")
            
        logging.info(f"Fetched {len(tickers)} NASDAQ tickers")
    except Exception as e:
        logging.error(f"Error fetching NASDAQ tickers: {e}")
        return []
    return tickers
def quick_pre_market_check(ticker, date, client):
    """Quick pre-market gain check with price, volume, and gains checks."""
    try:
        polygon_limiter.acquire()
        
        # Add retry logic for connection issues
        for attempt in range(3):
            try:
                aggs = client.get_aggs(
                    ticker,
                    multiplier=1,
                    timespan="minute", 
                    from_=date,
                    to=date,
                    adjusted=True
                )
                break
            except Exception as e:
                if "429" in str(e) or "Connection pool is full" in str(e):
                    if attempt < 2:
                        time.sleep(0.5)  # Wait before retry
                        continue
                raise e
        
        if not aggs or len(aggs) == 0:
            return None
            
        # Convert to DataFrame for analysis
        data = pd.DataFrame([
            {
                "timestamp": getattr(agg, 'timestamp', None),
                "open": getattr(agg, 'open', None),
                "high": getattr(agg, 'high', None),
                "low": getattr(agg, 'low', None),
                "close": getattr(agg, 'close', None),
                "volume": getattr(agg, 'volume', None)
            }
            for agg in aggs if hasattr(agg, 'timestamp')
        ])
        
        if data.empty:
            return None
            
        # Basic pre-market filtering
        data["timestamp"] = pd.to_datetime(data["timestamp"], unit="ms", errors="coerce")
        data["datetime"] = data["timestamp"].dt.tz_localize("UTC").dt.tz_convert(EASTERN_TZ)
        data.set_index("datetime", inplace=True)
        
        pre_market_start = pd.Timestamp(f"{date} 04:00:00").tz_localize(EASTERN_TZ)
        pre_market_end = pd.Timestamp(f"{date} 09:29:59").tz_localize(EASTERN_TZ)
        pre_market = data[pre_market_start:pre_market_end].copy()
        
        if pre_market.empty:
            return None
            
        total_volume = pre_market["volume"].sum()
        pm_high = pre_market["high"].max()
        
        # Quick volume check (50K+)
        if total_volume < 50000:
            return None
            
        # Quick price check (>$0.50)
        if pm_high <= 0.50:
            return None
            
        # Get previous close for gains calculation (weekend-aware)
        try:
            target_dt = datetime.strptime(date, "%Y-%m-%d")
            previous_close = None
            
            # Check up to 7 days back to find the last trading day
            for days in range(1, 8):
                check_date = (target_dt - timedelta(days=days)).strftime("%Y-%m-%d")
                polygon_limiter.acquire()
                prev_aggs = client.get_aggs(
                    ticker,
                    multiplier=1,
                    timespan="day",
                    from_=check_date,
                    to=check_date,
                    adjusted=True
                )
                if prev_aggs and len(prev_aggs) > 0:
                    close = prev_aggs[0].close
                    if close > 0 and close < 50000:
                        previous_close = close
                        break
                        
            if not previous_close:
                return None
        except:
            return None
            
        # Calculate percent gain
        if previous_close and previous_close > 0:
            percent_gain = safe_percentage_change(pm_high, previous_close)
            if percent_gain == "Not Available" or percent_gain < 10:
                return None
        else:
            return None
            
        return {
            "ticker": ticker,
            "total_volume": total_volume,
            "pm_high": pm_high,
            "previous_close": previous_close,
            "percent_gain": percent_gain,
            "pre_market_data": pre_market
        }
        
    except Exception as e:
        logging.warning(f"Quick check failed for {ticker}: {e}")
        return None

def parallel_float_check(ticker):
    """Parallel float check using YFinance for Stage 2 filtering."""
    try:
        # Add small delay to avoid YFinance rate limiting
        time.sleep(0.01)  # Reduced delay for parallel processing
        yf_ticker = yf.Ticker(ticker)
        float_shares = yf_ticker.info.get("floatShares")
        if float_shares and float_shares > 0:
            if float_shares >= 20_000_000:
                logging.info(colorize_yfinance(f"Stage 2: {ticker} filtered out - float too high: {format_number(float_shares)}"))
                return {"ticker": ticker, "float_shares": float_shares, "passed": False}
            else:
                logging.info(colorize_yfinance(f"Stage 2: {ticker} passed float check: {format_number(float_shares)}"))
                return {"ticker": ticker, "float_shares": float_shares, "passed": True}
        else:
            # If no float data, filter it out
            logging.info(colorize_yfinance(f"Stage 2: {ticker} filtered out - no float data available"))
            return {"ticker": ticker, "float_shares": None, "passed": False}
    except Exception as e:
        logging.warning(colorize_yfinance(f"Stage 2: Error checking float for {ticker}: {e}"))
        return {"ticker": ticker, "float_shares": None, "passed": False}  # Filter out on error

def optimized_filter_pre_market_gainers(tickers, date):
    """Optimized two-stage filtering to reduce API calls."""
    client = get_polygon_client()
    stage1_results = []
    total_tickers = len(tickers)
    
    logging.info(f"Stage 1: Quick pre-market checks (price, volume, gains) for {total_tickers} tickers")
    
    # Stage 1: Quick pre-market checks (price, volume, gains) - 25 workers
    with ThreadPoolExecutor(max_workers=25) as executor:
        future_to_ticker = {executor.submit(quick_pre_market_check, ticker, date, client): ticker for ticker in tickers}
        for i, future in enumerate(as_completed(future_to_ticker), 1):
            ticker = future_to_ticker[future]
            try:
                result = future.result()
                if result:
                    stage1_results.append(result)
                if i % 100 == 0:
                    logging.info(f"Stage 1: Completed {i}/{total_tickers} tickers, {len(stage1_results)} qualifiers")
            except Exception as e:
                logging.warning(f"Stage 1 error for {ticker}: {e}")
    
    logging.info(f"Stage 1 complete: {len(stage1_results)} tickers passed quick checks")
    
    # Stage 2: Parallel float checks and data collection
    logging.info(f"Stage 2: Processing {len(stage1_results)} tickers with parallel float checks")
    
    # Parallel float checks for all Stage 1 qualifiers
    float_results = {}
    with ThreadPoolExecutor(max_workers=22) as executor:
        future_to_ticker = {executor.submit(parallel_float_check, result["ticker"]): result["ticker"] for result in stage1_results}
        for future in as_completed(future_to_ticker):
            ticker = future_to_ticker[future]
            try:
                float_result = future.result()
                float_results[ticker] = float_result
            except Exception as e:
                logging.warning(f"Float check error for {ticker}: {e}")
                float_results[ticker] = {"ticker": ticker, "float_shares": None, "passed": True}
    
    # Filter out tickers that failed float check
    qualified_tickers = []
    for result in stage1_results:
        ticker = result["ticker"]
        float_result = float_results.get(ticker, {"passed": True})
        if float_result["passed"]:
            qualified_tickers.append(result)
    
    logging.info(f"Stage 2: {len(qualified_tickers)} tickers passed float checks")
    
    # Return only the basic qualifying ticker info - comprehensive data collection happens in single-phase
    final_results = []
    for result in qualified_tickers:
        ticker = result["ticker"]
        float_result = float_results.get(ticker, {})
        float_shares = float_result.get("float_shares")
        
        # Just return basic info - comprehensive data collection will happen in single-phase
        final_results.append({
            "ticker": ticker,
            "percent_gain": result["percent_gain"],
            "total_volume": result["total_volume"],
            "pm_high": result["pm_high"],
            "previous_close": result["previous_close"],
            "float_shares": float_shares
        })
        logging.info(f"Stage 2: {ticker} qualifies - {result['percent_gain']:.1f}% gain, {result['total_volume']} volume")
    
    logging.info(f"Final results: {len(final_results)} tickers passed all filters")
    return final_results

def process_phase1_fast_apis(ticker: str, date: str, float_shares: Optional[float] = None, outstanding_shares: Optional[float] = None) -> Optional[Dict[str, Any]]:
    """Phase 1: Process fast APIs (Polygon, YFinance, SEC) - 30 workers.
    If float_shares or outstanding_shares are provided (from Excel), they will be used with priority over API values."""
    logging.info(f"[PHASE 1] Processing fast APIs for {ticker} on {date}")
    
    if not is_trading_day(date):
        logging.info(f"{date} is not a trading day, skipping {ticker}")
        return None
    
    # Retry mechanism for connection pool issues
    max_retries = 2
    for attempt in range(max_retries):
        try:
            # Get static data (Polygon API, YFinance, SEC) - pass Excel values if available
            static_data = get_static_data(ticker, date, POLYGON_API_KEY, float_shares, outstanding_shares)
            
            # Get intraday data (Polygon API)
            intraday_data, used_yfinance = get_intraday_data(ticker, date, static_data["previous_close"], static_data["shares_outstanding_raw"], static_data["float_shares_raw"])
            
            return {
                'ticker': ticker,
                'date': date,
                'static_data': static_data,
                'intraday_data': intraday_data,
                'used_yfinance': used_yfinance
            }
        except Exception as e:
            if "Connection pool is full" in str(e) and attempt < max_retries - 1:
                logging.warning(f"[PHASE 1] Connection pool exhausted for {ticker} on {date}. Attempt {attempt + 1}/{max_retries}. Cleaning up...")
                cleanup_polygon_client()  # Force cleanup
                time.sleep(2)  # Longer pause between retries
                continue
            else:
                logging.error(f"[PHASE 1] Error processing {ticker} on {date}: {e}")
                return None

def process_phase2_wayback(ticker, date, phase1_result):
    """Phase 2: Process Wayback Machine (conservative) - 10 workers."""
    if not phase1_result:
        logging.info(f"[PHASE 2] Skipping {ticker} on {date} - no Phase 1 data")
        return None

    logging.info(f"[PHASE 2] Processing Wayback for {ticker} on {date}")
    
    try:
        # Get Wayback data (historical float)
        wayback_data = get_wayback_float_data(ticker, date)
        
        return {
            'ticker': ticker,
            'date': date,
            'wayback_data': wayback_data
        }
    except Exception as e:
        logging.error(f"[PHASE 2] Error processing {ticker} on {date}: {e}")
        return None
            
def process_phase3_final_assembly(ticker, date, phase1_result, phase2_result, news_data, news_source="Unknown", excel_row_index=None):
    """Phase 3: Final assembly and calculations - 30 workers."""
    if not phase1_result:
        logging.info(f"[PHASE 3] Skipping {ticker} on {date} - no Phase 1 data")
        return None

    logging.info(f"[PHASE 3] Final assembly for {ticker} on {date}")
    
    try:
        static_data = phase1_result['static_data']
        intraday_data = phase1_result['intraday_data']
        wayback_data = phase2_result.get('wayback_data') if phase2_result else None
        
        # Create the final row with all data
        row = create_final_row(ticker, date, static_data, intraday_data, wayback_data, news_data, news_source, excel_row_index)
        return row
    except Exception as e:
        logging.error(f"[PHASE 3] Error processing {ticker} on {date}: {e}")
        return None
    
def get_wayback_float_data(ticker, date):
    """Get float data from Wayback Machine for a ticker-date pair."""
    
    # Hard-coded skip for problematic tickers
    problematic_tickers = [
        "NEHC", "HWH", "INZY", "SAIH", "YJ", "RYET", "HWH", "NEHC", "ULY", "NAAS", "ZEO", "HTCO", "FEAM", "FMTO", "TVGN", "STAI", "RANI", "XLO", "SLDB", "PYXS", "TRVI", "ALLR", "LFWD", "BDTX", "OPTN", "CNTM", "SNOA", "SLXN", "VSTE", "KDLY", "SAG", "NXU", "TOPW", "DWSN", "AMOD", "TDTH", "GHRS", "SOPA", "EVAX", "GCTK", "QNTM", "KAPA", "SFHG", "LIPO", "PLRX", "WOK", "NAAS", "STEC", "ORIS", "BREA", "PBM", "CNSP", "TOI", "MBRX", "NIXX", "VMAR", "AMST", "BTAI", "VRME", "RPID", "BTOG", "HMR", "BTAI", "NIXX", "THAR", "YHC", "SCNX"
    ]
    logging.info(f"[WAYBACK] Checking ticker {ticker} against problematic list: {problematic_tickers}")
    if ticker in problematic_tickers:
        logging.info(f"[WAYBACK] SKIPPING {ticker} - known problematic ticker, going straight to fallback")
        return None
    else:
        logging.info(f"[WAYBACK] {ticker} not in problematic list, proceeding with Wayback")
    
    def wayback_worker():
        """Worker function to run Wayback scraping in a separate thread."""
        try:
            # Create WaybackFloatScraper instance to access Wayback functionality
            wayback_scraper = WaybackFloatScraper()
            wayback_result = wayback_scraper.get_float_from_wayback(ticker, date)
            
            if wayback_result and wayback_result.get('float'):
                return {
                    'float': wayback_result['float'],
                    'source': 'Wayback Machine',
                    'date': wayback_result.get('date', date)
                }
            else:
                return None
        except Exception as e:
            # Check if it's a timeout error and skip instead of retrying
            if "ReadTimeoutError" in str(e) or "Read timed out" in str(e) or "Timeout" in str(e):
                logging.info(f"Wayback timeout for {ticker} on {date} - skipping to fallback process")
                return None
            else:
                logging.error(f"Error getting Wayback data for {ticker} on {date}: {e}")
                return None
    
    try:
        # Use threading to implement 1-minute timeout
        import threading
        result = [None]
        exception = [None]
        
        def target():
            try:
                result[0] = wayback_worker()
            except Exception as e:
                exception[0] = e
        
        thread = threading.Thread(target=target)
        thread.daemon = True
        thread.start()
        thread.join(timeout=90)  # 90 seconds timeout
        
        if thread.is_alive():
            logging.warning(f"[WAYBACK] Timeout after 90 seconds for {ticker} on {date} - skipping to fallback")
            return None
        
        if exception[0]:
            raise exception[0]
        
        return result[0]
        
    except Exception as e:
        logging.error(f"Error in Wayback processing for {ticker} on {date}: {e}")
        return None

def create_failed_row(ticker, date, excel_row_index=None, error_message="Failed"):
    """Create a failed row with all required columns set to 'Not Available' to maintain Excel order."""
    return {
        "Ticker": ticker,
        "Date": date if date else "Invalid Date",
        "Excel_Row_Index": excel_row_index if excel_row_index is not None else 0,
        "News": "No news",
        "News Source": "Unknown",
        "Total Day MCAP": "Not Available",
        "Float": "Not Available",
        "Float Source": "Not Available",
        "Outstanding Shares": "Not Available",
        "Shares Source": "Not Available",
        "Total Day Vol": "Not Available",
        "Previous Close": "Not Available",
        "PM Open": "Not Available",
        "PM High": "Not Available",
        "PM High Time": "Not Available",
        "PM Low After High": "Not Available",
        "PM Low After High Time": "Not Available",
        "PM Low 15m After High": "Not Available",
        "PM Low 15m After High Time": "Not Available",
        "PM Low 30m After High": "Not Available",
        "PM Low 30m After High Time": "Not Available",
        "PM Low 1h After High": "Not Available",
        "PM Low 1h After High Time": "Not Available",
        "PM Low": "Not Available",
        "PM Low Time": "Not Available",
        "PM Close": "Not Available",
        "Reclaim VWAP PM": "Not Available",
        "VWAP Breakdown Time PM": "Not Available",
        "VWAP Deviation % PM High": "Not Available",
        "Vol @ PM High": "Not Available",
        "Vol @ PM Low": "Not Available",
        "Open": "Not Available",
        "Open High": "Not Available",
        "Open High Time": "Not Available",
        "Open Low After High": "Not Available",
        "Open Low After High Time": "Not Available",
        "Open Low 15m After High": "Not Available",
        "Open Low 15m After High Time": "Not Available",
        "Open Low 30m After High": "Not Available",
        "Open Low 30m After High Time": "Not Available",
        "Open Low 1h After High": "Not Available",
        "Open Low 1h After High Time": "Not Available",
        "Open Low": "Not Available",
        "Open Low Time": "Not Available",
        "Reclaim VWAP Open": "Not Available",
        "VWAP Breakdown Time Open": "Not Available",
        "VWAP Deviation % Open High": "Not Available",
        "Open Close": "Not Available",
        "Vol @ Open High": "Not Available",
        "Vol @ Open Low": "Not Available",
        "AH Open": "Not Available",
        "AH High": "Not Available",
        "AH High Time": "Not Available",
        "AH Low": "Not Available",
        "AH Low Time": "Not Available",
        "AH Close": "Not Available",
        "Reclaim VWAP AH": "Not Available",
        "VWAP Breakdown Time AH": "Not Available",
        "VWAP Deviation % AH High": "Not Available",
        "Vol @ AH High": "Not Available",
        "Vol @ 6AM": "Not Available",
        "Vol @ 7AM": "Not Available",
        "Vol @ 8AM": "Not Available",
        "Vol @ 9AM": "Not Available",
        "Vol @ 10AM": "Not Available",
        "Vol @ 11AM": "Not Available",
        "Vol @ 12PM": "Not Available",
        "% Gain PM": "Not Available",
        "% Gain Open": "Not Available",
        "% Gain AH": "Not Available",
        "Market Cap at PM High": "Not Available",
        "Market Cap at Open High": "Not Available",
        "Market Cap 10AM": "Not Available",
        "Market Cap 11AM": "Not Available",
        "Market Cap 12PM": "Not Available",
        "Market Cap AH": "Not Available",
        "PM Short P&L": "Not Available",
        "PM Short P&L %": "Not Available",
        "Open Short P&L": "Not Available",
        "Open Short P&L %": "Not Available",
        "PM to Open Gap": "Not Available",
        "Open to 12pm Range": "Not Available",
        "PM Volume Spike": "Not Available",
        "Open Volume Spike": "Not Available",
        "PM Momentum": "Not Available",
        "Open Momentum": "Not Available",
        "PM Reversal Signal": "Not Available",
        "Open Reversal Signal": "Not Available",
        "PM RSI at High 1min": "Not Available",
        "PM RSI at High 2min": "Not Available",
        "PM RSI at High 5min": "Not Available",
        "PM RSI at High 1hour": "Not Available",
        "Open RSI at High 1min": "Not Available",
        "Open RSI at High 2min": "Not Available",
        "Open RSI at High 5min": "Not Available",
        "Open RSI at High 1hour": "Not Available",
        "MACD Line PM High": "Not Available",
        "MACD Signal PM High": "Not Available",
        "MACD Histogram PM High": "Not Available",
        "MACD Line Open High": "Not Available",
        "MACD Signal Open High": "Not Available",
        "MACD Histogram Open High": "Not Available",
        "Short Float %": "Not Available",
        "Short Interest % of Outstanding": "Not Available",
        "Days to Cover": "Not Available",
        "Short Interest": "Not Available",
        "BB %B at Open High": "Not Available",
        "BB %B at PM High": "Not Available",
        "Short Volume": "Not Available",
        "Short Volume Ratio": "Not Available",
        "Price 10AM": "Not Available",
        "Price 10:30AM": "Not Available",
        "Price 11AM": "Not Available",
        "Price 11:30AM": "Not Available",
        "Price 12PM": "Not Available",
        "Active Warrant Shares": "Not Available",
        "Active Convertible Shares": "Not Available",
        "Total Active Dilution": "Not Available",
        "Warrant Count": "Not Available",
        "Convertible Count": "Not Available",
        "Dilution vs Outstanding %": "Not Available",
        "Dilution vs Float %": "Not Available",
        "Float Dilution Risk": "Not Available",
        "Float Rotation 5AM": "Not Available",
        "Float Rotation 6AM": "Not Available",
        "Float Rotation 7AM": "Not Available",
        "Float Rotation 8AM": "Not Available",
        "Float Rotation 9AM": "Not Available",
        "Float Rotation 9:30AM": "Not Available",
        "Float Rotation 10AM": "Not Available",
        "Float Rotation 11AM": "Not Available",
        "Float Rotation 12PM": "Not Available",
        "Float Rotation 1PM": "Not Available",
        "Float Rotation 4PM": "Not Available",
        "RVOL PM High": "Not Available",
        "RVOL Open High": "Not Available",
        "RVOL 7AM": "Not Available",
        "RVOL 8AM": "Not Available",
        "RVOL 9AM": "Not Available",
        "RVOL 9:30AM": "Not Available",
        "RVOL 10AM": "Not Available",
        "RVOL 11AM": "Not Available",
        "RVOL 12PM": "Not Available",
        "RVOL 1PM": "Not Available",
        "Used YFinance": "Not Available",
        "Error": error_message
    }

def create_final_row(ticker, date, static_data, intraday_data, wayback_data, news_data, news_source="Unknown", excel_row_index=None):
    """Create the final row with all assembled data."""
    # Use static data for float shares
    float_shares = static_data.get("float_shares")
    
    # Use API float data source
    float_source = static_data.get("float_source", "Unknown")

    row = {
        "Ticker": ticker,
        "Date": date,
        "Excel_Row_Index": excel_row_index if excel_row_index is not None else 0,
        "News": news_data,
        "News Source": news_source,
        "Total Day MCAP": static_data["total_day_mcap"],
        "Float": float_shares,
        "Float Source": float_source,
        "Outstanding Shares": static_data["shares_outstanding"],
        "Shares Source": static_data.get("shares_source", "Unknown"),
        "Total Day Vol": intraday_data["total_day_volume"],
        "Previous Close": static_data["previous_close"],
        "PM Open": intraday_data["pm_open"],
        "PM High": intraday_data["pm_high"],
        "PM High Time": intraday_data["pm_high_time"],
        "PM Low After High": intraday_data["pm_low_after_high"],
        "PM Low After High Time": intraday_data["pm_low_after_high_time"],
        "PM Low 15m After High": intraday_data["pm_low_15m_after_high"],
        "PM Low 15m After High Time": intraday_data["pm_low_15m_after_high_time"],
        "PM Low 30m After High": intraday_data["pm_low_30m_after_high"],
        "PM Low 30m After High Time": intraday_data["pm_low_30m_after_high_time"],
        "PM Low 1h After High": intraday_data["pm_low_1h_after_high"],
        "PM Low 1h After High Time": intraday_data["pm_low_1h_after_high_time"],
        "PM Low": intraday_data["pm_low"],
        "PM Low Time": intraday_data["pm_low_time"],
        "PM Close": intraday_data["pm_close"],
        "Reclaim VWAP PM": intraday_data["reclaim_vwap_pm"],
        "VWAP Breakdown Time PM": intraday_data["vwap_breakdown_time_pm"],
        "VWAP Deviation % PM High": intraday_data["vwap_deviation_pm_high"],
        "Vol @ PM High": intraday_data["vol_pm_high"],
        "Vol @ PM Low": intraday_data["vol_pm_low"],
        "Open": intraday_data["open"],
        "Open High": intraday_data["open_high"],
        "Open High Time": intraday_data["open_high_time"],
        "Open Low After High": intraday_data["open_low_after_high"],
        "Open Low After High Time": intraday_data["open_low_after_high_time"],
        "Open Low 15m After High": intraday_data["open_low_15m_after_high"],
        "Open Low 15m After High Time": intraday_data["open_low_15m_after_high_time"],
        "Open Low 30m After High": intraday_data["open_low_30m_after_high"],
        "Open Low 30m After High Time": intraday_data["open_low_30m_after_high_time"],
        "Open Low 1h After High": intraday_data["open_low_1h_after_high"],
        "Open Low 1h After High Time": intraday_data["open_low_1h_after_high_time"],
        "Open Low": intraday_data["open_low"],
        "Open Low Time": intraday_data["open_low_time"],
        "Reclaim VWAP Open": intraday_data["reclaim_vwap_open"],
        "VWAP Breakdown Time Open": intraday_data["vwap_breakdown_time_open"],
        "VWAP Deviation % Open High": intraday_data["vwap_deviation_open_high"],
        "Open Close": intraday_data["open_close"],
        "Vol @ Open High": intraday_data["vol_open_high"],
        "Vol @ Open Low": intraday_data["vol_open_low"],
        "AH Open": intraday_data["ah_open"],
        "AH High": intraday_data["ah_high"],
        "AH High Time": intraday_data["ah_high_time"],
        "AH Low": intraday_data["ah_low"],
        "AH Low Time": intraday_data["ah_low_time"],
        "AH Close": intraday_data["ah_close"],
        "Reclaim VWAP AH": intraday_data["reclaim_vwap_ah"],
        "VWAP Breakdown Time AH": intraday_data["vwap_breakdown_time_ah"],
        "VWAP Deviation % AH High": intraday_data["vwap_deviation_ah_high"],
        "Vol @ AH High": intraday_data["vol_ah_high"],
        "Vol @ 6AM": intraday_data["Vol @ 6AM"],
        "Vol @ 7AM": intraday_data["Vol @ 7AM"],
        "Vol @ 8AM": intraday_data["Vol @ 8AM"],
        "Vol @ 9AM": intraday_data["Vol @ 9AM"],
        "Vol @ 10AM": intraday_data["Vol @ 10AM"],
        "Vol @ 11AM": intraday_data["Vol @ 11AM"],
        "Vol @ 12PM": intraday_data["Vol @ 12PM"],
        "% Gain PM": intraday_data["percent_gain_pm_high"],
        "% Gain Open": intraday_data["percent_gain_open_high"],
        "% Gain AH": intraday_data["percent_gain_ah_high"],
        "Market Cap at PM High": intraday_data["market_cap_pm_high"],
        "Market Cap at Open High": intraday_data["market_cap_open_high"],
        "Market Cap 10AM": intraday_data["market_cap_10am"],
        "Market Cap 11AM": intraday_data["market_cap_11am"],
        "Market Cap 12PM": intraday_data["market_cap_12pm"],
        "Market Cap AH": intraday_data["market_cap_ah_high"],
        # NEW: Short strategy columns
        "PM Short P&L": intraday_data["pm_short_pnl"],
        "PM Short P&L %": intraday_data["pm_short_pnl_pct"],
        "Open Short P&L": intraday_data["open_short_pnl"],
        "Open Short P&L %": intraday_data["open_short_pnl_pct"],
        "PM to Open Gap": intraday_data["pm_to_open_gap"],
        "Open to 12pm Range": intraday_data["open_to_12pm_range"],
        "PM Volume Spike": intraday_data["pm_volume_spike"],
        "Open Volume Spike": intraday_data["open_volume_spike"],
        "PM Momentum": intraday_data["pm_momentum"],
        "Open Momentum": intraday_data["open_momentum"],
        "PM Reversal Signal": intraday_data["pm_reversal_signal"],
        "Open Reversal Signal": intraday_data["open_reversal_signal"],
        # RSI columns for PM and Open sessions
        "PM RSI at High 1min": intraday_data["PM RSI at High 1min"],
        "PM RSI at High 2min": intraday_data["PM RSI at High 2min"],
        "PM RSI at High 5min": intraday_data["PM RSI at High 5min"],
        "PM RSI at High 1hour": intraday_data["PM RSI at High 1hour"],
        "Open RSI at High 1min": intraday_data["Open RSI at High 1min"],
        "Open RSI at High 2min": intraday_data["Open RSI at High 2min"],
        "Open RSI at High 5min": intraday_data["Open RSI at High 5min"],
        "Open RSI at High 1hour": intraday_data["Open RSI at High 1hour"],
        # MACD columns for PM and Open sessions
        "MACD Line PM High": intraday_data["MACD_Line_PM_High"],
        "MACD Signal PM High": intraday_data["MACD_Signal_PM_High"],
        "MACD Histogram PM High": intraday_data["MACD_Histogram_PM_High"],
        "MACD Line Open High": intraday_data["MACD_Line_Open_High"],
        "MACD Signal Open High": intraday_data["MACD_Signal_Open_High"],
        "MACD Histogram Open High": intraday_data["MACD_Histogram_Open_High"],
        # Short interest and short float columns
        "Short Float %": intraday_data["Short Float %"],
        "Short Interest % of Outstanding": intraday_data["Short Interest % of Outstanding"],
        "Days to Cover": intraday_data["Days to Cover"],
        "Short Interest": intraday_data["Short Interest"],
        # Technical indicators
        "BB %B at Open High": intraday_data["BB %B at Open High"],
        "BB %B at PM High": intraday_data["BB %B at PM High"],
        "Short Volume": intraday_data["Short Volume"],
        "Short Volume Ratio": intraday_data["Short Volume Ratio"],
        # Time-based prices
        "Price 10AM": intraday_data["price_10am"],
        "Price 10:30AM": intraday_data["price_1030am"],
        "Price 11AM": intraday_data["price_11am"],
        "Price 11:30AM": intraday_data["price_1130am"],
        "Price 12PM": intraday_data["price_12pm"],
        # Dilution tracking columns
        "Active Warrant Shares": static_data.get("active_warrant_shares", 0),
        "Active Convertible Shares": static_data.get("active_convertible_shares", 0),
        "Total Active Dilution": static_data.get("total_active_dilution", 0),
        "Warrant Count": static_data.get("warrant_count", 0),
        "Convertible Count": static_data.get("convertible_count", 0),
        "Dilution vs Outstanding %": static_data.get("dilution_vs_outstanding_pct", 0.0),
        "Dilution vs Float %": static_data.get("dilution_vs_float_pct", 0.0),
        "Float Dilution Risk": static_data.get("float_dilution_risk", "Low"),
        # Float Rotation columns
        "Float Rotation 5AM": intraday_data.get("FloatRotation5AM", "Not Available"),
        "Float Rotation 6AM": intraday_data.get("FloatRotation6AM", "Not Available"),
        "Float Rotation 7AM": intraday_data.get("FloatRotation7AM", "Not Available"),
        "Float Rotation 8AM": intraday_data.get("FloatRotation8AM", "Not Available"),
        "Float Rotation 9AM": intraday_data.get("FloatRotation9AM", "Not Available"),
        "Float Rotation 9:30AM": intraday_data.get("FloatRotation9:30AM", "Not Available"),
        "Float Rotation 10AM": intraday_data.get("FloatRotation10AM", "Not Available"),
        "Float Rotation 11AM": intraday_data.get("FloatRotation11AM", "Not Available"),
        "Float Rotation 12PM": intraday_data.get("FloatRotation12PM", "Not Available"),
        "Float Rotation 1PM": intraday_data.get("FloatRotation1PM", "Not Available"),
        "Float Rotation 4PM": intraday_data.get("FloatRotation4PM", "Not Available"),
        # RVOL columns
        "RVOL PM High": intraday_data.get("RVOLPMHigh", "Not Available"),
        "RVOL Open High": intraday_data.get("RVOLOpenHigh", "Not Available"),
        "RVOL 7AM": intraday_data.get("RVOL7AM", "Not Available"),
        "RVOL 8AM": intraday_data.get("RVOL8AM", "Not Available"),
        "RVOL 9AM": intraday_data.get("RVOL9AM", "Not Available"),
        "RVOL 9:30AM": intraday_data.get("RVOL9:30AM", "Not Available"),
        "RVOL 10AM": intraday_data.get("RVOL10AM", "Not Available"),
        "RVOL 11AM": intraday_data.get("RVOL11AM", "Not Available"),
        "RVOL 12PM": intraday_data.get("RVOL12PM", "Not Available"),
        "RVOL 1PM": intraday_data.get("RVOL1PM", "Not Available"),
        "Used YFinance": static_data.get("used_yfinance", False)
    }
    
    # Add minute-by-minute data columns
    if intraday_data.get("raw_minute_data"):
        for time_key, price in intraday_data["raw_minute_data"].items():
            row[time_key] = price
    
    # Add per-minute VWAP % Deviation and VWAP Deviation Price columns
    for k, v in intraday_data.items():
        if isinstance(k, str) and (k.startswith("VWAP % Deviation ") or k.startswith("VWAP Deviation Price ")):
            row[k] = v
    
    
    return row

def process_pair(ticker, date, skip_trading_check=False, float_shares=None, outstanding_shares=None, news_data="No news", news_source="Unknown", excel_row_index=None):
    """Process a single ticker-date pair using the phase-based approach.
    If float_shares or outstanding_shares are provided (from Excel), they will be used with priority over API values."""
    if not skip_trading_check and not is_trading_day(date):
        logging.info(f"{date} is not a trading day, skipping {ticker}")
        return None
    
    try:
        # Phase 1: Fast APIs - pass Excel values if available
        phase1_result = process_phase1_fast_apis(ticker, date, float_shares, outstanding_shares)
        if not phase1_result:
            return None
        
        # Phase 2: Skipped - using API float data only
        phase2_result = None
        
        # Phase 3: Final assembly
        result = process_phase3_final_assembly(ticker, date, phase1_result, phase2_result, news_data, news_source, excel_row_index)
        return result
        
    except Exception as e:
        logging.error(f"Error processing {ticker} for {date}: {e}")
        return None

def process_single_date(date, all_tickers):
    """Process a single date with all tickers and return results."""
    logging.info(f"Starting processing for date: {date}")
    
    if not is_trading_day(date):
        logging.info(f"{date} is not a trading day, skipping")
        return []
    
    # Filter tickers for this specific date
    filtered_tickers = optimized_filter_pre_market_gainers(all_tickers, date)
    ticker_date_pairs = [(ticker["ticker"], date, ticker.get("float_shares"), None) for ticker in filtered_tickers]
    
    if not ticker_date_pairs:
        logging.info(f"No qualifying tickers found for {date}")
        return []
    
    # Fetch news data for all ticker-date pairs
    # news_dict = get_all_news(ticker_date_pairs)
    news_dict = {}  # News fetching disabled per request
    
    # Process tickers for this date
    results = []
    with ThreadPoolExecutor(max_workers=25) as executor:  # Increased to 25 workers for better performance
        # Submit all tasks and store futures in order
        futures = []
        for ticker, date, float_shares, outstanding_shares in ticker_date_pairs:
            news_tuple = news_dict.get((ticker, date), ("No news", "Unknown"))
            future = executor.submit(process_pair, ticker, date, True, float_shares, outstanding_shares, news_tuple[0], news_tuple[1])
            futures.append((future, ticker, date))
        
        # Process results in the same order as input (not completion order)
        for i, (future, ticker, date) in enumerate(futures, 1):
            try:
                result = future.result()
                if result:
                    results.append(result)
            except Exception as e:
                logging.error(f"Error processing {ticker} for {date}: {e}")
            if i % 50 == 0:  # Log progress every 50 items
                logging.info(f"Date {date}: Processed {i} of {len(ticker_date_pairs)} pairs")
    
    logging.info(f"Completed processing for {date}: {len(results)} results")
    return results
def main(dates=None):
    """Main function to process NASDAQ tickers with 10%+ pre-market gain, 50K+ volume, and float < 20M."""
    # Start timing the entire script
    script_start_time = time.time()
    
    # Get dates or Excel filename from command line arguments
    excel_filename = 'Tickers.xlsx'  # Default filename
    if dates is None:
        import sys
        if len(sys.argv) > 1:
            first_arg = sys.argv[1]
            # Check if first argument is a date (YYYY-MM-DD format) or Excel filename
            try:
                datetime.strptime(first_arg, "%Y-%m-%d")
                # It's a date, use date mode
                dates = sys.argv[1:]
                logging.info(f"Using dates from command line: {dates}")
            except ValueError:
                # Not a date, treat as Excel filename
                excel_filename = first_arg if first_arg.endswith('.xlsx') else f"{first_arg}.xlsx"
                dates = []
                logging.info(f"Using Excel file: {excel_filename}")
        else:
            dates = []
            logging.info("No command line arguments provided, using default Tickers.xlsx")

    excel_mode = not bool(dates)
    valid_dates = []
    ticker_date_pairs = []

    if dates:
        # Validate provided dates
        for date in dates:
            try:
                datetime.strptime(date, "%Y-%m-%d")
                valid_dates.append(date)
            except ValueError:
                logging.error(f"Invalid date format: {date}. Expected YYYY-MM-DD")
                continue

        if not valid_dates:
            logging.error("No valid dates provided, exiting")
            return

        # Fetch all NASDAQ tickers once (shared across all dates)
        all_tickers = get_nasdaq_tickers(test_mode=False)  # Set to False for full NASDAQ
        if not all_tickers:
            logging.error("No NASDAQ tickers retrieved, exiting")
            return

        logging.info(f"Processing {len(valid_dates)} dates with {len(all_tickers)} tickers")
        
        # Build ticker_date_pairs for all dates
        for date in valid_dates:
            filtered_tickers = optimized_filter_pre_market_gainers(all_tickers, date)
            ticker_date_pairs.extend([(ticker["ticker"], date, ticker.get("float_shares"), None) for ticker in filtered_tickers])
        
        logging.info(f"Built {len(ticker_date_pairs)} ticker-date pairs across {len(valid_dates)} dates")
        
    else:
        # Excel mode with optimized batch processing
        try:
            df = pd.read_excel(excel_filename)
            
            # Optimize Excel processing with vectorized operations where possible
            logging.info(f"Processing Excel with {len(df)} rows")
            
            for row_index, row in df.iterrows():
                ticker = str(row['Ticker']).strip().upper()  # Normalize ticker format
                date_str = str(row['Date']).strip()
                
                # Add row index to maintain Excel order
                excel_row_index = row_index
                
                # Net Return column removed - no longer needed
                
                # Get Float value if available
                float_shares = None
                if 'Float' in df.columns and pd.notna(row['Float']):
                    try:
                        float_value = str(row['Float']).strip()
                        # Handle values with 'M' suffix (millions)
                        if float_value.endswith('M'):
                            float_shares = float(float_value[:-1]) * 1_000_000
                        else:
                            float_shares = float(float_value)
                        logging.info(f"Using float value from Excel for {ticker}: {float_shares:,.0f}")
                    except Exception as e:
                        logging.warning(f"Error parsing float value for {ticker}: {e}")
                        float_shares = None
                
                # Get Outstanding Shares value if available
                outstanding_shares = None
                if 'Outstanding Shares' in df.columns and pd.notna(row['Outstanding Shares']):
                    try:
                        shares_value = str(row['Outstanding Shares']).strip()
                        if shares_value != 'Not Available' and shares_value != 'N/A':
                            # Handle values with 'M' suffix (millions)
                            if shares_value.endswith('M'):
                                outstanding_shares = float(shares_value[:-1]) * 1_000_000
                            elif shares_value.endswith('B'):
                                outstanding_shares = float(shares_value[:-1]) * 1_000_000_000
                            elif shares_value.endswith('K'):
                                outstanding_shares = float(shares_value[:-1]) * 1_000
                            else:
                                outstanding_shares = float(shares_value)
                            logging.info(f"Using outstanding shares from Excel for {ticker}: {outstanding_shares:,.0f}")
                    except Exception as e:
                        logging.warning(f"Error parsing outstanding shares for {ticker}: {e}")
                        outstanding_shares = None
                
                # Skip empty or NaN values
                if pd.isna(date_str) or date_str == '' or date_str.lower() == 'nan':
                    logging.warning(f"Skipping empty date for {ticker}")
                    continue
                    
                try:
                    # Try multiple date formats and handle datetime objects
                    dt = None
                    
                    # First try to handle datetime objects directly
                    if hasattr(date_str, 'date'):
                        dt = date_str
                    else:
                        # Handle string dates with time components
                        date_str_clean = str(date_str).split(' ')[0]  # Remove time part
                        date_formats = [
                            '%b %d, %Y',    # "Dec 05, 2023"
                            '%d/%m/%Y',     # "29/01/2025"
                            '%Y-%m-%d',     # "2025-01-29"
                            '%m/%d/%Y',     # "01/29/2025"
                            '%d-%m-%Y',     # "29-01-2025"
                            '%Y/%m/%d'      # "2025/01/29"
                        ]
                        
                        for date_format in date_formats:
                            try:
                                dt = datetime.strptime(date_str_clean, date_format)
                                break
                            except ValueError:
                                continue
                    
                    if dt is None:
                        logging.error(f"Invalid date format in Excel: {date_str} for {ticker}")
                        continue
                    
                    internal_date = dt.strftime('%Y-%m-%d')
                    ticker_date_pairs.append((ticker, internal_date, float_shares, outstanding_shares, excel_row_index))
                    valid_dates.append(internal_date)
                except Exception as e:
                    logging.error(f"Error processing date {date_str} for {ticker}: {e}")
            
            valid_dates = sorted(set(valid_dates))
            logging.info(f"Parsed {len(ticker_date_pairs)} valid ticker-date pairs across {len(valid_dates)} unique dates")
        except FileNotFoundError:
            logging.error(f"{excel_filename} not found, exiting")
            return
        except Exception as e:
            logging.error(f"Error reading {excel_filename}: {e}")
            return

    if not ticker_date_pairs:
        logging.error("No valid ticker-date pairs to process, exiting")
        return

    # Fetch news data for all ticker-date pairs
    # news_dict = get_all_news(ticker_date_pairs)
    news_dict = {}  # News fetching disabled per request

    # Start timing for performance monitoring
    start_time = time.time()

    # Calculate optimal number of workers based on system and workload
    import os
    cpu_count = os.cpu_count() or 4
    total_pairs = len(ticker_date_pairs)
    
    # Single-phase processing - combine everything into one phase
    workers = min(25, os.cpu_count() or 4)  # Set to 25 workers for better performance
    
    logging.info(f"Processing {total_pairs} ticker-date pairs with single-phase approach")
    logging.info(f"Workers: {workers} (single phase processing)")
    
    # Single phase: Process everything at once
    logging.info(f"ENTERING SINGLE PHASE: Processing all data with {workers} workers...")
    all_results = []
    
    with ThreadPoolExecutor(max_workers=workers) as executor:
        # Submit all tasks and store futures in order
        futures = []
        for pair in ticker_date_pairs:
            # Handle both old format (4 elements) and new format (5 elements with excel_row_index)
            if len(pair) == 5:
                ticker, date, float_shares, outstanding_shares, excel_row_index = pair
            else:
                ticker, date, float_shares, outstanding_shares = pair
                excel_row_index = None
            news_tuple = news_dict.get((ticker, date), ("No news", "Unknown"))
            future = executor.submit(process_pair, ticker, date, True, float_shares, outstanding_shares, news_tuple[0], news_tuple[1], excel_row_index)
            futures.append((future, ticker, date, excel_row_index))
        
        completed_count = 0
        progress_interval = max(1, min(50, total_pairs // 20))
        failed_tickers = []
        
        # Process results in the same order as input (not completion order)
        for future, ticker, date, excel_row_index in futures:
            completed_count += 1
            
            try:
                result = future.result()
                if result:
                    # Ensure Excel_Row_Index is set if not already present
                    if "Excel_Row_Index" not in result and excel_row_index is not None:
                        result["Excel_Row_Index"] = excel_row_index
                    all_results.append(result)
                else:
                    # Create failed row with Excel row index
                    failed_row = create_failed_row(ticker, date, excel_row_index, "Processing returned None")
                    failed_tickers.append({
                        "Ticker": ticker,
                        "Date": date,
                        "Excel_Row_Index": excel_row_index,
                        "Status": "Failed - No data returned",
                        "Error": "Processing returned None"
                    })
                    all_results.append(failed_row)
                    logging.warning(f"Failed to process {ticker} on {date} - no data returned")
            except Exception as e:
                # Create failed row with Excel row index
                failed_row = create_failed_row(ticker, date, excel_row_index, str(e))
                failed_tickers.append({
                    "Ticker": ticker,
                    "Date": date,
                    "Excel_Row_Index": excel_row_index,
                    "Status": "Failed - Exception",
                    "Error": str(e)
                })
                all_results.append(failed_row)
                logging.error(f"Error processing {ticker} on {date}: {e}")
            
            if completed_count % progress_interval == 0:
                progress_pct = (completed_count / total_pairs) * 100
                logging.info(f"Progress: {completed_count}/{total_pairs} pairs ({progress_pct:.1f}%)")
                
                if completed_count % (progress_interval * 2) == 0:  # More frequent cleanup for 20 workers
                    monitor_memory()
                    cleanup_memory()
                    cleanup_polygon_client()
                    
                    elapsed_time = time.time() - start_time
                    if elapsed_time > 0:
                        rate = completed_count / elapsed_time
                        eta = (total_pairs - completed_count) / rate if rate > 0 else 0
                        logging.info(f"Processing rate: {rate:.1f} pairs/sec, ETA: {eta/60:.1f} minutes")
    
    # Final progress update with performance summary
    total_time = time.time() - start_time
    avg_time_per_pair = total_time / len(all_results) if all_results else 0
    pairs_per_second = len(all_results) / total_time if total_time > 0 else 0
    
    logging.info(f"[SUCCESS] Completed processing {len(all_results)} of {total_pairs} pairs")
    logging.info(f"[TIMING] Total time: {total_time/60:.1f} minutes, Avg: {avg_time_per_pair:.2f}s per pair, Rate: {pairs_per_second:.1f} pairs/sec")
    print(f"[SUCCESS] Processed {len(all_results)} of {total_pairs} pairs in {total_time/60:.1f} minutes ({pairs_per_second:.1f} pairs/sec)")

    if all_results:
        try:
            output_df = pd.DataFrame(all_results)
            
            # Fill any missing columns (from minute-by-minute data or other dynamic columns) with "Not Available"
            # This ensures failed rows have all columns filled
            for col in output_df.columns:
                if col != 'Excel_Row_Index' and col != 'Error':
                    output_df[col] = output_df[col].fillna("Not Available")
            
            # Sort by Excel row index to maintain original Excel order
            if 'Excel_Row_Index' in output_df.columns:
                output_df = output_df.sort_values('Excel_Row_Index').reset_index(drop=True)
                # Remove the Excel_Row_Index column after sorting
                output_df = output_df.drop('Excel_Row_Index', axis=1)
                logging.info("Results sorted by Excel row order")
            
            # Format Date to d/m/y
            output_df['Date'] = pd.to_datetime(output_df['Date']).dt.strftime('%d/%m/%Y')
            
            # Reorder columns to put minute data at the end
            primary_columns = ['Ticker', 'Date']
            
            # Get minute columns (04:00, 04:01, etc.)
            minute_columns = sorted([col for col in output_df.columns if col.count(':') == 1 and col.count(' ') == 0])
            
            # Get all other columns (excluding minute columns)
            other_columns = [col for col in output_df.columns if col not in minute_columns and col not in primary_columns]
            
            # Reorder DataFrame
            final_columns = primary_columns + other_columns + minute_columns
            output_df = output_df.reindex(columns=final_columns)
            
            # Create filename with date range and timestamp to prevent conflicts
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if len(valid_dates) == 1:
                dt = datetime.strptime(valid_dates[0], '%Y-%m-%d')
                date_str = dt.strftime('%d%m%y')
                output_file = f"stock_data_pm_{date_str}_{timestamp}.csv"
            else:
                earliest = min(valid_dates)
                latest = max(valid_dates)
                dt_early = datetime.strptime(earliest, '%Y-%m-%d')
                dt_late = datetime.strptime(latest, '%Y-%m-%d')
                early_str = dt_early.strftime('%d%m%y')
                late_str = dt_late.strftime('%d%m%y')
                output_file = f"stock_data_pm_{early_str}to{late_str}_{timestamp}.csv"
            
            # Save as Excel with color coding
            excel_file = save_to_excel_with_colors(output_df, output_file)
            if excel_file:
                upload_to_s3(excel_file)

            # Option 2: Save separate files for each date (uncomment if preferred)
            # for date in valid_dates:
            #     date_df = output_df[output_df["Date"] == date]
            #     if not date_df.empty:
            #         output_file = f"stock_data_{date}.csv"
            #         date_df.to_csv(output_file, mode='w', header=True, index=False)
            #         logging.info(f"Wrote results for {date} to {output_file}")
            #         upload_to_s3(output_file)
        except Exception as e:
            logging.error(f"Error saving results: {e}")
    
    # Calculate and display total script execution time
    script_total_time = time.time() - script_start_time
    script_minutes = script_total_time / 60
    logging.info(f"[SCRIPT COMPLETION] Total execution time: {script_minutes:.1f} minutes")
    print(f"[SCRIPT COMPLETION] Total execution time: {script_minutes:.1f} minutes")

def lambda_handler(event, context):
    """AWS Lambda handler."""
    # Require event to provide dates explicitly
    dates = event.get('dates')
    if not dates:
        raise ValueError("No dates provided in event for Lambda handler.")
    main(dates)
    return {"status": "success"}

if __name__ == "__main__":
    main()